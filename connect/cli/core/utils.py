# -*- coding: utf-8 -*-

# This file is part of the Ingram Micro Cloud Blue Connect connect-cli.
# Copyright (c) 2019-2022 Ingram Micro. All Rights Reserved.
import os

import click
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import Color, WHITE

from connect.cli import get_version
from connect.cli.core.constants import PYPI_JSON_API_URL


def continue_or_quit():
    while True:
        click.echo('')
        click.echo("Press 'c' to continue or 'q' to quit ", nl=False)
        c = click.getchar()
        click.echo()
        if c == 'c':
            return True
        if c == 'q':
            return False


def check_for_updates(*args):
    try:
        res = requests.get(PYPI_JSON_API_URL)
        if res.status_code == 200:
            data = res.json()
            version = data['info']['version']
            current = get_version()
            if version != current:
                click.secho(
                    f'\nYou are running CloudBlue Connect CLI version {current}. '
                    f'A newer version is available: {version}.\n',
                    fg='yellow',
                )
    except requests.RequestException:
        pass


def validate_output_options(output_path, output_file, default_dir_name, default_file_name=None):
    """
    Common validation for commands using output path and file options.
    """
    if not default_file_name:
        default_file_name = default_dir_name

    output_path = output_path or os.getcwd()
    if not os.path.exists(output_path):
        raise click.ClickException("Output Path does not exist")

    output_path = os.path.join(output_path, default_dir_name)
    if not os.path.exists(output_path):
        os.mkdir(output_path)
    elif not os.path.isdir(output_path):
        raise click.ClickException(
            f"Exists a file with name '{os.path.basename(output_path)}' but a directory is "
            "expected, please rename it",
        )

    output_file = os.path.join(output_path, output_file or f'{default_file_name}.xlsx')

    return output_file


def set_ws_main_header(ws, title):
    """
    Set the header for a main worksheet
    """
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 180
    ws.merge_cells('A1:B1')
    cell = ws['A1']
    cell.fill = PatternFill('solid', start_color=Color('1565C0'))
    cell.font = Font(sz=24, color=WHITE)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value = title

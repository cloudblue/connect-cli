from datetime import datetime

from click import ClickException
from fs.tempfs import TempFS
from openpyxl import load_workbook

from connect.cli.plugins.product.export import dump_product
from connect.cli.plugins.product.sync import (
    ActionsSynchronizer,
    CapabilitiesSynchronizer,
    GeneralSynchronizer,
    ItemSynchronizer,
    MediaSynchronizer,
    ParamsSynchronizer,
    TemplatesSynchronizer,
)
from connect.client import ClientError
from connect.cli.plugins.shared.translations_synchronizers import sync_product_translations
from connect.cli.plugins.shared.utils import get_translation_attributes_sheets


class ProductCloner:
    def __init__(self, config, source_account, destination_account, product_id, progress, stats):
        self.fs = TempFS(identifier=f'_clone_{product_id}')
        self.config = config
        self.source_account = (source_account if source_account else config.active.id)
        self.destination_account = (destination_account if destination_account else config.active.id)
        self.product_id = product_id
        self.stats = stats
        self.progress = progress
        self.destination_product = None
        self.wb = None

    def dump(self):
        dump_product(
            client=self.config.active.client,
            product_id=self.product_id,
            output_path=self.fs.root_path,
            progress=self.progress,
            output_file='',
        )

    def inject(self):  # noqa: CCR001
        try:
            self.config.activate(self.destination_account)
            input_file = f'{self.fs.root_path}/{self.product_id}/{self.product_id}.xlsx'
            synchronizer = GeneralSynchronizer(
                self.config.active.client,
                None,
            )

            synchronizer.open(input_file, 'General Information')
            synchronizer.sync()
            synchronizer = ItemSynchronizer(
                self.config.active.client,
                self.progress,
                self.stats,
            )
            product_id = synchronizer.open(input_file, 'Items')
            items = self.config.active.client.products[product_id].items.all()
            for item in items:
                self.config.active.client.products[product_id].items[item['id']].delete()
            synchronizer.sync()

            synchronizer = CapabilitiesSynchronizer(
                self.config.active.client,
                self.progress,
                self.stats,
            )
            synchronizer.open(input_file, 'Capabilities')
            synchronizer.sync()

            templates = self.config.active.client.products[product_id].templates.all()
            sample_templates = []
            for template in templates:
                sample_templates.append(template['id'])

            synchronizer = TemplatesSynchronizer(
                self.config.active.client,
                self.progress,
                self.stats,
            )

            synchronizer.open(input_file, 'Templates')
            synchronizer.sync()

            for template in sample_templates:
                try:
                    self.config.active.client.products[product_id].templates[template].delete()
                except ClientError:
                    # done intentionally till fulfillment in progress template not on public api
                    pass

            synchronizer = ParamsSynchronizer(
                self.config.active.client,
                self.progress,
                self.stats,
            )

            synchronizer.open(input_file, "Ordering Parameters")
            synchronizer.sync()

            synchronizer.open(input_file, "Fulfillment Parameters")
            synchronizer.sync()

            synchronizer.open(input_file, "Configuration Parameters")
            synchronizer.sync()

            synchronizer = ActionsSynchronizer(
                self.config.active.client,
                self.progress,
                self.stats,
            )

            synchronizer.open(input_file, 'Actions')
            synchronizer.sync()

            synchronizer = MediaSynchronizer(
                self.config.active.client,
                self.progress,
                self.stats,
            )

            synchronizer.open(input_file, 'Media')
            synchronizer.sync()

            sync_product_translations(
                self.config.active.client,
                self.progress,
                input_file,
                self.stats,
                save=False,
                is_clone=True,
            )

            self.config.activate(self.source_account)
        except ClientError as e:
            raise ClickException(f"Error while cloning product: {str(e)}")

    def load_wb(self):
        self.wb = load_workbook(
            f'{self.fs.root_path}/{self.product_id}/{self.product_id}.xlsx',
            data_only=True,
        )

    def create_product(self, name=None):
        if not name:
            time = datetime.today().strftime('%Y-%m-%d-%H:%M:%S')
            name = f"Clone of {self.product_id} {time}"
        ws = self.wb['General Information']
        ws['B6'].value = name
        self.config.activate(self.destination_account)

        try:
            category = self._get_cat_id(self.config.active.client, ws['B8'].value)
            primary_locale_id = self._get_primary_locale_id(ws['B14'].value)
            product = self.config.active.client.products.create(
                {
                    "name": name,
                    "category": {
                        "id": category,
                    },
                    "translations": [
                        {"locale": {"id": primary_locale_id}, "primary": True},
                    ],
                },
            )
            if name:
                ws['B6'].value = name
            ws['B5'].value = product['id']
            self.destination_product = product['id']
            self.wb.save(f'{self.fs.root_path}/{self.product_id}/{self.product_id}.xlsx')
        except ClientError as e:
            raise ClickException(f'Error on product creation: {str(e)}')

    def clean_wb(self):
        ws = self.wb['Capabilities']
        for row in range(2, 11):
            ws[f'B{row}'].value = 'update'

        ws = self.wb['Embedding Static Resources']
        for row in range(2, ws.max_row + 1):
            ws[f'C{row}'].value = 'create'

        ws = self.wb['Media']
        for row in range(2, ws.max_row + 1):
            ws[f'B{row}'].value = ''
            ws[f'C{row}'].value = 'create'

        ws = self.wb['Templates']
        for row in range(2, ws.max_row + 1):
            ws[f'A{row}'].value = ''
            ws[f'C{row}'].value = 'create'

        ws = self.wb['Items']
        for row in range(2, ws.max_row + 1):
            ws[f'A{row}'].value = ''
            ws[f'C{row}'].value = 'create'

        ws = self.wb['Ordering Parameters']
        for row in range(2, ws.max_row + 1):
            ws[f'A{row}'].value = ''
            ws[f'C{row}'].value = 'create'

        ws = self.wb['Fulfillment Parameters']
        for row in range(2, ws.max_row + 1):
            ws[f'A{row}'].value = ''
            ws[f'C{row}'].value = 'create'

        ws = self.wb['Configuration Parameters']
        for row in range(2, ws.max_row + 1):
            ws[f'A{row}'].value = ''
            ws[f'C{row}'].value = 'create'

        ws = self.wb['Actions']
        for row in range(2, ws.max_row + 1):
            ws[f'A{row}'].value = ''
            ws[f'C{row}'].value = 'create'

        ws = self.wb['Translations']
        for row in range(2, ws.max_row + 1):
            ws[f'A{row}'].value = ''
            ws[f'B{row}'].value = 'create'
        for sheetname in get_translation_attributes_sheets(self.wb):
            ws = self.wb[sheetname]
            value = 'update'
            if sheetname.split()[0] == self.wb['General Information']['B14'].value.split()[0]:
                value = '-'
            for row in range(2, ws.max_row + 1):
                ws[f'C{row}'].value = value
        self.wb.save(f'{self.fs.root_path}/{self.product_id}/{self.product_id}.xlsx')

    @staticmethod
    def _get_cat_id(client, category_name):
        categories = client.categories.all()
        for category in categories:
            if category['name'] == category_name:
                return category['id']

    @staticmethod
    def _get_primary_locale_id(locale_repr):
        """
        `locale_repr` should be in the form of "<id> (<verbose name>)",
        e.g. "EN-GB (British English)"
        """
        return locale_repr.split(' ')[0]

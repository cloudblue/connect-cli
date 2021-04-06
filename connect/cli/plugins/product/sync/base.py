# -*- coding: utf-8 -*-

# This file is part of the Ingram Micro Cloud Blue Connect connect-cli.
# Copyright (c) 2019-2021 Ingram Micro. All Rights Reserved.

from zipfile import BadZipFile

from click import ClickException
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from connect.cli.plugins.exceptions import SheetNotFoundError
from connect.cli.plugins.product.utils import (
    get_col_headers_by_ws_type,
    get_col_limit_by_ws_type,
    get_ws_type_by_worksheet_name,
)


class ProductSynchronizer:
    def __init__(self, client, silent):
        self._client = client
        self._silent = silent
        self._product_id = None
        self._wb = None

    def open(self, input_file, worksheet):
        self._open_workbook(input_file)
        if worksheet not in self._wb.sheetnames:
            raise SheetNotFoundError(f'File does not contain {worksheet} to synchronize, skipping')
        ws = self._wb['General Information']
        product_id = ws['B5'].value
        if not self._client.products[product_id].exists():
            raise ClickException(f'Product {product_id} not found, create it first.')
        ws = self._wb[worksheet]
        self._validate_worksheet_sheet(ws, worksheet)

        self._product_id = product_id
        return self._product_id

    def sync(self):
        raise NotImplementedError("Not implemented")

    def save(self, output_file):
        self._wb.save(output_file)

    def _open_workbook(self, input_file):
        try:
            self._wb = load_workbook(
                input_file,
                data_only=True,
            )
        except InvalidFileException as ife:
            raise ClickException(str(ife))
        except BadZipFile:
            raise ClickException(f'{input_file} is not a valid xlsx file.')

    @staticmethod
    def _validate_worksheet_sheet(ws, worksheet):
        ws_type = get_ws_type_by_worksheet_name(worksheet)
        max_letter = get_col_limit_by_ws_type(ws_type)
        col_headers = get_col_headers_by_ws_type(ws_type)
        cels = ws['A1': f'{max_letter}1']
        for cel in cels[0]:
            if cel.value != col_headers[cel.column_letter]:
                raise ClickException(
                    f'Invalid input file: column {cel.column_letter} '
                    f'must be {col_headers[cel.column_letter]}',
                )

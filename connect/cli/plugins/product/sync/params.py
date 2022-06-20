# -*- coding: utf-8 -*-

# This file is part of the Ingram Micro Cloud Blue Connect connect-cli.
# Copyright (c) 2019-2021 Ingram Micro. All Rights Reserved.
import json
import re
from collections import namedtuple
from json.decoder import JSONDecodeError

from openpyxl.styles import Alignment

from connect.cli.plugins.product.constants import PARAM_TYPES
from connect.cli.plugins.shared.base import ProductSynchronizer
from connect.cli.plugins.shared.constants import PARAMS_COLS_HEADERS
from connect.cli.plugins.product.utils import get_json_object_for_param, ParamSwitchNotSupported
from connect.client import ClientError


fields = (v.replace(' ', '_').lower() for v in PARAMS_COLS_HEADERS.values())

_RowData = namedtuple('RowData', fields)


class ParamsSynchronizer(ProductSynchronizer):
    def __init__(self, client, progress, stats):
        self._param_type = None
        self._worksheet_name = None
        self.__stats = stats
        self._mstats = None
        super(ParamsSynchronizer, self).__init__(client, progress)

    def open(self, input_file, worksheet):
        if worksheet == "Ordering Parameters":
            self._param_type = 'ordering'
        elif worksheet == "Fulfillment Parameters":
            self._param_type = 'fulfillment'
        elif worksheet == "Configuration Parameters":
            self._param_type = 'configuration'
        self._worksheet_name = worksheet
        self._mstats = self.__stats[self._worksheet_name]
        return super(ParamsSynchronizer, self).open(input_file, worksheet)

    def sync(self):  # noqa: CCR001
        ws = self._wb[self._worksheet_name]

        task = self._progress.add_task('Processing param', total=ws.max_row - 1)
        for row_idx in range(2, ws.max_row + 1):
            data = _RowData(*[ws.cell(row_idx, col_idx).value for col_idx in range(1, 15)])
            self._progress.update(
                task,
                description=f'Processing param {data.id}',
                advance=1,
            )
            if data.action == '-':
                self._mstats.skipped()
                continue
            row_errors = self._validate_row(data)

            if row_errors:
                self._mstats.error(row_errors, row_idx)
                continue

            if data.action == 'delete':
                try:
                    self._client.products[self._product_id].parameters[data.verbose_id].delete()
                except ClientError:
                    pass
                self._mstats.deleted()
                continue

            param_payload = {}
            if data.json_properties:
                param_payload = json.loads(data.json_properties)
            param_payload['name'] = data.id
            param_payload['title'] = data.title
            param_payload['description'] = data.description
            param_payload['phase'] = data.phase
            param_payload['scope'] = data.scope
            param_payload['type'] = data.type
            if 'constraints' not in param_payload:
                param_payload['constraints'] = {}
            param_payload['constraints']['required'] = False if data.required == '-' else True
            param_payload['constraints']['unique'] = False if data.unique == '-' else True
            param_payload['constraints']['hidden'] = False if data.hidden == '-' else True

            if data.action == 'update':
                try:
                    original_param = self._get_original_param(data)
                    if original_param:
                        self._compare_param(original_param, data)

                    param = self._client.products[self._product_id].parameters[
                        data.verbose_id
                    ].update(
                        param_payload,
                    )
                    self._update_sheet_row(ws, row_idx, param)
                    self._mstats.updated()
                except Exception as e:
                    self._mstats.error(str(e), row_idx)

            if data.action == 'create':
                try:
                    original_param = self._get_original_param(data)
                    if original_param:
                        self._updated_or_skipped(ws, row_idx, original_param, param_payload)
                        continue
                    param = self._client.products[self._product_id].parameters.create(
                        param_payload,
                    )
                    self._update_sheet_row(ws, row_idx, param)
                    self._mstats.created()
                except Exception as e:
                    self._mstats.error(str(e), row_idx)
        self._progress.update(task, completed=ws.max_row - 1)

    @staticmethod
    def _update_sheet_row(ws, row_idx, param=None):
        ws.cell(row_idx, 3, value='-')
        if param:
            ws.cell(row_idx, 1, value=param['id']).alignment = Alignment(
                horizontal='left',
                vertical='top',
            )
            ws.cell(row_idx, 12, value=get_json_object_for_param(param))
            ws.cell(row_idx, 13, value=param['events']['created']['at']).alignment = Alignment(
                horizontal='left',
                vertical='top',
            )
            ws.cell(
                row_idx,
                14,
                value=param['events'].get('updated', {}).get('at'),
            ).alignment = Alignment(
                horizontal='left',
                vertical='top',
            )

    @staticmethod
    def _compare_param(original, data):
        if original['type'] != data.type:
            raise ParamSwitchNotSupported('Switching parameter type is not supported')
        if original['scope'] != data.scope:
            raise ParamSwitchNotSupported('switching scope is not supported')
        if original['phase'] != data.phase:
            raise ParamSwitchNotSupported('switching phase is not supported')

    def _validate_row(self, data):  # noqa: CCR001

        errors = []
        if not data.id:
            errors.append(
                'Parameter must have an id',
            )
            return errors
        id_pattern = "^[A-Za-z0-9_-]*$"

        if not bool(re.match(id_pattern, data.id)):
            errors.append(
                f'Parameter ID must contain only letters, numbers and `_`, provided {data.id}',
            )

        elif data.phase != self._param_type:
            errors.append(
                f'Parameters of type {self._param_type} are only supported when processing '
                f'{self._worksheet_name}. Has been provided {data.phase}.',
            )
        elif data.action in ('update', 'delete') and (
                not data.verbose_id or not data.verbose_id.startswith('PRM-')
        ):
            errors.append(
                'Verbose ID is required on update and delete actions.',
            )
        elif data.type not in PARAM_TYPES:
            errors.append(
                f'Parameter type {data.type} is not one of the supported ones:'
                f'{",".join(PARAM_TYPES)}',
            )
        elif self._param_type in ('ordering', 'fulfillment') and data.scope not in (
            'asset', 'tier1', 'tier2',
        ):
            errors.append(
                f'Only asset, tier1 and tier2 scopes are supported for {self._worksheet_name}',
            )
        elif self._param_type == 'configuration' and data.scope not in (
            'item', 'item_marketplace', 'marketplace', 'product',
        ):
            errors.append(
                'Only item, item_marketplace, marketplace and product scopes are supported for '
                f'{self._worksheet_name}',
            )
        elif data.required not in (True, 'True', '-'):
            errors.append(
                'Required must be either True or `-`',
            )
        elif data.unique not in (True, 'True', '-'):
            errors.append(
                'Unique must be either True or `-`',
            )
        elif data.hidden not in (True, 'True', '-'):
            errors.append(
                'Hidden must be either True or `-`',
            )

        if len(errors) > 0:
            return errors

        if data.json_properties:
            try:
                json.loads(data.json_properties)
            except JSONDecodeError:
                errors.append(
                    'JSON properties must have json format',
                )
        return errors

    def _get_original_param(self, data):
        filter_kwargs = {}
        if data.verbose_id:
            filter_kwargs['id'] = data.verbose_id
        else:
            filter_kwargs['name'] = data.id
        return (
            self._client.products[self._product_id].parameters
            .filter(**filter_kwargs)
            .first()
        )

    def _updated_or_skipped(self, ws, row_idx, original, payload):
        original_filter = {k: v for k, v in original.items() if k in payload.keys()}

        if original_filter == payload:
            self._update_sheet_row(ws, row_idx)
            self._mstats.skipped()
        else:
            param = self._client.products[self._product_id].parameters[
                original["id"]
            ].update(
                payload,
            )
            self._update_sheet_row(ws, row_idx, param)
            self._mstats.updated()

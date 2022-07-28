# -*- coding: utf-8 -*-

# This file is part of the Ingram Micro Cloud Blue Connect connect-cli.
# Copyright (c) 2019-2022 Ingram Micro. All Rights Reserved.

import os
import copy
import json
from datetime import datetime
from urllib import parse

import requests
from click import ClickException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import Color, WHITE
from openpyxl.utils import quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

from connect.cli.core.http import (
    format_http_status,
    handle_http_error,
)
from connect.cli.core.utils import validate_output_options
from connect.cli.plugins.product.constants import PARAM_TYPES
from connect.cli.plugins.product.utils import get_json_object_for_param
from connect.cli.plugins.shared.export import (
    alter_attributes_sheet,
    get_translation_workbook,
)
from connect.cli.plugins.shared.utils import (
    fill_translation_row,
    get_col_headers_by_ws_type,
    get_col_limit_by_ws_type,
    setup_locale_data_validation,
)
from connect.client import ClientError, R


def _setup_locales_list(ws, client):
    """
    Fill list of locales to use with DataValidation.
    """
    locales = client.ns('localization').locales.all()
    locales_list = [f"{locale['id']} ({locale['name']})" for locale in locales]
    ws['AB1'].value = 'Locales'
    for idx, loc in enumerate(locales_list, 2):
        ws[f'AB{idx}'].value = loc


def _setup_cover_sheet(ws, product, location, client, media_path):
    ws.title = 'General Information'
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 180
    ws.merge_cells('A1:B1')
    cell = ws['A1']
    cell.fill = PatternFill('solid', start_color=Color('1565C0'))
    cell.font = Font(sz=24, color=WHITE)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value = 'Product information'
    for i in range(3, 13):
        ws[f'A{i}'].font = Font(sz=12)
        ws[f'B{i}'].font = Font(sz=12)
    ws['A3'].value = 'Account ID'
    ws['B3'].value = product['owner']['id']
    ws['A4'].value = 'Account Name'
    ws['B4'].value = product['owner']['name']
    ws['A5'].value = 'Product ID'
    ws['B5'].value = product['id']
    ws['A6'].value = 'Product Name'
    ws['B6'].value = product['name']
    ws['A7'].value = 'Export datetime'
    ws['B7'].value = datetime.now().isoformat()
    ws['A8'].value = 'Product Category'
    ws['B8'].value = product['category']['name']
    ws['A9'].value = 'Product Icon file name'
    ws['A9'].font = Font(sz=14)
    ws['B9'].value = f'{product["id"]}.{product["icon"].split(".")[-1]}'
    _dump_image(
        f'{location}{product["icon"]}',
        f'{product["id"]}.{product["icon"].split(".")[-1]}',
        media_path,
    )
    ws['A10'].value = 'Product Short Description'
    ws['A10'].alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws['B10'].value = product['short_description']
    ws['B10'].alignment = Alignment(
        wrap_text=True,
    )
    ws['A11'].value = 'Product Detailed Description'
    ws['A11'].alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws['B11'].value = product['detailed_description']
    ws['B11'].alignment = Alignment(
        wrap_text=True,
    )
    ws['A12'].value = 'Embedding description'
    ws['B12'].value = product['customer_ui_settings']['description']
    ws['B12'].alignment = Alignment(
        wrap_text=True,
    )
    ws['A13'].value = 'Embedding getting started'
    ws['B13'].value = product['customer_ui_settings']['getting_started']
    ws['B13'].alignment = Alignment(
        wrap_text=True,
    )
    ws['A14'].value = 'Primary Translation Locale'
    primary_translation = (
        client.ns('localization').translations
        .filter(context__instance_id=product['id'], primary=True).first()
    )
    ws['B14'].value = (
        f'{ primary_translation["locale"]["id"] } '
        f'({ primary_translation["locale"]["name"] })'
    )

    categories = client.categories.all()
    unassignable_cat = ['Cloud Services', 'All Categories']
    categories_list = [
        cat['name'] for cat in categories if cat['name'] not in unassignable_cat
    ]
    ws['AA1'].value = 'Categories'
    for idx, cat in enumerate(categories_list, 2):
        ws[f'AA{idx}'].value = cat
    categories_validation = DataValidation(
        type='list',
        formula1=f'{quote_sheetname("General Information")}!$AA$2:$AA${idx}',
        allow_blank=False,
    )
    ws.add_data_validation(categories_validation)
    categories_validation.add('B8')


def _dump_image(image_location, image_name, media_path):
    image = requests.get(image_location)
    if image.status_code == 200:
        with open(os.path.join(media_path, image_name), 'wb') as f:
            f.write(image.content)
    else:
        raise ClickException(f"Error obtaining image from {image_location}")


def _setup_ws_header(ws, ws_type=None):  # noqa: CCR001
    if not ws_type:
        ws_type = 'items'

    color = Color('d3d3d3')
    fill = PatternFill('solid', color)
    cels = ws['A1': '{cell}1'.format(
        cell=get_col_limit_by_ws_type(ws_type),
    )]
    col_headers = get_col_headers_by_ws_type(ws_type)
    if ws_type == '_attributes':
        col_headers = {c.column_letter: c.value for c in next(ws.iter_rows(min_row=1, max_row=1))}
    for cel in cels[0]:
        ws.column_dimensions[cel.column_letter].width = 25
        ws.column_dimensions[cel.column_letter].auto_size = True
        cel.fill = fill
        cel.value = col_headers[cel.column_letter]
        if ws_type == 'params' and cel.value == 'JSON Properties':
            ws.column_dimensions[cel.column_letter].width = 100
        elif ws_type == 'capabilities' and cel.value == 'Capability':
            ws.column_dimensions[cel.column_letter].width = 50
        elif ws_type == 'static_links' and cel.value == 'Url':
            ws.column_dimensions[cel.column_letter].width = 100
        elif ws_type == 'templates':
            if cel.value == 'Content':
                ws.column_dimensions[cel.column_letter].width = 100
            if cel.value == 'Title':
                ws.column_dimensions[cel.column_letter].width = 50
        elif ws_type == '_attributes':
            if cel.column_letter in ['A', 'B', 'D']:
                ws.column_dimensions[cel.column_letter].width = 100


def _calculate_commitment(item):
    period = item.get('period')
    if not period:
        return '-'
    commitment = item.get('commitment')
    if not commitment:
        return '-'
    count = commitment['count']
    if count == 1:
        return '-'

    multiplier = commitment['multiplier']

    if multiplier == 'billing_period':
        if period == 'monthly':
            years = count // 12
            return '{quantity} year{plural}'.format(
                quantity=years,
                plural='s' if years > 1 else '',
            )
        else:
            return '{years} years'.format(
                years=count,
            )

    # One-time
    return '-'


def _fill_param_row(ws, row_idx, param):
    ws.cell(row_idx, 1, value=param['id']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 2, value=param['name']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 3, value='-').alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 4, value=param['title']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 5, value=param['description']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 6, value=param['phase']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 7, value=param['scope']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 8, value=param['type']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(
        row_idx, 9,
        value=param['constraints']['required'] if param['constraints']['required'] else '-',
    ).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(
        row_idx, 10,
        value=param['constraints']['unique'] if param['constraints']['unique'] else '-',
    ).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(
        row_idx, 11,
        value=param['constraints']['hidden'] if param['constraints']['hidden'] else '-',
    ).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(
        row_idx, 12,
        value=get_json_object_for_param(param),
    ).alignment = Alignment(
        wrap_text=True,
    )
    events = param.get('events', {})
    ws.cell(
        row_idx, 13, value=events.get('created', {}).get('at', '-'),
    ).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(
        row_idx, 14, value=events.get('updated', {}).get('at', '-'),
    ).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )


def _fill_media_row(ws, row_idx, media, location, product, media_path):
    ws.cell(row_idx, 1, value=media['position'])
    ws.cell(row_idx, 2, value=media['id'])
    ws.cell(row_idx, 3, value='-')
    ws.cell(row_idx, 4, value=media['type'])
    ws.cell(row_idx, 5, value=f'{media["id"]}.{media["thumbnail"].split(".")[-1]}')
    _dump_image(
        f'{location}{media["thumbnail"]}',
        f'{media["id"]}.{media["thumbnail"].split(".")[-1]}',
        media_path,
    )
    ws.cell(row_idx, 6, value='-' if media['type'] == 'image' else media['url'])


def _fill_template_row(ws, row_idx, template):
    ws.cell(row_idx, 1, value=template['id']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 2, value=template['title']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 3, value='-').alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 4, value=template['scope']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(
        row_idx, 5, value=template['type'] if 'type' in template else 'fulfillment',
    ).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 6, value=template['body']).alignment = Alignment(
        wrap_text=True,
    )
    events = template.get('events', {})
    ws.cell(
        row_idx, 7, value=events.get('created', {}).get('at', '-'),
    ).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(
        row_idx, 8, value=events.get('updated', {}).get('at', '-'),
    ).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )


def _fill_action_row(ws, row_idx, action):
    ws.cell(row_idx, 1, value=action['id'])
    ws.cell(row_idx, 2, value=action['action'])
    ws.cell(row_idx, 3, value='-')
    ws.cell(row_idx, 4, value=action['name'])
    ws.cell(row_idx, 5, value=action['title'])
    ws.cell(row_idx, 6, value=action['description'])
    ws.cell(row_idx, 7, value=action['scope'])
    events = action.get('events', {})
    ws.cell(row_idx, 8, value=events.get('created', {}).get('at', '-'))
    ws.cell(row_idx, 9, value=events.get('updated', {}).get('at', '-'))


def _fill_configuration_row(ws, row_idx, configuration, conf_id):
    ws.cell(row_idx, 1, value=conf_id)
    ws.cell(row_idx, 2, value=configuration['parameter']['id'])
    ws.cell(row_idx, 3, value=configuration['parameter']['scope'])
    ws.cell(row_idx, 4, value='-')
    ws.cell(row_idx, 5, value=configuration['item']['id'] if 'item' in configuration else '-')
    ws.cell(row_idx, 6, value=configuration['item']['name'] if 'item' in configuration else '-')
    ws.cell(row_idx, 7, value=configuration['marketplace']['id'] if 'marketplace' in configuration else '-')
    ws.cell(row_idx, 8,
            value=configuration['marketplace']['name'] if 'marketplace' in configuration else '-')
    if 'structured_value' in configuration:
        value = configuration['structured_value']
        value = json.dumps(value, indent=4, sort_keys=True)
        ws.cell(row_idx, 9, value=value).alignment = Alignment(wrap_text=True)
    elif 'value' in configuration:
        ws.cell(row_idx, 9, value=configuration['value'])
    else:
        ws.cell(row_idx, 9, value='-')


def _fill_item_row(ws, row_idx, item):
    ws.cell(row_idx, 1, value=item['id'])
    ws.cell(row_idx, 2, value=item['mpn'])
    ws.cell(row_idx, 3, value='-')
    ws.cell(row_idx, 4, value=item['display_name'])
    ws.cell(row_idx, 5, value=item['description'])
    ws.cell(row_idx, 6, value=item['type'])
    ws.cell(row_idx, 7, value=item['precision'])
    ws.cell(row_idx, 8, value=item['unit']['unit'])
    period = item.get('period', 'monthly')
    if period.startswith('years_'):
        period = f'{period.rsplit("_")[-1]} years'
    ws.cell(row_idx, 9, value=period)
    ws.cell(row_idx, 10, value=_calculate_commitment(item))
    ws.cell(row_idx, 11, value=item['status'])
    events = item.get('events', {})
    ws.cell(row_idx, 12, value=events.get('created', {}).get('at', '-'))
    ws.cell(row_idx, 13, value=events.get('updated', {}).get('at', '-'))


def _calculate_configuration_id(configuration):
    conf_id = configuration['parameter']['id']
    if 'item' in configuration and 'id' in configuration['item']:
        conf_id = f'{conf_id}#{configuration["item"]["id"]}'
    else:
        conf_id = f'{conf_id}#'
    if 'marketplace' in configuration and 'id' in configuration['marketplace']:
        conf_id = f'{conf_id}#{configuration["marketplace"]["id"]}'
    else:
        conf_id = f'{conf_id}#'

    return conf_id


def _dump_actions(ws, client, product_id, progress):
    _setup_ws_header(ws, 'actions')

    row_idx = 2

    actions = client.products[product_id].actions.all()
    count = actions.count()

    action_validation = DataValidation(
        type='list',
        formula1='"-,create,update,delete"',
        allow_blank=False,
    )

    scope_validation = DataValidation(
        type='list',
        formula1='"asset,tier1,tier2"',
        allow_blank=False,
    )

    if count > 0:
        ws.add_data_validation(action_validation)
        ws.add_data_validation(scope_validation)

    task = progress.add_task('Processing action', total=count)

    for action in actions:
        progress.update(task, description=f'Processing action {action["id"]}', advance=1)
        _fill_action_row(ws, row_idx, action)
        action_validation.add(f'C{row_idx}')
        scope_validation.add(f'G{row_idx}')
        row_idx += 1

    progress.update(task, completed=count)


def _dump_configuration(ws, client, product_id, progress):
    _setup_ws_header(ws, 'configurations')

    row_idx = 2

    configurations = client.products[product_id].configurations.all()
    count = configurations.count()

    action_validation = DataValidation(
        type='list',
        formula1='"-,update,delete"',
        allow_blank=False,
    )

    if count == 0:
        return

    ws.add_data_validation(action_validation)

    task = progress.add_task('Processing parameter configuration', total=count)

    for configuration in configurations:
        conf_id = _calculate_configuration_id(configuration)
        progress.update(task, description=f'Processing parameter configuration {conf_id}', advance=1)
        _fill_configuration_row(ws, row_idx, configuration, conf_id)
        action_validation.add(f'D{row_idx}')
        row_idx += 1

    progress.update(task, completed=count)


def _dump_parameters(ws, client, product_id, param_type, progress):
    _setup_ws_header(ws, 'params')

    rql = R().phase.eq(param_type)

    row_idx = 2

    params = client.products[product_id].parameters.filter(rql)
    count = params.count()

    if count == 0:
        # Product without params is strange, but may exist
        return

    action_validation = DataValidation(
        type='list',
        formula1='"-,create,update,delete"',
        allow_blank=False,
    )
    type_validation = DataValidation(
        type='list',
        formula1='"{params}"'.format(
            params=','.join(PARAM_TYPES),
        ),
        allow_blank=False,
    )
    ordering_fulfillment_scope_validation = DataValidation(
        type='list',
        formula1='"asset,tier1,tier2"',
        allow_blank=False,
    )
    configuration_scope_validation = DataValidation(
        type='list',
        formula1='"product,marketplace,item,item_marketplace"',
        allow_blank=False,
    )
    bool_validation = DataValidation(
        type='list',
        formula1='"True,-"',
        allow_blank=False,
    )
    ws.add_data_validation(action_validation)
    ws.add_data_validation(type_validation)
    ws.add_data_validation(ordering_fulfillment_scope_validation)
    ws.add_data_validation(configuration_scope_validation)
    ws.add_data_validation(bool_validation)

    task = progress.add_task(f'Processing {param_type} parameter', total=count)

    for param in params:
        progress.update(
            task,
            description=f'Processing {param_type} parameter {param["id"]}',
            advance=1,
        )
        _fill_param_row(ws, row_idx, param)
        action_validation.add(f'C{row_idx}')
        if param['phase'] == 'configuration':
            configuration_scope_validation.add(f'G{row_idx}')
        else:
            ordering_fulfillment_scope_validation.add(f'G{row_idx}')
        type_validation.add(f'H{row_idx}')
        bool_validation.add(f'I{row_idx}')
        bool_validation.add(f'J{row_idx}')
        bool_validation.add(f'K{row_idx}')
        row_idx += 1

    progress.update(task, completed=count)


def _dump_media(ws, client, product_id, media_location, media_path, progress):
    _setup_ws_header(ws, 'media')
    row_idx = 2

    medias = client.products[product_id].media.all()
    count = medias.count()
    action_validation = DataValidation(
        type='list',
        formula1='"-,create,update,delete"',
        allow_blank=False,
    )
    type_validation = DataValidation(
        type='list',
        formula1='"image,video"',
        allow_blank=False,
    )
    if count > 0:
        ws.add_data_validation(action_validation)
        ws.add_data_validation(type_validation)

    task = progress.add_task('Processing media', total=count)
    for media in medias:
        progress.update(task, description=f'Processing media {media["id"]}', advance=1)
        _fill_media_row(ws, row_idx, media, media_location, product_id, media_path)
        action_validation.add(f'C{row_idx}')
        type_validation.add(f'D{row_idx}')
        row_idx += 1

    progress.update(task, completed=count)


def _dump_external_static_links(ws, product, progress):
    _setup_ws_header(ws, 'static_links')
    row_idx = 2
    count = len(product['customer_ui_settings']['download_links'])
    count = count + len(product['customer_ui_settings']['documents'])

    action_validation = DataValidation(
        type='list',
        formula1='"-,create,delete"',
        allow_blank=False,
    )
    link_type = DataValidation(
        type='list',
        formula1='"Download,Documentation"',
        allow_blank=False,
    )
    if count > 0:
        ws.add_data_validation(action_validation)
        ws.add_data_validation(link_type)

    task = progress.add_task('Processing static links', total=count)

    for link in product['customer_ui_settings']['download_links']:
        progress.update(task, advance=1)
        ws.cell(row_idx, 1, value='Download')
        ws.cell(row_idx, 2, value=link['title'])
        ws.cell(row_idx, 3, value='-')
        ws.cell(row_idx, 4, value=link['url'])
        action_validation.add(f'C{row_idx}')
        link_type.add(f'A{row_idx}')
        row_idx += 1

    for link in product['customer_ui_settings']['documents']:
        progress.update(task, advance=1)
        ws.cell(row_idx, 1, value='Documentation')
        ws.cell(row_idx, 2, value=link['title'])
        ws.cell(row_idx, 3, value='-')
        ws.cell(row_idx, 4, value=link['url'])
        action_validation.add(f'C{row_idx}')
        link_type.add(f'A{row_idx}')
        row_idx += 1

    progress.update(task, completed=count)


def _dump_capabilities(ws, product, progress):  # noqa: CCR001
    _setup_ws_header(ws, 'capabilities')
    task = progress.add_task('Processing product capabilities', total=1)
    capabilities = product['capabilities']
    ppu = product['capabilities'].get('ppu', False)
    tiers = capabilities['tiers']
    subscription = capabilities['subscription']
    change = subscription['change']

    action_validation = DataValidation(
        type='list',
        formula1='"-,update"',
        allow_blank=False,
    )
    ppu_validation = DataValidation(
        type='list',
        formula1='"Disabled,QT,TR,PR"',
        allow_blank=False,
    )
    disabled_enabled = DataValidation(
        type='list',
        formula1='"Disabled,Enabled"',
        allow_blank=False,
    )
    tier_validation = DataValidation(
        type='list',
        formula1='"Disabled,1,2"',
        allow_blank=False,
    )
    ws.add_data_validation(action_validation)
    ws.add_data_validation(ppu_validation)
    ws.add_data_validation(disabled_enabled)
    ws.add_data_validation(tier_validation)

    ws['A2'].value = 'Pay-as-you-go support and schema'
    ws['B2'].value = '-'
    ws['C2'].value = (ppu['schema'] if ppu else 'Disabled')
    ppu_validation.add(ws['C2'])
    ws['A3'].value = 'Pay-as-you-go dynamic items support'
    ws['B3'].value = '-'
    ws['C3'].value = (
        'Enabled' if ppu and 'dynamic' in ppu and ppu['dynamic'] else 'Disabled'
    )
    disabled_enabled.add(ws['C3'])
    ws['A4'].value = 'Pay-as-you-go future charges support'
    ws['B4'].value = '-'
    ws['C4'].value = (
        'Enabled' if ppu and 'future' in ppu and ppu['future'] else 'Disabled'
    )
    disabled_enabled.add(ws['C4'])
    ws['A5'].value = 'Consumption reporting for Reservation Items'
    ws['B5'].value = '-'
    ws['C5'].value = (
        'Enabled' if capabilities['reservation']['consumption'] else 'Disabled'
    )
    disabled_enabled.add(ws['C5'])

    def _get_reporting_consumption(reservation_cap):
        if 'consumption' in reservation_cap and reservation_cap['consumption']:
            return 'Enabled'
        return 'Disabled'

    ws['C5'].value = _get_reporting_consumption(capabilities['reservation'])
    disabled_enabled.add(ws['C5'])
    ws['A6'].value = 'Dynamic Validation of the Draft Requests'
    ws['B6'].value = '-'

    def _get_dynamic_validation_draft(capabilities_cart):
        if 'validation' in capabilities_cart and capabilities['cart']['validation']:
            return 'Enabled'
        return 'Disabled'
    ws['C6'].value = _get_dynamic_validation_draft(capabilities['cart'])
    disabled_enabled.add(ws['C6'])
    ws['A7'].value = 'Dynamic Validation of the Inquiring Form'
    ws['B7'].value = '-'

    def _get_validation_inquiring(capabilities_inquiring):
        if 'validation' in capabilities_inquiring and capabilities_inquiring['validation']:
            return 'Enabled'
        return 'Disabled'

    ws['C7'].value = _get_validation_inquiring(capabilities['inquiring'])
    disabled_enabled.add(ws['C7'])
    ws['A8'].value = 'Reseller Authorization Level'
    ws['B8'].value = '-'

    def _get_reseller_authorization_level(tiers):
        if tiers and 'configs' in tiers and tiers['configs']:
            return tiers['configs']['level']
        return 'Disabled'

    ws['C8'].value = _get_reseller_authorization_level(tiers)
    tier_validation.add(ws['C8'])
    ws['A9'].value = 'Tier Accounts Sync'
    ws['B9'].value = '-'
    ws['C9'].value = (
        'Enabled' if tiers and 'updates' in tiers and tiers['updates'] else 'Disabled'
    )
    disabled_enabled.add(ws['C9'])
    ws['A10'].value = 'Administrative Hold'
    ws['B10'].value = '-'

    ws['A11'].value = 'Dynamic Validation of Tier Requests'
    ws['B11'].value = '-'
    ws['C11'].value = (
        'Enabled' if capabilities['tiers']['validation'] else 'Disabled'
    )
    disabled_enabled.add(ws['C11'])
    ws['A12'].value = 'Editable Ordering Parameters in Change Request'
    ws['B12'].value = '-'
    ws['C12'].value = (
        'Enabled' if subscription['change']['editable_ordering_parameters'] else 'Disabled'
    )
    disabled_enabled.add(ws['C12'])
    ws['A13'].value = 'Validation of Draft Change Request'
    ws['B13'].value = '-'
    ws['C13'].value = (
        'Enabled' if 'validation' in change and change['validation'] else 'Disabled'
    )
    disabled_enabled.add(ws['C13'])
    ws['A14'].value = 'Validation of inquiring form for Change Requests'
    ws['B14'].value = '-'
    ws['C14'].value = (
        'Enabled' if 'inquiring_validation' in change and change['inquiring_validation'] else 'Disabled'
    )
    disabled_enabled.add(ws['C14'])

    def _get_administrative_hold(capabilities):
        if 'hold' in capabilities['subscription'] and capabilities['subscription']['hold']:
            return 'Enabled'
        return 'Disabled'

    ws['C10'].value = _get_administrative_hold(capabilities)
    disabled_enabled.add(ws['C10'])
    idx = 2
    while idx < 11:
        action_validation.add(f'B{idx}')
        idx = idx + 1

    progress.update(task, advance=1)


def _dump_templates(ws, client, product_id, progress):
    _setup_ws_header(ws, 'templates')

    row_idx = 2

    action_validation = DataValidation(
        type='list',
        formula1='"-,create,update,delete"',
        allow_blank=False,
    )
    scope_validation = DataValidation(
        type='list',
        formula1='"asset,tier1,tier2"',
        allow_blank=False,
    )
    type_validation = DataValidation(
        type='list',
        formula1='"pending,fulfillment,inquire"',
        allow_blank=False,
    )

    templates = client.products[product_id].templates.all()
    count = templates.count()

    if count > 0:
        ws.add_data_validation(action_validation)
        ws.add_data_validation(scope_validation)
        ws.add_data_validation(type_validation)

    task = progress.add_task('Processing template', total=count)

    for template in templates:
        progress.update(task, description=f'Processing template {template["id"]}', advance=1)
        _fill_template_row(ws, row_idx, template)
        action_validation.add(f'C{row_idx}')
        scope_validation.add(f'D{row_idx}')
        type_validation.add(f'E{row_idx}')
        row_idx += 1

    progress.update(task, completed=count)


def _dump_items(ws, client, product_id, progress):
    _setup_ws_header(ws, 'items')

    row_idx = 2

    items = client.products[product_id].items.all()
    count = items.count()

    if count == 0:
        raise ClickException(f'The product {product_id} doesn\'t have items.')

    action_validation = DataValidation(
        type='list',
        formula1='"-,create,update,delete"',
        allow_blank=False,
    )
    type_validation = DataValidation(
        type='list',
        formula1='"reservation,ppu"',
        allow_blank=False,
    )
    period_validation = DataValidation(
        type='list',
        formula1='"onetime,monthly,yearly,2 years,3 years,4 years,5 years"',
        allow_blank=False,
    )

    precision_validation = DataValidation(
        type='list',
        formula1='"integer,decimal(1),decimal(2),decimal(4),decimal(8)"',
        allow_blank=False,
    )

    commitment_validation = DataValidation(
        type='list',
        formula1='"-,1 year,2 years,3 years,4 years,5 years"',
        allow_blank=False,
    )

    ws.add_data_validation(action_validation)
    ws.add_data_validation(type_validation)
    ws.add_data_validation(period_validation)
    ws.add_data_validation(precision_validation)
    ws.add_data_validation(commitment_validation)

    task = progress.add_task('Processing item', total=count)

    for item in items:
        progress.update(task, description=f'Processing item {item["id"]}', advance=1)
        _fill_item_row(ws, row_idx, item)
        action_validation.add(f'C{row_idx}')
        type_validation.add(f'F{row_idx}')
        precision_validation.add(f'G{row_idx}')
        period_validation.add(f'I{row_idx}')
        commitment_validation.add(f'J{row_idx}')
        row_idx += 1

    progress.update(task, completed=count)


def _dump_translations(wb, client, product_id, progress):
    ws = wb.create_sheet('Translations')
    _setup_ws_header(ws, 'translations')
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15

    rql = R().context.instance_id.eq(product_id)

    translations = (
        client.ns('localization')
        .translations
        .filter(rql)
    )
    count = translations.count()

    action_validation = DataValidation(
        type='list',
        formula1='"-,delete,update,create"',
        allow_blank=False,
    )
    no_action_validation = DataValidation(
        type='list',
        formula1='"-"',
        allow_blank=False,
    )
    disabled_enabled = DataValidation(
        type='list',
        formula1='"Disabled,Enabled"',
        allow_blank=False,
    )
    ws.add_data_validation(action_validation)
    ws.add_data_validation(no_action_validation)
    ws.add_data_validation(disabled_enabled)

    task = progress.add_task('Processing translation', total=count)

    for row_idx, translation in enumerate(translations, 2):
        progress.update(task, description=f'Processing translation {translation["id"]}', advance=1)
        fill_translation_row(ws, row_idx, translation)
        if translation['primary']:
            no_action_validation.add(ws[f'B{row_idx}'])
        else:
            action_validation.add(ws[f'B{row_idx}'])
        disabled_enabled.add(ws[f'I{row_idx}'])
        _dump_translation_attr(wb, client, translation)

    setup_locale_data_validation(wb['General Information'], ws)
    progress.update(task, completed=count)


def _dump_translation_attr(wb, client, translation):
    external_wb = get_translation_workbook(client, translation['id'])
    attr_ws = wb.create_sheet(f'{translation["locale"]["id"]} ({translation["id"]})')
    for row in external_wb['Attributes']:
        for cell in row:
            attr_ws[cell.coordinate].value = cell.value
            attr_ws[cell.coordinate].alignment = copy.copy(cell.alignment)
    alter_attributes_sheet(attr_ws)
    _setup_ws_header(attr_ws, '_attributes')


def dump_product(  # noqa: CCR001
    client, product_id, output_file, progress, output_path=None,
):
    output_file = validate_output_options(output_path, output_file, default_dir_name=product_id)
    media_path = os.path.join(os.path.dirname(output_file), 'media')
    if not os.path.exists(media_path):
        os.mkdir(media_path)
    try:

        product = client.products[product_id].get()
        wb = Workbook()

        _setup_locales_list(wb.active, client)

        connect_api_location = parse.urlparse(client.endpoint)
        media_location = f'{connect_api_location.scheme}://{connect_api_location.netloc}'
        _setup_cover_sheet(
            wb.active,
            product,
            media_location,
            client,
            media_path,
        )
        _dump_capabilities(wb.create_sheet('Capabilities'), product, progress)
        _dump_external_static_links(wb.create_sheet('Embedding Static Resources'), product, progress)
        _dump_media(
            wb.create_sheet('Media'),
            client,
            product_id,
            media_location,
            media_path,
            progress,
        )
        _dump_templates(wb.create_sheet('Templates'), client, product_id, progress)
        _dump_items(wb.create_sheet('Items'), client, product_id, progress)
        _dump_parameters(
            wb.create_sheet('Ordering Parameters'),
            client,
            product_id,
            'ordering',
            progress,
        )
        _dump_parameters(
            wb.create_sheet('Fulfillment Parameters'),
            client,
            product_id,
            'fulfillment',
            progress,
        )
        _dump_parameters(
            wb.create_sheet('Configuration Parameters'),
            client,
            product_id,
            'configuration',
            progress,
        )
        _dump_actions(wb.create_sheet('Actions'), client, product_id, progress)
        _dump_configuration(wb.create_sheet('Configuration'), client, product_id, progress)
        _dump_translations(wb, client, product_id, progress)
        wb.save(output_file)

    except ClientError as error:
        status = format_http_status(error.status_code)
        if error.status_code == 404:
            raise ClickException(f'{status}: Product {product_id} not found.')

        handle_http_error(error)

    return output_file

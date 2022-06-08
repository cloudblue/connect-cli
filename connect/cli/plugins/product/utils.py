import re
import json
from copy import deepcopy

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

from connect.cli.plugins.product.constants import (
    ACTIONS_HEADERS,
    CAPABILITIES_COLS_HEADERS,
    CONFIGURATION_HEADERS,
    ITEMS_COLS_HEADERS,
    MEDIA_COLS_HEADERS,
    PARAMS_COLS_HEADERS,
    STATIC_LINK_HEADERS,
    TEMPLATES_HEADERS,
    TRANSLATION_HEADERS,
)


def get_col_limit_by_ws_type(ws_type):
    if ws_type == 'items':
        return 'M'
    elif ws_type == 'params':
        return 'N'
    elif ws_type == 'media':
        return 'F'
    elif ws_type == 'capabilities':
        return 'C'
    elif ws_type == 'static_links':
        return 'D'
    elif ws_type == 'templates':
        return 'H'
    elif ws_type == 'configurations':
        return 'I'
    elif ws_type == 'actions':
        return 'I'
    elif ws_type == 'translations':
        return 'N'
    elif ws_type == '_attributes':
        return 'F'
    return 'Z'


def get_ws_type_by_worksheet_name(ws_name):
    if ws_name == 'Items':
        return 'items'
    elif ws_name == "Ordering Parameters":
        return 'params'
    elif ws_name == 'Fulfillment Parameters':
        return 'params'
    elif ws_name == 'Configuration Parameters':
        return 'params'
    elif ws_name == 'Media':
        return 'media'
    elif ws_name == 'Capabilities':
        return 'capabilities'
    elif ws_name == 'Embedding Static Resources':
        return 'static_links'
    elif ws_name == 'Templates':
        return 'templates'
    elif ws_name == 'Configuration':
        return 'configurations'
    elif ws_name == 'Actions':
        return 'actions'
    elif ws_name == 'Translations':
        return 'translations'
    return None


def get_col_headers_by_ws_type(ws_type):
    if ws_type == 'items':
        return ITEMS_COLS_HEADERS
    elif ws_type == 'params':
        return PARAMS_COLS_HEADERS
    elif ws_type == 'media':
        return MEDIA_COLS_HEADERS
    elif ws_type == 'capabilities':
        return CAPABILITIES_COLS_HEADERS
    elif ws_type == 'static_links':
        return STATIC_LINK_HEADERS
    elif ws_type == 'templates':
        return TEMPLATES_HEADERS
    elif ws_type == 'configurations':
        return CONFIGURATION_HEADERS
    elif ws_type == 'actions':
        return ACTIONS_HEADERS
    elif ws_type == 'translations':
        return TRANSLATION_HEADERS


def cleanup_product_for_update(product):
    del product['icon']
    ppu = product['capabilities'].get('ppu', False)
    if product['capabilities']['subscription'] and 'schema' in product['capabilities']['subscription']:
        del product['capabilities']['subscription']['schema']
    if ppu and 'predictive' in ppu:
        del product['capabilities']['ppu']['predictive']
    return product


def get_json_object_for_param(original_param):
    param = deepcopy(original_param)
    del param['id']
    del param['name']
    del param['title']
    del param['description']
    del param['phase']
    del param['scope']
    del param['type']
    del param['constraints']['required']
    del param['constraints']['unique']
    del param['constraints']['hidden']
    del param['position']
    del param['events']

    return json.dumps(param, indent=4, sort_keys=True)


def fill_translation_row(ws, row_idx, translation, update_mode=False):
    """
    Fill translations worksheet row with data from a translation.
    if 'update_mode' is True, then skip some cells that should not be overwritten.
    """
    ws.cell(row_idx, 1, value=translation['id'])
    ws.cell(row_idx, 3, value=translation['context']['instance_id'])
    ws.cell(row_idx, 4, value=translation['context']['name'])
    ws.cell(row_idx, 5, value=translation['owner']['id'])
    ws.cell(row_idx, 6, value=translation['owner']['name'])
    ws.cell(row_idx, 7, value=f"{translation['locale']['id']} ({translation['locale']['name']})")
    ws.cell(row_idx, 10, value=_calculate_translation_completion(translation)).number_format = '0%'
    ws.cell(row_idx, 11, value=translation['status'])
    ws.cell(row_idx, 12, value='Yes' if translation['primary'] else 'No')
    ws.cell(row_idx, 13, value=translation['events'].get('created', {}).get('at', '-'))
    ws.cell(row_idx, 14, value=translation['events'].get('updated', {}).get('at', '-'))
    if not update_mode:
        ws.cell(row_idx, 2, value='-')
        ws.cell(row_idx, 8, value=translation.get('description', '-') or '-').alignment = Alignment(wrap_text=True)
        ws.cell(row_idx, 9, value='Enabled' if translation['auto']['enabled'] else 'Disabled')


def _calculate_translation_completion(translation):
    stats = translation['stats']
    try:
        return (
            stats.get('translated') / stats.get('total')
        )
    except TypeError:
        return '-'


def setup_locale_data_validation(general_ws, translations_ws):
    """
    Setup DataValidation on locale column.
    It is necessary to setup DataValidation every time the worksheet is saved to ensure that it is
    not removed.
    This is related to the openpyxl warning: DataValidation extension not supported and will be
    removed.
    """
    row_idx = 2
    while general_ws[f'AB{row_idx + 1}'].value:
        row_idx += 1

    locales_validation = DataValidation(
        type='list',
        formula1='{sheet_name}!${column}$2:${column}${last_row_idx}'.format(
            sheet_name=quote_sheetname(general_ws.title),
            column='AB',
            last_row_idx=row_idx,
        ),
        allow_blank=False,
    )
    translations_ws.add_data_validation(locales_validation)
    locales_validation.add(f'G2:G{translations_ws.max_row}')


def get_translation_attributes_sheets(wb_filename):
    """
    return the list of worksheet names corresponding to translation attributes sheets in the given
    workbook file.
    """
    pattern = r'^[\w-]{2,} \(TRN-\d{4}-\d{4}-\d{4}\)$'
    wb = load_workbook(wb_filename)
    return [sheetname for sheetname in wb.sheetnames if re.match(pattern, sheetname)]


class ParamSwitchNotSupported(Exception):
    pass

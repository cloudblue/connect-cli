# -*- coding: utf-8 -*-

# This file is part of the Ingram Micro Cloud Blue Connect connect-cli.
# Copyright (c) 2019-2022 Ingram Micro. All Rights Reserved.

from openpyxl.worksheet.datavalidation import DataValidation

from connect.cli.core.utils import validate_output_options
from connect.cli.plugins.shared.export import (
    alter_attributes_sheet,
    get_translation_workbook,
)


def dump_translation(
    client, translation_id, output_file, output_path=None,
):
    output_file = validate_output_options(output_path, output_file, default_dir_name=translation_id)
    wb = get_translation_workbook(client, translation_id)
    _alter_general_sheet(wb['General'])
    alter_attributes_sheet(wb['Attributes'])
    wb.save(output_file)
    return output_file


def _alter_general_sheet(ws):
    for row_idx in range(1, ws.max_row + 1):  # pragma: no branch
        if ws.cell(row_idx, 1).value.lower() == 'auto-translation':
            disabled_enabled = DataValidation(
                type='list',
                formula1='"Disabled,Enabled"',
                allow_blank=False,
            )
            ws.add_data_validation(disabled_enabled)
            disabled_enabled.add(ws.cell(row_idx, 2))
            break

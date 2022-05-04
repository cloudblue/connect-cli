# -*- coding: utf-8 -*-

# This file is part of the Ingram Micro Cloud Blue Connect connect-cli.
# Copyright (c) 2019-2022 Ingram Micro. All Rights Reserved.

from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.worksheet.datavalidation import DataValidation
from tqdm import tqdm

from connect.cli.core.constants import DEFAULT_BAR_FORMAT
from connect.cli.plugins.translation.constants import ATTRIBUTES_COLS


def dump_translation_attributes(ws, client, translation_id, silent):
    _prepare_attributes_worksheet(ws)
    attributes = client.ns('localization').translations[translation_id].attributes.all()
    count = attributes.count()
    action_validation = DataValidation(
        type='list',
        formula1='"-,update"',
        allow_blank=False,
    )
    if count > 0:
        ws.add_data_validation(action_validation)

    progress = tqdm(
        enumerate(attributes, start=2), total=count, disable=silent, leave=True,
        bar_format=DEFAULT_BAR_FORMAT,
    )
    for row_idx, attribute in progress:
        progress.set_description(f"Processing attribute {attribute['key']}")
        _fill_attribute_row(ws, row_idx, attribute)
        action_validation.add(f'B{row_idx}')


def _prepare_attributes_worksheet(ws):
    color = Color('d3d3d3')
    fill = PatternFill('solid', color)
    cels = ws['A1': 'D1']
    for cel in cels[0]:
        cel.fill = fill
        ws.column_dimensions[cel.column_letter].width = ATTRIBUTES_COLS[cel.column_letter].width
        ws.column_dimensions[cel.column_letter].auto_size = True
        cel.value = ATTRIBUTES_COLS[cel.column_letter].header


def _fill_attribute_row(ws, row_idx, attribute):
    ws.cell(row_idx, 1, value=attribute['key'])
    ws.cell(row_idx, 2, value='-')
    ws.cell(row_idx, 3, value=attribute.get('value', '-')).alignment = Alignment(wrap_text=True)
    ws.cell(row_idx, 4, value=attribute.get('comment', '-'))

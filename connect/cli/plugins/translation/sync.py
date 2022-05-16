from collections import namedtuple
from time import sleep
from types import SimpleNamespace

import click
from tqdm import tqdm, trange

from connect.client import ClientError

from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from connect.cli.core.constants import DEFAULT_BAR_FORMAT
from connect.cli.core.http import handle_http_error
from connect.cli.plugins.shared.sync_stats import SynchronizerStats
from connect.cli.plugins.shared.exceptions import SheetNotFoundError
from connect.cli.plugins.translation.constants import (
    ATTRIBUTES_SHEET_COLUMNS, GENERAL_SHEET_FIELDS,
)


class TranslationSynchronizer:
    """
    Synchronize a translation from excel file. It may update an existing
    translation or create a new one depending on some checks.
    """
    DEFAULT_WAIT_SECONDS = 1

    def __init__(self, client, silent, account_id):
        self._client = client
        self._silent = silent
        self._wb = None
        self.account_id = account_id
        self.stats = SynchronizerStats()

    def open(self, input_file):
        self._open_workbook(input_file)
        if not {'General', 'Attributes'}.issubset(self._wb.sheetnames):
            raise SheetNotFoundError(
                "File must contain worksheets 'General' and 'Attributes' to synchronize, skipping",
            )
        self._validate_general_worksheet(self._wb['General'])
        self._validate_attributes_worksheet(self._wb['Attributes'])

    def save(self, output_file):
        self._wb.save(output_file)

    def sync(self, yes):
        ws = self._wb['General']
        general_data = self._read_general_data(ws)
        current_translation = self._get_translation(general_data.translation_id)
        do_create = (
            not current_translation
            or self._check_create(current_translation, general_data)
        )
        translation_id = None
        if do_create:
            if not yes:
                click.confirm(
                    click.style("A new translation will be created.\n", fg='yellow')
                    + "The owner will be the current active account, "
                    f"locale {general_data.locale_id} "
                    f"and context '{general_data.context_name}' ({general_data.context_id}).\n"
                    "Do you want to continue?",
                    abort=True,
                )
            new_translation = self._create_translation(general_data)
            if new_translation:
                self._update_general_sheet(ws, new_translation)
                translation_id = new_translation['id']
                if not self._silent:  # pragma: no cover
                    click.secho(f"\nCreated new translation {translation_id}\n", fg='yellow')
        else:
            translation_id = self._update_translation(general_data)

        if not translation_id:
            return

        if general_data.auto_enabled == 'Enabled' and (
            do_create or not current_translation['auto']['enabled']
        ):
            self._wait_for_autotranslation(translation_id)
        attributes = self._collect_attributes_to_update(self._wb['Attributes'])
        if attributes:
            self._update_attributes(translation_id, attributes, self._wb['Attributes'])

    def _open_workbook(self, input_file):
        try:
            self._wb = load_workbook(input_file, data_only=True)
        except InvalidFileException as ife:
            raise click.ClickException(str(ife))
        except BadZipFile:
            raise click.ClickException(f'{input_file} is not a valid xlsx file.')

    @staticmethod
    def _validate_general_worksheet(ws):
        for settings in GENERAL_SHEET_FIELDS.values():
            cell = ws.cell(settings.row_idx, 1)
            if cell.value != settings.title:
                raise click.ClickException(
                    f"{cell.coordinate} must be '{settings.title}', but it is '{cell.value}'",
                )

    @staticmethod
    def _validate_attributes_worksheet(ws):
        for col_idx, header in enumerate(ATTRIBUTES_SHEET_COLUMNS, 1):
            if header == 'original value':
                continue
            cell = ws.cell(1, col_idx)
            if cell.value != header:
                raise click.ClickException(
                    f"Column '{cell.coordinate}' must be '{header}', but it is '{cell.value}'",
                )

    def _get_translation(self, translation_id, raise_404=False):
        try:
            return self._client.ns('localization').translations[translation_id].get()
        except ClientError as error:
            if error.status_code == 404 and not raise_404:
                return None
            handle_http_error(error)

    def _check_create(self, translation, general_data):
        """
        Check if should create a new translation or update an existing one.
        """
        return (
            translation['owner']['id'] != self.account_id
            or translation['locale']['id'] != general_data.locale_id
            or translation['context']['id'] != general_data.context_id
        )

    def _create_translation(self, general_data):
        try:
            translation = self._client.ns('localization').translations.create({
                'context': {
                    'id': general_data.context_id,
                },
                'locale': {
                    'id': general_data.locale_id,
                },
                'description': general_data.description,
                'auto': {
                    'enabled': general_data.auto_enabled == 'Enabled',
                },
            })
            self.stats['Translation'].created()
            return translation
        except ClientError as e:
            self.stats['Translation'].error(
                f'Error while updating general translation information: {str(e)}',
            )

    @staticmethod
    def _update_general_sheet(ws, translation):
        ws['B1'].value = translation['id']
        ws['B2'].value = translation['owner']['id']
        ws['B3'].value = translation['owner']['name']
        ws['B5'].value = translation['context']['id']
        ws['B6'].value = translation['context']['instance_id']
        ws['B7'].value = translation['context']['name']

    def _update_translation(self, general_data):
        try:
            translation_resource = self._client.ns('localization').translations[general_data.translation_id]
            translation = translation_resource.update({
                'description': general_data.description,
                'auto': {
                    'enabled': general_data.auto_enabled == 'Enabled',
                },
            })
            self.stats['Translation'].updated()
            return translation['id']
        except ClientError as e:
            self.stats['Translation'].error(
                f'Error while updating general translation information: {str(e)}',
            )

    def _wait_for_autotranslation(self, translation_id, wait_seconds=None, max_counts=5):
        progress = trange(0, max_counts, disable=self._silent, leave=False, bar_format=DEFAULT_BAR_FORMAT)
        for _ in progress:
            progress.set_description('Waiting for pending translation tasks')
            sleep(wait_seconds or self.DEFAULT_WAIT_SECONDS)
            translation = self._get_translation(translation_id, raise_404=True)
            status = translation['auto']['status']
            if status == 'processing':
                pass
            elif status in ('on', 'off'):
                break
            elif status == 'error':
                translation['auto']['error_message']
                raise click.ClickException(
                    'The auto-translation task failed with error: '
                    + translation['auto']['error_message'],
                )
            else:
                raise click.ClickException(f'Unknown auto-translation status: {status}')
        else:
            raise click.ClickException('Timeout waiting for pending translation tasks')

    def _collect_attributes_to_update(self, ws):
        AttributeRow = namedtuple(
            'AttributeRow',
            (header.replace(' ', '_').lower() for header in ATTRIBUTES_SHEET_COLUMNS),
        )

        progress = tqdm(
            enumerate(ws.iter_rows(min_row=2, values_only=True), 2),
            total=ws.max_row - 1, disable=self._silent, leave=True, bar_format=DEFAULT_BAR_FORMAT,
        )

        attributes = {}
        for row_idx, row in progress:
            row = AttributeRow(*row)
            progress.set_description(f'Process attribute {row.key}')
            if row.action == 'update':
                attributes[row_idx] = {'key': row.key, 'value': row.value, 'comment': row.comment}
            else:
                self.stats['Attributes'].skipped()

        return attributes

    def _update_attributes(self, translation_id, attributes, ws):
        try:
            translation_res = self._client.ns('localization').translations[translation_id]
            translation_res.attributes.bulk_update(list(attributes.values()))
            self.stats['Attributes'].updated(len(attributes))
            for row_idx in attributes.keys():
                self._update_attributes_sheet_row(ws, row_idx)
        except ClientError as e:
            self.stats['Attributes'].error(
                f'Error while updating attributes: {str(e)}',
                range(1, len(attributes) + 1),
            )

    @staticmethod
    def _update_attributes_sheet_row(ws, row_idx):
        ws.cell(row_idx, 3, value='-')

    @staticmethod
    def _read_general_data(ws):
        return SimpleNamespace(**{
            field: ws.cell(settings.row_idx, 2).value for field, settings in GENERAL_SHEET_FIELDS.items()
        })

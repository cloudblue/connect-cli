from openpyxl.utils import get_column_letter


def insert_column_ws(ws, column, width=None):
    """
    Insert a column in a worksheet keeping the width of existing columns.
    """
    column_widths = _get_column_widths(ws, total_columns=max(column, ws.max_column))
    if width is None:
        width = column_widths[column - 1]
    column_widths.insert(column - 1, width)
    ws.insert_cols(idx=column, amount=1)
    _set_column_widths(ws, column_widths)


def _get_column_widths(ws, total_columns):
    return [
        ws.column_dimensions[
            get_column_letter(column)
        ].width for column in range(1, total_columns + 1)
    ]


def _set_column_widths(ws, column_widths):
    for column, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(column)].width = column_width

# -*- coding: utf-8 -*-

# This file is part of the Ingram Micro Cloud Blue Connect connect-cli.
# Copyright (c) 2019-2022 Ingram Micro. All Rights Reserved.

from types import SimpleNamespace

import click

from connect.client import ClientError

from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from connect.cli.core.http import handle_http_error
from connect.cli.core.terminal import console
from connect.cli.plugins.shared.sync_stats import SynchronizerStats
from connect.cli.plugins.shared.exceptions import SheetNotFoundError
from connect.cli.plugins.translation.constants import GENERAL_SHEET_FIELDS


class TranslationSynchronizer:
    """
    Synchronize a translation from excel file. It may update an existing
    translation or create a new one depending on some checks.
    """
    def __init__(self, client, account_id, stats=None):
        self._client = client
        self._wb = None
        self.account_id = account_id
        if stats is None:
            stats = SynchronizerStats()
        self._mstats = stats['Translation']

    def open(self, input_file):
        self._open_workbook(input_file)
        if 'General' not in self._wb.sheetnames:
            raise SheetNotFoundError("File does not contain worksheet 'General' to synchronize, skipping")
        self._validate_general_worksheet(self._wb['General'])

    def save(self, output_file):
        self._wb.save(output_file)

    def sync(self):
        """
        Updates or creates the translation. Return the Translation ID and a boolean indicating
        if should wait for autotranslation to finish.
        If the operation is not successful then returns (None, False).
        """
        ws = self._wb['General']
        general_data = self._read_general_data(ws)
        current_translation = self._get_translation(general_data.translation_id)
        do_create = (
            not current_translation
            or self._check_create(current_translation, general_data)
        )
        translation_id = None
        if do_create:
            new_translation = self._try_create_translation(current_translation, general_data)
            if new_translation:
                self._update_general_sheet(ws, new_translation)
                translation_id = new_translation['id']

                console.secho(f"\nCreated new translation {translation_id}\n", fg='yellow')
        else:
            translation_id = self._update_translation(general_data)

        should_wait_for_autotranslation = (
            translation_id and general_data.auto_enabled == 'Enabled'
            and (do_create or not current_translation['auto']['enabled'])
        )
        return translation_id, should_wait_for_autotranslation

    def _open_workbook(self, input_file):
        try:
            self._wb = load_workbook(input_file, data_only=True)
        except InvalidFileException as ife:
            raise click.ClickException(str(ife))
        except BadZipFile:
            raise click.ClickException(f'{input_file} is not a valid xlsx file.')

    @staticmethod
    def _validate_general_worksheet(ws):
        for settings in GENERAL_SHEET_FIELDS.values():
            cell = ws.cell(settings.row_idx, 1)
            if cell.value != settings.title:
                raise click.ClickException(
                    f"{cell.coordinate} must be '{settings.title}', but it is '{cell.value}'",
                )

    def _get_translation(self, translation_id):
        try:
            return self._client.ns('localization').translations[translation_id].get()
        except ClientError as error:
            if error.status_code == 404:
                return None
            handle_http_error(error)

    def _check_create(self, translation, general_data):
        """
        Check if should create a new translation or update an existing one.
        """
        return (
            translation['owner']['id'] != self.account_id
            or translation['locale']['id'] != general_data.locale_id
            or self._is_different_context(translation, general_data)
        )

    def _try_create_translation(self, current_translation, general_data):
        if (
            current_translation
            and self._is_different_context(current_translation, general_data)
        ):
            self._resolve_new_context(general_data)

        console.confirm(
            click.style("A new translation will be created.\n", fg='yellow')
            + "The owner will be the current active account, "
            f"locale {general_data.locale_id} "
            f"and context '{general_data.context_name}' ({general_data.context_id}).\n"
            "Do you want to continue?",
            abort=True,
        )

        try:
            translation = self._client.ns('localization').translations.create({
                'context': {
                    'id': general_data.context_id,
                },
                'locale': {
                    'id': general_data.locale_id,
                },
                'description': general_data.description,
                'auto': {
                    'enabled': general_data.auto_enabled == 'Enabled',
                },
            })
            self._mstats.created()
            return translation
        except ClientError as e:
            self._mstats.error(
                f'Error while creating translation: {str(e)}',
            )

    def _is_different_context(self, translation, general_data):
        return (
            general_data.context_id and general_data.context_id != translation['context']['id']
            or (
                general_data.context_instance_id
                and general_data.context_instance_id != translation['context']['instance_id']
            )
        )

    def _resolve_new_context(self, general_data):
        """Check that there is no ambiguity on the new specified context,
        and update the context_id accordingly."""
        try:
            if general_data.context_id:
                ctx = self._get_context(general_data.context_id)
                if not ctx:
                    raise click.ClickException(
                        f"The Context ID ({general_data.context_id}) doesn't exist",
                    )
                if (
                    general_data.context_instance_id
                    and general_data.context_instance_id != ctx['instance_id']
                ):
                    raise click.ClickException(
                        f"The Instance ID ({general_data.context_instance_id}) doesn't correspond "
                        f"to the Context ID ({general_data.context_id})",
                    )
                general_data.context_name = ctx['name']
            elif general_data.context_instance_id:  # pragma: no branch
                ctx = self._client.ns('localization').contexts.filter(
                    instance_id=general_data.context_instance_id,
                ).first()
                if not ctx:
                    raise click.ClickException(
                        f"The Instance ID ({general_data.context_instance_id}) doesn't exist",
                    )
                general_data.context_id = ctx['id']
                general_data.context_name = ctx['name']
        except ClientError as error:
            handle_http_error(error)

    def _get_context(self, context_id):
        try:
            return self._client.ns('localization').contexts[context_id].get()
        except ClientError as error:
            if error.status_code == 404:
                return None
            raise

    @staticmethod
    def _update_general_sheet(ws, translation):
        ws['B1'].value = translation['id']
        ws['B2'].value = translation['owner']['id']
        ws['B3'].value = translation['owner']['name']
        ws['B5'].value = translation['context']['id']
        ws['B6'].value = translation['context']['instance_id']
        ws['B7'].value = translation['context']['name']

    def _update_translation(self, general_data):
        try:
            translation_resource = self._client.ns('localization').translations[general_data.translation_id]
            translation = translation_resource.update({
                'description': general_data.description,
                'auto': {
                    'enabled': general_data.auto_enabled == 'Enabled',
                },
            })
            self._mstats.updated()
            return translation['id']
        except ClientError as e:
            self._mstats.error(
                f'Error while updating translation information: {str(e)}',
            )

    @staticmethod
    def _read_general_data(ws):
        return SimpleNamespace(**{
            field: ws.cell(settings.row_idx, 2).value for field, settings in GENERAL_SHEET_FIELDS.items()
        })

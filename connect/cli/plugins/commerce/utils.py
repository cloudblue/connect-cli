import os
import json
from datetime import datetime
import string
from urllib.parse import urlparse
import tempfile
from mimetypes import guess_type

from click import ClickException
from connect.cli.core.terminal import console
from connect.client import ClientError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import WHITE, Color

from connect.cli.core.utils import validate_output_options


def display_streams_table(
    query_billing_streams,
    query_pricing_streams,
    active_account_id,
):
    rows = []
    for origin, query in (
        ('Billing', query_billing_streams),
        ('Pricing', query_pricing_streams),
    ):
        for resource in query:
            rows.append(
                (
                    resource['id'],
                    origin,
                    resource['name'],
                    'Computed' if ('sources' in resource and resource['sources']) else 'Simple',
                    'Inbound' if resource['owner']['id'] != active_account_id else 'Outbound',
                    resource['status'].capitalize(),
                    resource['visibility'].capitalize(),
                ),
            )
    if rows:
        console.table(
            columns=[
                'ID',
                'Business scope',
                'Name',
                'Type',
                'Category',
                'Status',
                'Visibility',
            ],
            rows=rows,
        )
    else:
        console.secho(
            f'Results not found for the current account {active_account_id}.',
            fg='yellow',
        )


def guess_if_billing_or_pricing_stream(client, stream_id):
    if client.ns('billing').streams.filter(id=stream_id).count() > 0:
        return 'billing'
    elif client.ns('pricing').streams.filter(id=stream_id).count() > 0:
        return 'pricing'
    return None


def fill_general_information(ws, data, progress):
    ws.title = 'General Information'
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 180
    ws.merge_cells('A1:B1')
    cell = ws['A1']
    cell.fill = PatternFill('solid', start_color=Color('1565C0'))
    cell.font = Font(sz=24, color=WHITE)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value = 'Stream information'
    count = len(data)
    task = progress.add_task('Filling general information', total=count)
    for line, key in enumerate(data, 2):
        ws[f'A{line}'] = key
        ws[f'B{line}'] = data[key]
        progress.update(task, advance=1)


def _fill_headers(ws, columns):
    color = Color('d3d3d3')
    fill = PatternFill('solid', color)
    for n, value in enumerate(columns):
        letter = string.ascii_uppercase[n]
        cell = ws[f'{letter}1']
        cell.fill = fill
        cell.value = value


def fill_columns(ws, columns, progress):
    columns = list(columns)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    _fill_headers(
        ws,
        (
            'ID',
            'Name',
            'Description',
            'Type',
            'Position',
            'Required',
            'Output',
        ),
    )
    alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    if len(columns) > 0:
        task = progress.add_task('Filling columns', total=len(columns))
        for n, col in enumerate(columns, 2):
            ws.cell(n, 1, value=col['id']).alignment = alignment
            ws.cell(n, 2, value=col['name']).alignment = alignment
            ws.cell(n, 3, value=col.get('description', '')).alignment = alignment
            ws.cell(n, 4, value=col.get('type', '')).alignment = alignment
            ws.cell(n, 5, value=col['position']).alignment = alignment
            ws.cell(n, 6, value=col['required']).alignment = alignment
            ws.cell(n, 7, value=col['output']).alignment = alignment
            progress.update(task, advance=1)


def fill_transformations(ws, transformations, progress):
    try:
        transformations = list(transformations)
    except ClientError:
        return
    ws.column_dimensions['A'].width = 30
    for column in ('B', 'C', 'D', 'E', 'F', 'G'):
        ws.column_dimensions[column].width = 22
    ws.column_dimensions['H'].width = 40
    _fill_headers(
        ws,
        (
            'ID',
            'Function ID',
            'Function Name',
            'Description',
            'Overview',
            'Input Columns',
            'Output Columns',
            'Position',
            'Settings',
        ),
    )
    alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    if len(transformations) > 0:
        task = progress.add_task('Filling transformations', total=len(transformations))
        for n, transformation in enumerate(transformations, 2):
            ws.cell(n, 1, value=transformation['id']).alignment = alignment
            ws.cell(n, 2, value=transformation['function']['id']).alignment = alignment
            ws.cell(n, 3, value=transformation['function']['name']).alignment = alignment
            ws.cell(n, 4, value=transformation.get('description', '')).alignment = alignment
            ws.cell(n, 5, value=transformation['overview']).alignment = alignment
            col_alignment = Alignment(
                horizontal='left',
                vertical='top',
                wrap_text=True,
            )
            ws.cell(
                n,
                6,
                value='\n'.join([c['id'] for c in transformation['columns']['input']]),
            ).alignment = col_alignment
            ws.cell(
                n,
                7,
                value='\n'.join([c['id'] for c in transformation['columns']['output']]),
            ).alignment = col_alignment
            ws.cell(n, 8, value=transformation['position']).alignment = alignment
            ws.cell(
                n,
                9,
                value=json.dumps(transformation['settings'], indent=4, sort_keys=True),
            ).alignment = col_alignment
            ws.row_dimensions[n].height = 120
            progress.update(task, advance=1)


def _download_file(client, folder_type, folder_name, file_id, file_destination):
    try:
        response = (
            client.ns(
                'media',
            )
            .ns(
                'folders',
            )
            .collection(
                folder_type,
            )[folder_name]
            .collection(
                'files',
            )[file_id]
            .get()
        )
    except ClientError:
        raise ClickException(f'Error obtaining file {file_id} -> {file_destination}')
    with open(file_destination, 'wb') as f:
        f.write(response)


def fill_and_download_attachments(ws, attachments, client, base_path, progress):
    attachments = list(attachments)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    _fill_headers(
        ws,
        ('ID', 'Name'),
    )
    if len(attachments) > 0:
        task = progress.add_task('Filling and downloading attachments', total=len(attachments))
        for n, attachment in enumerate(attachments, 2):
            ws.cell(n, 1, value=attachment['id'])
            ws.cell(n, 2, value=attachment['name'])
            _download_file(
                client=client,
                folder_type=attachment['folder']['type'],
                folder_name=attachment['folder']['name'],
                file_id=attachment['id'],
                file_destination=os.path.join(base_path, attachment['name']),
            )
            progress.update(task, advance=1)


def export_stream(
    client,
    stream_id,
    active_account_id,
    output_file,
    output_path=None,
):
    output_file = validate_output_options(output_path, output_file, default_dir_name=stream_id)
    attachments_path = os.path.join(os.path.dirname(output_file), 'attachments')
    if not os.path.exists(attachments_path):
        os.mkdir(attachments_path)
    sample_input_path = os.path.join(os.path.dirname(output_file), 'sample', 'input')
    if not os.path.exists(sample_input_path):
        os.makedirs(sample_input_path)

    with console.status_progress() as (status, progress):
        collection = guess_if_billing_or_pricing_stream(client, stream_id)
        if not collection:
            console.secho(
                f'Stream {stream_id} not found for the current account {active_account_id}.',
                fg='red',
            )
            return

        response = (
            client.ns(collection)
            .streams.filter(id=stream_id)
            .select(
                'context',
                'samples',
                'sources',
            )
            .first()
        )

        wb = Workbook()

        stream_type = 'Computed' if 'sources' in response and response['sources'] else 'Simple'
        input = response['samples'].get('input', {})
        input_file_name = urlparse(input.get('name', '/')).path.split('/')[-1]
        status.update('Extracting general information', fg='blue')
        fill_general_information(
            ws=wb.active,
            data={
                'Stream ID': stream_id,
                'Stream Name': response['name'],
                'Stream Description': response.get('description'),
                'Stream Type': stream_type,
                'Stream Category': 'Inbound'
                if response['owner']['id'] != active_account_id
                else 'Outbound',
                'Computed Stream Source ID': response['sources'][0]['id']
                if stream_type == 'Computed'
                else '',
                'Computed Stream Source Name': response['sources'][0]['name']
                if stream_type == 'Computed'
                else '',
                'Computed Stream Source Type': response['sources'][0]['type']
                if stream_type == 'Computed'
                else '',
                'Product ID': response['context'].get('product', {}).get('id', ''),
                'Product Name': response['context'].get('product', {}).get('name', ''),
                'Partner ID': response['context'].get('account', {}).get('id', ''),
                'Partner Name': response['context'].get('account', {}).get('name', ''),
                'Marketplace ID': response['context'].get('marketplace', {}).get('id', ''),
                'Marketplace Name': response['context'].get('marketplace', {}).get('name', ''),
                'Pricelist ID': response['context'].get('pricelist', {}).get('id', ''),
                'Pricelist Name': response['context'].get('pricelist', {}).get('name', ''),
                'Listing ID': response['context'].get('listing', {}).get('id', ''),
                'Visibility': response['visibility'],
                'Input Data': input_file_name,
                'Export datetime': datetime.now().strftime('%Y-%m-%dT%H:%M:%M.%f'),
            },
            progress=progress,
        )

        if input:
            extracted_stream = input['name'].split('/')[-4]
            status.update('Downloading sample file', fg='blue')
            _download_file(
                client=client,
                folder_type='streams_samples',
                folder_name=extracted_stream,
                file_id=input.get('id', None),
                file_destination=os.path.join(sample_input_path, input_file_name),
            )

        status.update('Extracting columns', fg='blue')
        fill_columns(
            wb.create_sheet('Columns'),
            client.ns(collection).streams[stream_id].columns.all(),
            progress=progress,
        )

        status.update('Extracting transformations', fg='blue')
        fill_transformations(
            wb.create_sheet('Transformations'),
            client.ns(collection).streams[stream_id].transformations.all().select('columns'),
            progress=progress,
        )

        status.update('Extracting attachments', fg='blue')
        fill_and_download_attachments(
            wb.create_sheet('Attachments'),
            client.ns('media')
            .ns('folders')
            .collection('streams_attachments')[stream_id]
            .files.all(),
            client,
            base_path=attachments_path,
            progress=progress,
        )

        wb.save(output_file)

    console.echo('')

    console.secho(
        f'Stream {stream_id} exported properly to {output_file}.',
        fg='green',
    )


def get_destination_account(
    config,
    destination_account_id,
):
    if not destination_account_id or destination_account_id == config.active.id:
        return config.active
    if destination_account_id in config.accounts:
        return config.accounts[destination_account_id]
    else:
        raise ClickException(f'Error obtaining the destination account id {destination_account_id}')


def create_stream_from_origin(
    client,
    origin_stream,
    collection,
    stream_name=None,
    validate_context_objects=False,
):
    context = origin_stream['context']
    if validate_context_objects:
        for obj, ns in (
            ('product', 'products'),
            ('marketplace', 'marketplaces'),
            ('account', 'accounts'),
            ('pricelist', 'pricelists'),
        ):
            if obj in context:
                if client.collection(ns).filter(id=context[obj]['id']).count() == 0:
                    console.secho(
                        f'The {obj} {context[obj]["id"]} does not exists.',
                        fg='yellow',
                    )
                    del context[obj]
    body = {
        'name': (
            stream_name
            or f'Clone of {origin_stream["name"]} {origin_stream["id"]}) {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
        ),
        'description': origin_stream.get('description', ''),
        'context': context,
        'sources': origin_stream['sources'],
        'status': 'configuring',
        'type': collection,
        'visibility': 'private',
    }
    try:
        destination_stream = client.ns(collection).streams.create(
            json=body,
        )
    except ClientError as e:
        raise ClickException(e.errors[0])
    return destination_stream['id']


def _upload_file(
    client,
    folder,
    stream_id,
    file,
):
    with open(file, 'rb') as f:
        name = f.name
        data = {
            'file': (name, f, guess_type(name)),
        }
        return (
            client.ns(
                'media',
            )
            .ns(
                'folders',
            )
            .collection(
                folder,
            )[stream_id]
            .files.create(
                files=data,
            )
        )


def upload_sample(
    client,
    collection,
    stream_id,
    file,
):
    response = _upload_file(client, 'streams_samples', stream_id, file)
    body = {'samples': {'input': {'id': response['id']}}}
    client.ns(collection).streams[stream_id].update(
        json=body,
    )
    return response['id']


def clone_sample(
    origin_client,
    destination_client,
    destination_stream_id,
    collection,
    origin_stream_input_sample,
    progress,
):
    task_sample = progress.add_task('Processing input sample', total=1)
    with tempfile.TemporaryDirectory() as tmpdir:
        sample_path = os.path.join(tmpdir, 'sample')
        os.makedirs(sample_path)
        name = urlparse(origin_stream_input_sample['name']).path.split('/')[-1]
        stream = urlparse(origin_stream_input_sample['name']).path.split('/')[-4]
        destination = os.path.join(sample_path, name)
        _download_file(
            origin_client,
            'streams_samples',
            stream,
            origin_stream_input_sample['id'],
            destination,
        )
        upload_sample(
            destination_client,
            collection,
            destination_stream_id,
            destination,
        )
    progress.update(task_sample, completed=1)


def upload_attachment(
    client,
    stream_id,
    attachment,
):
    return _upload_file(client, 'streams_attachments', stream_id, attachment)


def clone_attachments(
    origin_client,
    destination_client,
    attachments,
    stream_id,
    destination_stream_id,
    progress,
):
    file_mapping = {}
    task_attachment = progress.add_task('Processing attachments', total=len(attachments))
    with tempfile.TemporaryDirectory() as tmpdir:
        for attachment in attachments:
            destination = os.path.join(
                tmpdir,
                attachment['name'],
            )
            _download_file(
                client=origin_client,
                folder_type=attachment['folder']['type'],
                folder_name=attachment['folder']['name'],
                file_id=attachment['id'],
                file_destination=destination,
            )
            response = upload_attachment(
                destination_client,
                destination_stream_id,
                destination,
            )
            file_mapping[
                f'/public/v1/media/folders/streams_attachments/{stream_id}/files/{attachment["id"]}'
            ] = f'/public/v1/media/folders/streams_attachments/{destination_stream_id}/files/{response["id"]}'
            progress.update(task_attachment, advance=1)
    return file_mapping


def _sort_list_by_id(list_to_sort):
    return sorted(list_to_sort, key=lambda c: c['id'])


def generate_column_mapping(client, collection, stream_id):
    mapping_by_name = {}
    columns_by_id = {}
    columns = _sort_list_by_id(list(client.ns(collection).streams[stream_id].columns.all()))
    for col in columns:
        if col['name'] in mapping_by_name:
            mapping_by_name[col['name']].append(col['id'])
        else:
            mapping_by_name[col['name']] = [col['id']]
        columns_by_id[col['id']] = col
    return mapping_by_name, columns_by_id


def clone_transformations(
    destination_client,
    destination_stream_id,
    collection,
    file_mapping,
    transformations,
    o_mapping_by_name,
    progress,
):
    transformation_task = progress.add_task(
        'Processing transformations', total=len(transformations)
    )
    for transformation in transformations:
        d_mapping_by_name, columns_by_id = generate_column_mapping(
            destination_client,
            collection,
            destination_stream_id,
        )
        input_columns = []
        for col in transformation['columns']['input']:
            column_ids = d_mapping_by_name[col['name']]
            position = 0
            if len(column_ids) > 1:
                for id in o_mapping_by_name[col['name']]:
                    if id == col['id']:
                        break
                    position += 1
            input_column_id = column_ids[position]
            input_columns.append(columns_by_id[input_column_id])
        payload = {
            'description': transformation.get('description', ''),
            'function': {'id': transformation['function']['id']},
            'settings': transformation['settings'],
            'overview': transformation['overview'],
            'position': transformation['position'],
            'columns': {
                'input': input_columns,
                'output': transformation['columns']['output'],
            },
        }

        if transformation['function']['name'] == 'Lookup Data from a stream attached Excel':
            if 'file' in payload['settings']:
                payload['settings']['file'] = file_mapping[payload['settings']['file']]
        destination_client.ns(
            collection,
        ).streams[destination_stream_id].transformations.create(
            payload=payload,
        )
        progress.update(transformation_task, advance=1)


def align_column_output(
    collection, origin_client, origin_stream_id, destination_client, destination_stream_id, progress
):
    origin_columns = _sort_list_by_id(
        list(origin_client.ns(collection).streams[origin_stream_id].columns.all())
    )
    columns_task = progress.add_task('Processing columns', total=len(origin_columns))
    dest_columns = _sort_list_by_id(
        list(destination_client.ns(collection).streams[destination_stream_id].columns.all())
    )
    updated = 0
    for n in range(len(origin_columns)):
        if origin_columns[n]['output'] is False and dest_columns[n]['output'] is True:
            destination_client.ns(collection).streams[destination_stream_id].columns[
                dest_columns[n]['id']
            ].update(
                payload={'output': False},
            )
            updated += 1
        progress.update(columns_task, advance=1)
    return len(origin_columns), updated


def print_results(results):
    COLUMNS = (
        'Module',
        ('right', 'Processed'),
        ('right', 'Created'),
        ('right', 'Updated'),
        ('right', 'Deleted'),
        ('right', 'Skipped'),
        ('right', 'Errors'),
    )
    console.table(
        columns=COLUMNS,
        rows=results,
        expand=True,
    )


def print_errors(errors):  # pragma: no cover
    if errors:
        console.confirm(
            'Are you sure you want to display errors?',
            abort=True,
        )
        console.echo('')
        for error in errors:
            console.secho(
                error,
                fg='red',
            )


def clone_stream(
    origin_account,
    stream_id,
    destination_account,
    stream_name=None,
    validate=False,
):
    results = []

    with console.progress() as progress:
        collection = guess_if_billing_or_pricing_stream(origin_account.client, stream_id)
        if not collection:
            raise ClickException(
                f'Stream {stream_id} not found for the current account {origin_account.id}.'
            )

        origin_stream = (
            origin_account.client.ns(collection)
            .streams.filter(id=stream_id)
            .select(
                'context',
                'samples',
                'sources',
            )
            .first()
        )

        stream_type = (
            'Computed' if 'sources' in origin_stream and origin_stream['sources'] else 'Simple'
        )

        if stream_type == 'Computed' and origin_account != destination_account:
            raise ClickException('You cannot clone a Computed stream between different accounts.')

        category = 'Inbound' if origin_stream['owner']['id'] != origin_account.id else 'Outbound'

        if category == 'Inbound':
            raise ClickException('Inbound streams cannot be cloned.')

        destination_stream_id = create_stream_from_origin(
            destination_account.client,
            origin_stream,
            collection,
            stream_name,
            origin_account != destination_account,
        )

        file_mapping = {}
        if (
            stream_type == 'Simple'
            and 'samples' in origin_stream
            and 'input' in origin_stream['samples']
        ):
            clone_sample(
                origin_account.client,
                destination_account.client,
                destination_stream_id,
                collection,
                origin_stream['samples']['input'],
                progress,
            )
            results.append(('Sample input', 1, 1, 0, 0, 0, 0))

        attachments = list(
            origin_account.client.ns('media')
            .ns('folders')
            .collection('streams_attachments')[stream_id]
            .files.all()
        )
        if attachments:
            file_mapping = clone_attachments(
                origin_account.client,
                destination_account.client,
                attachments,
                stream_id,
                destination_stream_id,
                progress,
            )
            results.append(('Attachments', len(attachments), len(attachments), 0, 0, 0, 0))

        transformations = list(
            origin_account.client.ns(
                collection,
            )
            .streams[stream_id]
            .transformations.all()
            .select(
                'columns',
            )
        )

        if transformations:
            o_mapping_by_name, _ = generate_column_mapping(
                origin_account.client, collection, stream_id
            )
            clone_transformations(
                destination_account.client,
                destination_stream_id,
                collection,
                file_mapping,
                transformations,
                o_mapping_by_name,
                progress,
            )
            results.append(
                ('Transformations', len(transformations), len(transformations), 0, 0, 0, 0)
            )

        processed, updated = align_column_output(
            collection,
            origin_account.client,
            stream_id,
            destination_account.client,
            destination_stream_id,
            progress,
        )
        results.append(('Columns', processed, 0, updated, 0, 0, 0))

    if validate:
        console.secho('')
        destination_account.client.ns(collection).streams[destination_stream_id]('validate').post()
        console.secho(
            f'Stream {destination_stream_id} validation executed properly.',
            fg='green',
        )

    return destination_stream_id, results


def validate_sheet_names(wb):
    for sheet in (
        'General Information',
        'Columns',
        'Transformations',
        'Attachments',
    ):
        if sheet not in wb.sheetnames:
            raise ClickException(
                'The file must contain `General Information`, `Columns`, `Transformations` and '
                '`Attachments` sheets.'
            )


def _validate_header(
    current_headers,
    expected_headers,
    sheet_name,
):
    for header in expected_headers:
        if header not in current_headers:
            raise ClickException(
                f'The {sheet_name} sheet header does not contain `{header}` header.'
            )


def validate_headers(wb):
    _validate_header(
        [c.value for c in wb['General Information'][1]],
        (
            'Stream information',
            None,
        ),
        'General Information',
    )
    _validate_header(
        [c.value for c in wb['Columns'][1]],
        (
            'ID',
            'Name',
            'Description',
            'Type',
            'Position',
            'Required',
            'Output',
        ),
        'Columns',
    )
    _validate_header(
        [c.value for c in wb['Transformations'][1]],
        (
            'ID',
            'Function ID',
            'Function Name',
            'Description',
            'Overview',
            'Input Columns',
            'Output Columns',
            'Position',
            'Settings',
        ),
        'Transformations',
    )
    _validate_header(
        [c.value for c in wb['Attachments'][1]],
        (
            'ID',
            'Name',
        ),
        'Attachments',
    )


def get_work_book(input_file):
    if not os.path.exists(input_file):
        raise ClickException(f'The file {input_file} does not exists.')
    if 'xlsx' not in input_file:
        raise ClickException(f'The file {input_file} has invalid format, must be xlsx.')
    wb = load_workbook(input_file, read_only=True)

    validate_sheet_names(wb)

    validate_headers(wb)

    return wb


def update_general_information(
    client,
    collection,
    stream_id,
    sheet,
    results,
    errors,
    progress,
):
    task = progress.add_task('Updating general information', total=5)
    stream = (
        client.ns(collection)
        .streams.filter(id=stream_id)
        .select(
            'context',
            'samples',
            'sources',
            'validation',
        )
        .first()
    )

    if stream['status'] == 'active' or stream.get('validation', {}).get('status') == 'processing':
        raise ClickException(
            f'Stream {stream_id} cannot be updated because it is in "active" status '
            f'or validation is processing.',
        )

    body = {'context': {}}
    updated = 0
    errors_on_update = 0
    for n in range(2, sheet.max_row + 1):
        h, v = sheet[n]
        if h.value == 'Stream Name' and stream['name'] != v.value:
            body['name'] = v.value
            updated += 1
        elif h.value == 'Stream Description' and stream.get('description') != v.value:
            body['description'] = v.value
            updated += 1
        elif (
            h.value == 'Product ID'
            and stream.get('context', {}).get('product', {}).get('id', None) != v.value
        ):
            body['context']['product'] = {'id': v.value}
            updated += 1
        elif (
            h.value == 'Partner ID'
            and stream.get('context', {}).get('account', {}).get('id', None) != v.value
        ):
            body['context']['account'] = {'id': v.value}
            updated += 1
        elif (
            h.value == 'Marketplace ID'
            and stream.get('context', {}).get('marketplace', {}).get('id', None) != v.value
        ):
            body['context']['marketplace'] = {'id': v.value}
            updated += 1

    if updated:
        if not body['context']:
            del body['context']
        try:
            client.ns(collection).streams[stream_id].update(
                json=body,
            )
        except ClientError as e:
            errors.append(str(e))
            updated = 0
            errors_on_update = 5

    results.append(('General information', 5, 0, updated, 0, 0, errors_on_update))
    progress.update(task, advance=5)


def update_transformations(
    client,
    collection,
    stream_id,
    sheet,
    results,
    errors,
    progress,
):
    task = progress.add_task('Updating transformation information', total=sheet.max_row - 1)

    updated = 0
    deleted = 0
    ids = []
    for n in range(2, sheet.max_row + 1):
        id, fid, fname, descr, over, input, output, position, settings = sheet[n]
        ids.append(id.value)
        origin_trf = None
        try:
            origin_trf = client.ns(collection).streams[stream_id].transformations[id.value].get()
        except ClientError:
            errors.append(
                f'The transformation {id.value} cannot be updated because it does not exist.'
            )
            progress.update(task, advance=1)
            continue

        try:
            to_update = {}
            if origin_trf['settings'] != json.loads(settings.value):
                to_update['settings'] = json.loads(settings.value)
            if origin_trf.get('description') != descr.value:
                to_update['description'] = descr.value
            if int(origin_trf['position']) != position.value:
                to_update['position'] = int(position.value)

            if to_update:
                client.ns(collection).streams[stream_id].transformations[id.value].update(
                    json=to_update,
                )
                updated += 1
            progress.update(task, advance=1)
        except ClientError:
            errors.append(f'Error updating the transformation {id.value} with data {to_update}.')

    try:
        current_ids = [
            t['id'] for t in list(client.ns(collection).streams[stream_id].transformations.all())
        ]
        for cid in current_ids:
            if cid not in ids:
                client.ns(collection).streams[stream_id].transformations[cid].delete()
                deleted += 1
    except ClientError as e:
        errors.append(f'Error deleting the transformation {e}.')

    results.append(('Transformations', sheet.max_row - 1, 0, updated, deleted, 0, len(errors)))


def update_attachments(
    client,
    stream_id,
    sheet,
    results,
    errors,
    progress,
):
    task = progress.add_task('Updating attachment files', total=sheet.max_row - 1)

    created = 0
    deleted = 0
    errored = 0
    skipped = 0
    ids = []
    for n in range(2, sheet.max_row + 1):
        id, file_name = sheet[n]
        try:
            if (
                client.ns('media')
                .ns('folders')
                .collection('streams_attachments')[stream_id]
                .collection('files')
                .filter(id=id.value)
                .count()
                == 0
            ):
                response = upload_attachment(
                    client,
                    stream_id,
                    os.path.join(stream_id, 'attachments', file_name.value),
                )
                ids.append(response['id'])
                created += 1
            else:
                ids.append(id.value)
                skipped += 1

        except ClientError as e:
            errors.append(str(e))
            errored += 1

    for attachment in list(
        client.ns('media').ns('folders').collection('streams_attachments')[stream_id].files.all()
    ):
        try:
            if attachment['id'] not in ids:
                client.ns('media').ns('folders').collection('streams_attachments')[stream_id].files[
                    attachment['id']
                ].delete()
                deleted += 1
        except ClientError as e:
            errors.append(str(e))
            errored += 1
    progress.update(task, advance=sheet.max_row - 1)
    results.append(('Attachment files', sheet.max_row - 1, created, 0, deleted, skipped, errored))


def sync_stream(
    account,
    stream_id,
    input_file,
):
    wb = get_work_book(input_file)

    collection = guess_if_billing_or_pricing_stream(account.client, stream_id)
    if not collection:
        raise ClickException(f'Stream {stream_id} not found for the current account {account.id}.')

    results = []
    errors = []
    with console.status_progress() as (status, progress):
        status.update('Updating general information', fg='blue')
        update_general_information(
            client=account.client,
            collection=collection,
            stream_id=stream_id,
            sheet=wb['General Information'],
            results=results,
            errors=errors,
            progress=progress,
        )

        status.update('Updating transformations', fg='blue')
        update_transformations(
            client=account.client,
            collection=collection,
            stream_id=stream_id,
            sheet=wb['Transformations'],
            results=results,
            errors=errors,
            progress=progress,
        )

        status.update('Updating attachments', fg='blue')
        if wb['Attachments'].max_row > 1:
            update_attachments(
                client=account.client,
                stream_id=stream_id,
                sheet=wb['Attachments'],
                results=results,
                errors=errors,
                progress=progress,
            )

    return results, errors

from connect.client import ClientError
from iso3166 import countries
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.worksheet.datavalidation import DataValidation

from connect.cli.core.http import handle_http_error
from connect.cli.core.terminal import console
from connect.cli.core.utils import validate_output_options
from connect.cli.plugins.customer.constants import COL_HEADERS


def dump_customers(client, account_id, output_file, output_path=None):  # noqa: CCR001
    output_file = validate_output_options(
        output_path, output_file, default_dir_name=account_id, default_file_name='customers',
    )
    try:
        wb = Workbook()
        _prepare_worksheet(wb.create_sheet('Customers'))
        _add_countries(wb.create_sheet('Countries'))

        customers = client.ns('tier').accounts.all()
        row_idx = 2
        count = customers.count()
        with console.progress() as progress:
            task = progress.add_task('Processing customer', total=count)
            for customer in customers:
                progress.update(task, description=f'Processing customer {customer["id"]}', advance=1)
                _fill_customer_row(wb['Customers'], row_idx, customer)
                row_idx += 1
            progress.update(task, completed=count)
    except ClientError as error:
        handle_http_error(error)

    default_sheet = wb['Sheet']
    wb.remove(default_sheet)
    wb.save(output_file)

    return output_file


def _fill_customer_row(ws, row_idx, customer):
    action_validation = DataValidation(
        type='list',
        formula1='"-,create,update"',
        allow_blank=False,
    )
    action_validation.error = str('Action must be from list')
    action_validation.errorTitle = str('Invalid action')
    action_validation.prompt = str('Please choose action from list')
    action_validation.promptTitle = str('List of choices')
    search_criteria_validation = DataValidation(
        type='list',
        formula1='"-,id,external_id,external_uid"',
        allow_blank=False,
    )
    search_criteria_validation.error = str('Search criteria must be one from list')
    search_criteria_validation.errorTitle = str('Invalid search criteria')
    search_criteria_validation.prompt = str('Please choose search criteria from list')
    search_criteria_validation.promptTitle = str('List of choices')

    ws.add_data_validation(action_validation)
    ws.add_data_validation(search_criteria_validation)

    ws.cell(row_idx, 1, value=customer.get('id', '-'))
    ws.cell(row_idx, 2, value=customer.get('external_id', '-'))
    ws.cell(row_idx, 3, value=customer.get('external_uid', '-'))
    ws.cell(row_idx, 4, value='-')
    ws.cell(row_idx, 5, value=customer['hub'].get('id', '-') if 'hub' in customer else '-')
    ws.cell(row_idx, 6, value='id' if 'parent' in customer else '-')
    ws.cell(row_idx, 7, value=customer['parent'].get('id', '-') if 'parent' in customer else '-')
    ws.cell(row_idx, 8, value=customer.get('type', '-'))
    ws.cell(row_idx, 9, value=customer.get('tax_id', '-'))
    ws.cell(row_idx, 10, value=customer.get('name', '-'))
    ws.cell(row_idx, 11, value=customer['contact_info'].get('address_line1', '-'))
    ws.cell(row_idx, 12, value=customer['contact_info'].get('address_line2', '-'))
    ws.cell(row_idx, 13, value=customer['contact_info'].get('city', '-'))
    ws.cell(row_idx, 14, value=customer['contact_info'].get('state', '-'))
    ws.cell(row_idx, 15, value=customer['contact_info'].get('zip', '-'))
    ws.cell(row_idx, 16, value=customer['contact_info'].get('country', '-'))
    ws.cell(row_idx, 17, value=customer['contact_info']['contact'].get('first_name', '-'))
    ws.cell(row_idx, 18, value=customer['contact_info']['contact'].get('last_name', '-'))
    ws.cell(row_idx, 19, value=customer['contact_info']['contact'].get('email', '-'))
    ws.cell(
        row_idx, 20, value=_get_phone_number(
            customer['contact_info']['contact'].get(
                'phone_number',
                '-',
            ),
        ),
    )

    action_validation.add(f'D{row_idx}')
    search_criteria_validation.add(f'F{row_idx}')


def _get_phone_number(number):
    if number == '-':
        return number
    output_number = ''
    if 'country_code' in number:
        output_number = number['country_code']
    if 'area_code' in number:
        output_number += number['area_code']
    if 'phone_number' in number:
        output_number += number['phone_number']
    if 'extension' in number:
        output_number += number['extension']
    return output_number


def _prepare_worksheet(ws):
    color = Color('d3d3d3')
    fill = PatternFill('solid', color)
    cels = ws['A1': 'T1']
    for cel in cels[0]:
        if cel.column_letter in ['J', 'K', 'L']:
            ws.column_dimensions[cel.column_letter].width = 50
        elif cel.column_letter in ['B', 'D', 'E', 'F']:
            ws.column_dimensions[cel.column_letter].width = 15
        else:
            ws.column_dimensions[cel.column_letter].width = 20
        ws.column_dimensions[cel.column_letter].auto_size = True
        cel.fill = fill
        cel.value = COL_HEADERS[cel.column_letter]


def _add_countries(ws):
    col_headers = {
        'A': '2 letters country code',
        'B': 'Country name',
    }
    color = Color('d3d3d3')
    fill = PatternFill('solid', color)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['A'].auto_size = True
    ws.column_dimensions['B'].width = 50
    cels = ws['A1': 'B1']
    for cel in cels[0]:
        cel.fill = fill
        cel.value = col_headers[cel.column_letter]
    row_idx = 2
    for country in countries:
        ws.cell(row_idx, 1, value=country.alpha2)
        ws.cell(row_idx, 2, value=country.name)
        row_idx += 1
    ws.cell(row_idx, 1, value='-')
    ws.cell(row_idx, 2, value='Not Selected')

import uuid
from collections import namedtuple

import phonenumbers
from click import ClickException

from connect.client import ClientError, R

from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from connect.cli.core.terminal import console
from connect.cli.plugins.shared.sync_stats import SynchronizerStatsSingleModule
from connect.cli.plugins.shared.exceptions import SheetNotFoundError
from connect.cli.plugins.customer.constants import COL_HEADERS

fields = (v.replace(' ', '_').lower() for v in COL_HEADERS.values())

_RowData = namedtuple('RowData', fields)


class CustomerSynchronizer:
    def __init__(self, client, account_id):
        self._client = client
        self._wb = None
        self.account_id = account_id
        self.hubs = ['HB-0000-0000']
        self.stats = SynchronizerStatsSingleModule('Customers')

    def populate_hubs(self):
        if self.account_id.startswith('PA-'):
            hubs = self._client.hubs.all()
            for hub in hubs:
                if hub['instance']['type'] != 'OA':
                    self.hubs.append(hub['id'])

    def open(self, input_file, worksheet):
        self._open_workbook(input_file)
        if worksheet not in self._wb.sheetnames:
            raise SheetNotFoundError(f'File does not contain {worksheet} to synchronize, skipping')
        ws = self._wb[worksheet]
        self._validate_worksheet_sheet(ws)

    def save(self, output_file):
        self._wb.save(output_file)

    def _open_workbook(self, input_file):
        try:
            self._wb = load_workbook(
                input_file,
                data_only=True,
            )
        except InvalidFileException as ife:
            raise ClickException(str(ife))
        except BadZipFile:
            raise ClickException(f'{input_file} is not a valid xlsx file.')

    @staticmethod
    def _validate_worksheet_sheet(ws):
        cels = ws['A1': 'T1']
        for cel in cels[0]:
            if cel.value != COL_HEADERS[cel.column_letter]:
                raise ClickException(
                    f'Column `{cel.coordinate}` must be `{COL_HEADERS[cel.column_letter]}` '
                    f'and is `{cel.value}`.',
                )

    def sync(self):  # noqa: CCR001
        ws = self._wb['Customers']
        self.stats.reset()
        parent_uuid = {}
        parent_external_id = {}
        parent_id = []

        self.populate_hubs()
        with console.progress() as progress:
            task = progress.add_task('Processing item', total=ws.max_row - 1)
            for row_idx in range(2, ws.max_row + 1):
                data = _RowData(*[ws.cell(row_idx, col_idx).value for col_idx in range(1, 21)])
                progress.update(
                    task,
                    description=f'Processing item {data.id or data.external_id or data.external_uid}',
                    advance=1,
                )
                if data.action == '-':
                    self.stats.skipped()
                    continue
                row_errors = self._validate_row(data)
                if row_errors:
                    self.stats.error(row_errors, row_idx)
                    continue
                if data.parent_search_criteria and not data.parent_search_value:
                    self.stats.error("Parent search value is needed if criteria is set", row_idx)
                    continue
                if data.hub_id and (data.hub_id != '' or data.hub_id != '-'):
                    if data.hub_id not in self.hubs:
                        self.stats.error(f"Accounts on hub {data.hub_id} can not be modified", row_idx)
                        continue
                name = f'{data.technical_contact_first_name} {data.technical_contact_last_name}'
                model = {
                    "type": data.type,
                    "name": data.company_name if data.company_name else name,
                    "contact_info": {
                        "address_line1": data.address_line_1,
                        "address_line2": data.address_line_2,
                        "city": data.city,
                        "country": data.country,
                        "postal_code": data.zip,
                        "state": data.state,
                        "contact": {
                            "first_name": data.technical_contact_first_name,
                            "last_name": data.technical_contact_last_name,
                            "email": data.technical_contact_email,
                        },
                    },
                }
                if data.external_id:
                    model['external_id'] = data.external_id
                if data.external_uid:
                    model['external_uid'] = data.external_uid
                else:
                    model['external_uid'] = str(uuid.uuid4())
                if data.technical_contact_phone:
                    try:
                        phone = phonenumbers.parse(data.technical_contact_phone, data.country)
                        phone_number = {
                            "country_code": f'+{str(phone.country_code)}',
                            "area_code": '',
                            "extension": str(phone.extension) if phone.extension else '-',
                            'phone_number': str(phone.national_number),
                        }
                        model['contact_info']['contact']['phone_number'] = phone_number
                    except Exception:
                        pass
                if data.parent_search_criteria != '-':
                    model['parent'] = {}
                    if data.parent_search_criteria == 'id':
                        if data.parent_search_value not in parent_id:
                            try:
                                pacc = self._client.ns('tier').accounts[data.parent_search_value].get()
                                parent_id.append(pacc['id'])
                            except ClientError:
                                self.stats.error(
                                    f'Parent with id {data.parent_search_value} does not exist',
                                    row_idx,
                                )
                                continue
                        model['parent']['id'] = data.parent_search_value
                    elif data.parent_search_criteria == 'external_id':
                        if data.parent_search_value not in parent_external_id:
                            try:
                                r = R().external_id.eq(str(data.parent_search_value))
                                pacc = self._client.ns('tier').accounts.filter(r).all()
                                pacc_count = pacc.count()
                                if pacc_count == 0:
                                    self.stats.error(
                                        f'Parent with external_id {data.parent_search_value} not found',
                                        row_idx,
                                    )
                                    continue
                                elif pacc_count > 1:
                                    self.stats.error(
                                        f'More than one Parent with external_id {data.parent_search_value}',
                                        row_idx,
                                    )
                                    continue
                                parent_external_id[data.parent_search_value] = pacc[0]['id']
                            except ClientError:
                                self.stats.error(
                                    'Error when obtaining parent data from Connect', row_idx,
                                )
                                continue
                        model['parent']['id'] = parent_external_id[data.parent_search_value]
                    else:
                        if data.parent_search_value not in parent_uuid:
                            try:
                                r = R().external_uid.eq(str(data.parent_search_value))
                                pacc = self._client.ns('tier').accounts.filter(r).all()
                                if pacc.count() == 0:
                                    self.stats.error(
                                        f'Parent with external_uid {data.parent_search_value} not found',
                                        row_idx,
                                    )
                                    continue
                                elif pacc.count() > 1:
                                    self.stats.error(
                                        f'More than one Parent with external_uid {data.parent_search_value}',
                                        row_idx,
                                    )
                                    continue
                                parent_uuid[data.parent_search_value] = pacc[0]['id']
                            except ClientError:
                                self.stats.error(
                                    'Error when obtaining parent data from Connect', row_idx,
                                )
                                continue
                        model['parent']['id'] = parent_uuid[data.parent_search_value]
                if data.action == 'create':
                    try:
                        account = self._client.ns('tier').accounts.create(model)
                    except ClientError as e:
                        self.stats.error(
                            f'Error when creating account: {str(e)}', row_idx,
                        )
                        continue
                    self.stats.created()
                    self._update_sheet_row(ws, row_idx, account)
                else:
                    try:
                        model['id'] = data.id
                        account = self._client.ns('tier').accounts[data.id].update(model)
                    except ClientError as e:
                        self.stats.error(
                            f'Error when updating account: {str(e)}', row_idx,
                        )
                        continue
                    self.stats.updated()
                    self._update_sheet_row(ws, row_idx, account)
            progress.update(task, completed=ws.max_row - 1)

    @staticmethod
    def _update_sheet_row(ws, row_idx, account):
        ws.cell(row_idx, 1, value=account['id'])
        ws.cell(row_idx, 3, value=account['external_uid'])
        ws.cell(row_idx, 4, value='-')

    def _validate_row(self, row):
        errors = []
        if row.action not in ('-', 'create', 'update'):
            errors.append(f'Action {row.action} is not supported')
            return errors
        if row.action == '-':
            return
        if row.type not in ('customer', 'reseller'):
            errors.append(f'Customer type must be customer or reseller, not {row.type}')
            return errors
        if not all(
                [
                    row.address_line_1,
                    row.city,
                    row.state,
                    row.zip,
                    row.technical_contact_first_name,
                    row.technical_contact_last_name,
                    row.technical_contact_email,
                ],
        ):
            errors.append('Address line 1, city, state and zip are mandatory')
            return errors
        if row.action == 'create' and row.id is not None:
            errors.append(f'Create action must not have account id, is set to {row.id}')
            return errors
        if row.action == 'create' and row.type == 'customer' and (
                row.parent_search_criteria == '' or row.parent_search_criteria == '-'
        ):
            errors.append('Customers requires a parent account')
            return errors
        if row.action == 'update' and not row.id.startswith('TA-'):
            errors.append('Update operation requires account ID to be set')
            return errors
        if row.action == 'update':
            try:
                self._client.ns('tier').accounts[row.id].get()
            except ClientError:
                errors.append(
                    f'Account with id {row.id} does not exist',
                )
                return errors

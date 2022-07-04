from copy import copy
from io import BytesIO

from click import ClickException
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.utils import get_column_letter

from connect.client import ClientError
from connect.cli.core.http import format_http_status, handle_http_error
from connect.cli.plugins.translation.utils import insert_column_ws

EXCEL_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def get_translation_workbook(client, translation_id):
    try:
        xls_data = client.ns('localization').translations[translation_id].action('attributes').get(
            headers={'Accept': EXCEL_CONTENT_TYPE},
        )
        return load_workbook(filename=BytesIO(xls_data))
    except ClientError as error:
        if error.status_code == 404:
            status = format_http_status(error.status_code)
            raise ClickException(f'{status}: Translation {translation_id} not found.')
        handle_http_error(error)


def alter_attributes_sheet(ws):
    # Search for the 'value' column
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(1, col_idx).value.lower() == 'value':
            new_column = col_idx
            break
    else:
        return

    # Add new 'action' column before 'value' column
    fill = copy(ws.cell(1, 1).fill)
    font = copy(ws.cell(1, 1).font)
    insert_column_ws(ws, new_column, 20)
    header_cell = ws.cell(1, new_column)
    header_cell.value = 'action'
    header_cell.fill = fill
    header_cell.font = font

    # Setup 'action' column cells
    if ws.max_row > 1:
        action_validation = DataValidation(type='list', formula1='"-,update"', allow_blank=False)
        ws.add_data_validation(action_validation)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, new_column)
            cell.alignment = Alignment(vertical='top')
            cell.value = '-'
            action_validation.add(cell)

    # (re)set auto filter
    ws.auto_filter = AutoFilter(ref=f'A:{get_column_letter(ws.max_column)}')
    for col_idx in range(1, ws.max_column + 1):
        ws.auto_filter.add_filter_column(col_idx, [], blank=False)

# -*- coding: utf-8 -*-

# This file is part of the Ingram Micro Cloud Blue Connect connect-cli.
# Copyright (c) 2019-2022 Ingram Micro. All Rights Reserved.
import math
from collections import namedtuple

import click

from connect.client import ClientError

from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from connect.cli.plugins.shared.sync_stats import SynchronizerStats
from connect.cli.plugins.shared.exceptions import SheetNotFoundError
from connect.cli.plugins.shared.constants import ATTRIBUTES_SHEET_COLUMNS


class TranslationAttributesSynchronizer:
    """
    Synchronize the attributes of a translation from excel file.
    """
    _MAX_BATCH_SIZE = 10

    def __init__(self, client, progress, stats=None):
        self._client = client
        self._progress = progress
        self._wb = None
        self._ws = None
        if stats is None:
            stats = SynchronizerStats()
        self._mstats = stats['Translations Attributes']

    @property
    def max_batch_size(self):
        return self._MAX_BATCH_SIZE

    def open(self, input_file, worksheet):
        self._open_workbook(input_file)
        if worksheet not in self._wb.sheetnames:
            raise SheetNotFoundError(f"File does not contain worksheet '{worksheet}' to synchronize, skipping")
        self._ws = self._wb[worksheet]
        self._validate_attributes_worksheet(self._ws)

    def save(self, output_file):
        self._wb.save(output_file)

    def sync(self, translation, is_clone=False):
        translation_id = self._get_translation_id(translation)
        attributes = self._collect_attributes_to_update(self._ws, translation, is_clone)
        if attributes:
            self._update_attributes(translation_id, attributes, self._ws)

    def _open_workbook(self, input_file):
        try:
            self._wb = load_workbook(input_file, data_only=True)
        except InvalidFileException as ife:
            raise click.ClickException(str(ife))
        except BadZipFile:
            raise click.ClickException(f'{input_file} is not a valid xlsx file.')

    @staticmethod
    def _validate_attributes_worksheet(ws):
        for col_idx, header in enumerate(ATTRIBUTES_SHEET_COLUMNS, 1):
            if header == 'original value':
                continue
            cell = ws.cell(1, col_idx)
            if cell.value != header:
                raise click.ClickException(
                    f"Column '{cell.coordinate}' must be '{header}', but it is '{cell.value}'",
                )

    def _collect_attributes_to_update(self, ws, translation, is_clone):
        AttributeRow = namedtuple(
            'AttributeRow',
            (header.replace(' ', '_').lower() for header in ATTRIBUTES_SHEET_COLUMNS),
        )
        task = self._progress.add_task('Process attribute', total=ws.max_row - 1)
        new_attrs = None
        if is_clone and translation is not None:
            translation_res = self._client.ns('localization').translations[translation['id']]
            new_attrs = translation_res.attributes.all()
        attributes = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            row = AttributeRow(*row)
            self._progress.update(
                task,
                description=f'Process attribute {row.key}',
                advance=1,
            )
            if row.action == 'update':
                attributes[row_idx] = {'key': row.key, 'value': row.value, 'comment': row.comment}
                if new_attrs:
                    new_attrs_idx = row_idx - 2
                    if (
                        self._is_equal_attribute(new_attrs_idx, new_attrs, row)
                        or translation['auto']['enabled']
                    ):
                        attributes.pop(row_idx)
                        self._mstats.skipped()
                        continue
                    attributes[row_idx]['key'] = new_attrs[new_attrs_idx]['key']

            else:
                self._mstats.skipped()
        self._progress.update(task, completed=ws.max_row - 1)

        return attributes

    def _update_attributes(self, translation_id, attributes, ws):
        max_batch_size = self.max_batch_size
        try:
            translation_res = self._client.ns('localization').translations[translation_id]
            attr_value_list = list(attributes.values())
            chunk = 0
            # bulk update only support 10 items at a time
            for _ in range((math.ceil(len(attributes) / max_batch_size))):
                translation_res.attributes.bulk_update(attr_value_list[chunk:chunk + max_batch_size])
                chunk += max_batch_size
            self._mstats.updated(len(attributes))
            for row_idx in attributes.keys():
                self._update_attributes_sheet_row(ws, row_idx)
        except ClientError as e:
            self._mstats.error(
                f'Error while updating attributes: {str(e)}',
                range(1, len(attributes) + 1),
            )

    @staticmethod
    def _is_equal_attribute(row_idx, attributes, row):
        new_value = attributes[row_idx].get('value', None)
        new_comment = attributes[row_idx].get('comment', None)
        return all(
            [new_value == row.value, new_comment == row.comment],
        )

    @staticmethod
    def _update_attributes_sheet_row(ws, row_idx):
        ws.cell(row_idx, 3, value='-')

    @staticmethod
    def _get_translation_id(translation):
        try:
            return translation['id']
        except TypeError:
            return translation

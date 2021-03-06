import pytest
import warnings

from cnctcli.actions.customers import dump_customers

from openpyxl import load_workbook


def test_dump_customers(fs, mocked_responses, mocked_customer, mocked_reseller):

    warnings.filterwarnings("ignore", category=UserWarning)
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/tier/accounts',
        json=[mocked_customer, mocked_reseller],
        headers={
            'Content-Range': 'items 0-1/1',
        }
    )

    output_file = dump_customers(
        api_url='https://localhost/public/v1',
        api_key='ApiKey XXX',
        silent=True,
        account_id='PA-1234',
        output_file=f'{fs.root_path}/Customers.xlsx'
    )
    customers_wb = load_workbook(
        output_file,
        data_only=True,
    )
    ws = customers_wb['Customers']
    assert len(ws['A']) == 3
    assert ws['A2'].value == mocked_customer['id']

from cnctcli.actions.products.clone_product import ProductCloner
from cnctcli.config import Config
from click import ClickException
from openpyxl import load_workbook

import os
import pytest

def test_dump(mocker, config_mocker):
    mock = mocker.patch(
        'cnctcli.actions.products.clone_product.dump_product'
    )
    config = Config()
    config.load('/tmp')
    cloner = ProductCloner(
        config=config,
        source_account='VA-000',
        destination_account='VA-000',
        product_id='PRD-123'
    )

    cloner.dump()

    mock.assert_called_once()


def test_clean_wb(
    config_mocker,
    fs,
):
    config = Config()
    config.load('/tmp')

    cloner = ProductCloner(
        config=config,
        source_account='VA-000',
        destination_account='VA-000',
        product_id='PRD-123'
    )

    fs.add_real_file('./tests/fixtures/comparation_product.xlsx')
    os.mkdir(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123'
        )
    )
    wb = load_workbook('./tests/fixtures/comparation_product.xlsx')
    wb.save(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
            'PRD-123.xlsx'
        )
    )
    cloner.load_wb()
    cloner.clean_wb()

    cloned_wb = load_workbook(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
            'PRD-123.xlsx'
        )
    )

    for row in range(2, 11):
        assert cloned_wb['Capabilities'][f'B{row}'].value == 'update'


def test_create_product(
    config_mocker,
    fs,
    mocked_responses,
    mocked_categories_response,
    mocked_product_response,
):
    config = Config()
    config.load('/tmp')
    config.add_account('VA-000', 'Account 0', 'Api 0', 'https://localhost/public/v1')

    cloner = ProductCloner(
        config=config,
        source_account='VA-000',
        destination_account='VA-000',
        product_id='PRD-123'
    )

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/categories?limit=100&offset=0',
        json=mocked_categories_response,
    )
    mocked_responses.add(
        method='POST',
        url='https://localhost/public/v1/products',
        json=mocked_product_response
    )

    fs.add_real_file('./tests/fixtures/comparation_product.xlsx')
    os.mkdir(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123'
        )
    )
    wb = load_workbook('./tests/fixtures/comparation_product.xlsx')
    wb.save(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
            'PRD-123.xlsx'
        )
    )
    cloner.load_wb()
    cloner.create_product()

    wb = load_workbook(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
            'PRD-123.xlsx'
        )
    )

    assert wb['General Information']['B5'].value == 'PRD-276-377-545'


def test_create_product_errordef(
    config_mocker,
    fs,
    mocked_responses,
    mocked_categories_response,
    mocked_product_response,
):
    config = Config()
    config.load('/tmp')
    config.add_account('VA-000', 'Account 0', 'Api 0', 'https://localhost/public/v1')

    cloner = ProductCloner(
        config=config,
        source_account='VA-000',
        destination_account='VA-000',
        product_id='PRD-123'
    )

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/categories?limit=100&offset=0',
        json=mocked_categories_response,
    )
    mocked_responses.add(
        method='POST',
        url='https://localhost/public/v1/products',
        status=500
    )

    fs.add_real_file('./tests/fixtures/comparation_product.xlsx')
    os.mkdir(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123'
        )
    )
    wb = load_workbook('./tests/fixtures/comparation_product.xlsx')
    wb.save(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
            'PRD-123.xlsx'
        )
    )
    cloner.load_wb()
    with pytest.raises(ClickException) as e:
        cloner.create_product()

    assert 'Error on product creation' in str(e)


def test_inject(
    config_mocker,
    fs,
    mocked_responses,
    mocker
):
    config = Config()
    config.load('/tmp')
    config.add_account('VA-000', 'Account 0', 'Api 0', 'https://localhost/public/v1')

    cloner = ProductCloner(
        config=config,
        source_account='VA-000',
        destination_account='VA-000',
        product_id='PRD-123'
    )

    fs.add_real_file('./tests/fixtures/comparation_product.xlsx')
    os.mkdir(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123'
        )
    )
    wb = load_workbook('./tests/fixtures/comparation_product.xlsx')
    wb.save(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
            'PRD-123.xlsx'
        )
    )

    mocker.patch(
        'cnctcli.actions.products.clone_product.GeneralSynchronizer'
    )
    mocker.patch(
        'cnctcli.actions.products.clone_product.CapabilitiesSynchronizer'
    )
    mocker.patch(
        'cnctcli.actions.products.clone_product.TemplatesSynchronizer'
    )
    mocker.patch(
        'cnctcli.actions.products.clone_product.ParamsSynchronizer'
    )
    mocker.patch(
        'cnctcli.actions.products.clone_product.ActionsSynchronizer'
    )
    mocker.patch(
        'cnctcli.actions.products.clone_product.MediaSynchronizer'
    )
    mocker.patch(
        'cnctcli.actions.products.clone_product.ItemSynchronizer',
        return_value=FakeItemSynchronizer
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545/items',
        json=[],
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545/templates?limit=100&offset=0',
        json=[]
    )
    cloner.inject()


class FakeItemSynchronizer:
    @staticmethod
    def open(first, second):
        return 'PRD-276-377-545'

    @staticmethod
    def sync():
        pass

from cnctcli.actions.products.utils import SheetNotFoundError
from click import ClickException
from cnctcli.actions.products.sync import ProductSynchronizer
from cnct import ConnectClient
from openpyxl import load_workbook
from shutil import copy2

import pytest
import os


def test_open(fs, mocked_responses, mocked_product_response):
    fs.add_real_file('./tests/fixtures/comparation_product.xlsx')
    synchronizer = ProductSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545',
        json=mocked_product_response,
    )
    product_id = synchronizer.open(
        './tests/fixtures/comparation_product.xlsx', 'Items'
    )

    assert product_id == 'PRD-276-377-545'


def test_invalid_file_open(fs, mocked_responses, mocked_product_response):
    fs.add_real_file('./tests/fixtures/configurations_response.json')
    copy2('./tests/fixtures/configurations_response.json', 'fake.xlsx')

    synchronizer = ProductSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    with pytest.raises(ClickException) as e:
        synchronizer.open('./fake.xlsx', 'Items')

    assert 'is not a valid xlsx file' in str(e.value)


def test_invalid_zip_open(fs, mocked_responses, mocked_product_response):
    fs.add_real_file('./tests/fixtures/configurations_response.json')
    synchronizer = ProductSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    with pytest.raises(ClickException) as e:
        synchronizer.open('./tests/fixtures/configurations_response.json', 'Items')

    assert 'openpyxl does not support .json' in str(e.value)


def test_open_product_not_found(fs, mocked_responses, mocked_product_response):
    fs.add_real_file('./tests/fixtures/comparation_product.xlsx')
    synchronizer = ProductSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545',
        json={},
        status=404,
    )
    with pytest.raises(ClickException) as e:
        synchronizer.open(
            './tests/fixtures/comparation_product.xlsx', 'Items'
        )
    assert str(e.value) == "Product PRD-276-377-545 not found, create it first."


def test_sheet_not_found(fs):
    fs.add_real_file('./tests/fixtures/comparation_product.xlsx')
    wb = load_workbook('./tests/fixtures/comparation_product.xlsx')
    wb.remove(wb['Items'])
    wb.save('./test.xlsx')

    synchronizer = ProductSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    with pytest.raises(SheetNotFoundError) as e:
        synchronizer.open(
            './test.xlsx', 'Items'
        )

    assert str(e.value) == 'File does not contain Items to synchronize, skipping'


def test_invalid_items_sheet(fs, mocked_responses, mocked_product_response):
    fs.add_real_file('./tests/fixtures/comparation_product.xlsx')
    wb = load_workbook('./tests/fixtures/comparation_product.xlsx')
    ws = wb['Items']
    ws['A1'].value = 'Modified'
    wb.save('./test.xlsx')

    synchronizer = ProductSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545',
        json=mocked_product_response,
    )

    with pytest.raises(ClickException) as e:
        synchronizer.open(
            './test.xlsx', 'Items'
        )

    assert str(e.value) == 'Invalid input file: column A must be ID'


def test_no_sync():
    synchronizer = ProductSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    with pytest.raises(NotImplementedError) as e:
        synchronizer.sync()

    assert str(e.value) == "Not implemented"


def test_save(fs, mocked_responses, mocked_product_response):

    fs.add_real_file('./tests/fixtures/comparation_product.xlsx')
    synchronizer = ProductSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545',
        json=mocked_product_response,
    )

    synchronizer.open(
        './tests/fixtures/comparation_product.xlsx', 'Items'
    )

    synchronizer.save('./test.xlsx')

    assert os.path.isfile('./test.xlsx')

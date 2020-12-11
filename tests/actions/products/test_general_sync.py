from click import ClickException
from cnctcli.actions.products.general_sync import GeneralSynchronizer
from cnct import ConnectClient
from openpyxl import load_workbook

import pytest
import json

GENERAL_ERROR = "Errors has been detected on General Information tab:"


def test_sheet_not_found(fs):
    fs.add_real_file('./tests/fixtures/comparation_product.xlsx')
    wb = load_workbook('./tests/fixtures/comparation_product.xlsx')
    wb.remove(wb['General Information'])
    wb.save('./test.xlsx')

    synchronizer = GeneralSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    with pytest.raises(ClickException) as e:
        synchronizer.open(
            './test.xlsx', 'General Information'
        )

    assert str(e.value) == 'File does not contain General Information to synchronize'


def test_no_product_id(fs):
    fs.add_real_file('./tests/fixtures/comparation_product.xlsx')
    wb = load_workbook('./tests/fixtures/comparation_product.xlsx')
    ws = wb['General Information']
    ws['A5'].value = 'Modified'
    wb.save('./test.xlsx')

    synchronizer = GeneralSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    with pytest.raises(ClickException) as e:
        synchronizer.open(
            './test.xlsx', 'General Information'
        )

    assert str(e.value) == 'Input file has invalid format and could not read product id from it'


def test_open_product_not_found(fs, mocked_responses, mocked_product_response):
    fs.add_real_file('./tests/fixtures/comparation_product.xlsx')
    synchronizer = GeneralSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545',
        json={},
        status=404,
    )
    with pytest.raises(ClickException) as e:
        synchronizer.open(
            './tests/fixtures/comparation_product.xlsx', 'General Information'
        )
    assert str(e.value) == "Product PRD-276-377-545 not found, create it first."


def test_validate_no_prod_category(get_general_env, mocked_responses):
    get_general_env['General Information']['A8'] = None
    get_general_env['General Information']['B8'] = None
    get_general_env.save('./test.xlsx')
    synchronizer = GeneralSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    mocked_responses.remove(
        method_or_response='GET',
        url='https://localhost/public/v1/categories',
    )
    with pytest.raises(ClickException) as e:
        synchronizer.open(
            './test.xlsx', 'General Information'
        )
    assert str(e.value) == f'{GENERAL_ERROR} A8 must be `Product Category` and B8 contain the value'


def test_validate_invalid_prod_category(get_general_env):
    get_general_env['General Information']['B8'] = 'invalid'
    get_general_env.save('./test.xlsx')
    synchronizer = GeneralSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    with pytest.raises(ClickException) as e:
        synchronizer.open(
            './test.xlsx', 'General Information'
        )
    assert str(e.value) == f'{GENERAL_ERROR} Product category invalid is a not known category'


def test_validate_invalid_icon(get_general_env):
    get_general_env['General Information']['A9'] = 'invalid'
    get_general_env.save('./test.xlsx')
    synchronizer = GeneralSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    with pytest.raises(ClickException) as e:
        synchronizer.open(
            './test.xlsx', 'General Information'
        )
    assert str(e.value) == (
        f'{GENERAL_ERROR} A9 must be `Product Icon file name` and B9 contain the value'
    )


def test_validate_invalid_icon_file(get_general_env):
    get_general_env['General Information']['B9'] = 'invalid.png'
    get_general_env.save('./test.xlsx')
    synchronizer = GeneralSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    with pytest.raises(ClickException) as e:
        synchronizer.open(
            './test.xlsx', 'General Information'
        )
    assert str(e.value) == (
        f'{GENERAL_ERROR} File invalid.png does not exist in the media folder'
    )


def test_validate_long_short_description(get_general_env):
    get_general_env['General Information']['B10'] = \
        get_general_env['General Information']['B11'].value
    get_general_env.save('./test.xlsx')
    synchronizer = GeneralSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    with pytest.raises(ClickException) as e:
        synchronizer.open(
            './test.xlsx', 'General Information'
        )
    assert str(e.value) == (
        f'{GENERAL_ERROR} Short description is mandatory and must be on B10, short description '
        f'can not exceed 512 characters'
    )


def test_validate_detailed_description(get_general_env):
    get_general_env['General Information']['A11'] = 'invalid'
    get_general_env.save('./test.xlsx')
    synchronizer = GeneralSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    with pytest.raises(ClickException) as e:
        synchronizer.open(
            './test.xlsx', 'General Information'
        )
    assert str(e.value) == f'{GENERAL_ERROR} Product detailed description is required'


def test_validate_embedding_description(get_general_env):
    get_general_env['General Information']['A12'] = 'invalid'
    get_general_env.save('./test.xlsx')
    synchronizer = GeneralSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    with pytest.raises(ClickException) as e:
        synchronizer.open(
            './test.xlsx', 'General Information'
        )
    assert str(e.value) == f'{GENERAL_ERROR} Embedding description is required'


def test_validate_embedding_long_description(get_general_env):
    get_general_env['General Information']['A13'] = 'invalid'
    get_general_env.save('./test.xlsx')
    synchronizer = GeneralSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )

    with pytest.raises(ClickException) as e:
        synchronizer.open(
            './test.xlsx', 'General Information'
        )
    assert str(e.value) == f'{GENERAL_ERROR} Embedding getting started is required'


def test_open(get_general_env):
    get_general_env.save('./test.xlsx')
    synchronizer = GeneralSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )
    product_id = synchronizer.open(
        './test.xlsx', 'General Information'
    )

    assert product_id == 'PRD-276-377-545'


def test_sync_409(get_general_env, mocked_responses):
    get_general_env.save('./test.xlsx')
    synchronizer = GeneralSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        silent=True,
    )
    product_id = synchronizer.open(
        './test.xlsx', 'General Information'
    )

    mocked_responses.add(
        method='PUT',
        url='https://localhost/public/v1/products/PRD-276-377-545',
        status=409,
    )

    errors = synchronizer.sync()

    assert product_id == 'PRD-276-377-545'
    assert errors == ['Error while updating general product information: 409 Conflict']


def test_sync(fs, get_general_env, mocked_responses):
    with open('./tests/fixtures/product_response.json') as prod_response:
        mocked_responses.add(
            method='PUT',
            url='https://localhost/public/v1/products/PRD-276-377-545',
            json=json.load(prod_response),
        )
        get_general_env.save('./test.xlsx')
        synchronizer = GeneralSynchronizer(
            client=ConnectClient(
                use_specs=False,
                api_key='ApiKey SU:123',
                endpoint='https://localhost/public/v1',
            ),
            silent=True,
        )
        product_id = synchronizer.open(
            './test.xlsx', 'General Information'
        )

        errors = synchronizer.sync()

        assert product_id == 'PRD-276-377-545'
        assert errors == []
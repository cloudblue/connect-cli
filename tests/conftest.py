import functools
import json
import os
from shutil import copy2

import pytest
import responses
from fs.tempfs import TempFS
from openpyxl import load_workbook
from responses.registries import OrderedRegistry

from tests.data import (
    CONFIG_DATA,
    EXTENSION_BG_EVENT,
    EXTENSION_CLASS_DECLARATION,
    EXTENSION_IMPORTS,
    EXTENSION_INTERACTIVE_EVENT,
    EXTENSION_SCHEDULABLE_EVENT,
    EXTENSION_VARIABLES_DECLARATION,
    TEST_BG_EVENT,
    TEST_INTERACTIVE_EVENT,
    TEST_SCHEDULABLE_EVENT,
)


@pytest.fixture(scope='function')
def fs():
    return TempFS()


@pytest.fixture(scope='session', autouse=True)
def patch_console():
    os.environ['COLUMNS'] = '132'
    os.environ['NO_COLOR'] = '1'


@pytest.fixture(scope='session')
def ccli():
    from connect.cli.core.base import cli
    from connect.cli.core.plugins import load_plugins

    load_plugins(cli)
    return cli


@pytest.fixture(scope='function')
def config_vendor():
    from connect.cli.core.config import Config

    config = Config()
    config.add_account('VA-001-002', 'name', 'api_key', 'https://localhost/public/v1')
    return config


@pytest.fixture(scope='function')
def config_provider():
    from connect.cli.core.config import Config

    config = Config()
    config.add_account('PA-001-002', 'name', 'api_key', 'https://localhost/public/v1')
    return config


@pytest.fixture(scope='function')
def config_mocker(mocker):
    mocker.patch('os.path.isfile', return_value=True)
    return mocker.patch(
        'connect.cli.core.config.open',
        mocker.mock_open(read_data=json.dumps(CONFIG_DATA)),
    )


@pytest.fixture(scope='function')
def mocked_responses():
    with responses.RequestsMock() as rsps:
        yield rsps


@pytest.fixture(scope='function')
def mocked_responses_ordered():
    with responses.RequestsMock(registry=OrderedRegistry) as rsps:
        yield rsps


@pytest.fixture(scope='function')
def mocked_product_response():
    with open('./tests/fixtures/product_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def mocked_categories_response():
    with open('./tests/fixtures/categories_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def mocked_media_response(fs):
    with open('./tests/fixtures/media_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def mocked_templates_response(fs):
    with open('./tests/fixtures/templates_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def mocked_items_response(fs):
    with open('./tests/fixtures/items_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def mocked_ordering_params_response(fs):
    with open('./tests/fixtures/ordering_parameters_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def mocked_fulfillment_params_response(fs):
    with open('./tests/fixtures/fulfillment_parameters_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def mocked_configuration_params_response(fs):
    with open('./tests/fixtures/configuration_parameters_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def mocked_actions_response(fs):
    with open('./tests/fixtures/actions_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def mocked_configurations_response(fs):
    with open('./tests/fixtures/configurations_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def mocked_locales_response():
    with open('./tests/fixtures/locales_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def mocked_product_translations_response():
    with open('./tests/fixtures/product_translations_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def mocked_new_translation_response():
    with open('./tests/fixtures/new_translation_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def sample_product_workbook(fs):
    return load_workbook('./tests/fixtures/comparation_product.xlsx')


@pytest.fixture(scope='function')
def get_sync_items_env(fs, mocked_responses):
    with open('./tests/fixtures/units_response.json') as response:
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/settings/units',
            json=json.load(response),
        )
        with open('./tests/fixtures/product_response.json') as prod_response:
            mocked_responses.add(
                method='GET',
                url='https://localhost/public/v1/products/PRD-276-377-545',
                json=json.load(prod_response),
            )

            return load_workbook('./tests/fixtures/items_sync.xlsx')


@pytest.fixture(scope='function')
def get_sync_actions_env(fs, mocked_responses):
    with open('./tests/fixtures/product_response_modifications.json') as prod_response:
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/products/PRD-276-377-545',
            json=json.load(prod_response),
        )

        return load_workbook('./tests/fixtures/actions_sync.xlsx')


@pytest.fixture(scope='function')
def get_sync_capabilities_env(fs, mocked_responses):
    with open('./tests/fixtures/product_response_modifications.json') as prod_response:
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/products/PRD-276-377-545',
            json=json.load(prod_response),
        )

        return load_workbook('./tests/fixtures/capabilities_sync.xlsx')


@pytest.fixture(scope='function')
def get_sync_params_env(fs, mocked_responses):
    with open('./tests/fixtures/product_response_modifications.json') as prod_response:
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/products/PRD-276-377-545',
            json=json.load(prod_response),
        )

        return load_workbook('./tests/fixtures/params_sync.xlsx')


@pytest.fixture(scope='function')
def get_sync_config_env(fs, mocked_responses):
    with open('./tests/fixtures/product_response_modifications.json') as prod_response:
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/products/PRD-276-377-545',
            json=json.load(prod_response),
        )

        return load_workbook('./tests/fixtures/configuration_sync.xlsx')


@pytest.fixture(scope='function')
def get_sync_templates_env(fs, mocked_responses):
    with open('./tests/fixtures/product_response.json') as prod_response:
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/products/PRD-276-377-545',
            json=json.load(prod_response),
        )

        return load_workbook('./tests/fixtures/templates_sync.xlsx')


@pytest.fixture(scope='function')
def get_general_env(fs, mocked_responses):
    os.mkdir(f'{fs.root_path}/media')
    copy2('./tests/fixtures/image.png', f'{fs.root_path}/media/PRD-276-377-545.png')
    with open('./tests/fixtures/categories_response.json') as cat_response:
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/categories',
            json=json.load(cat_response),
        )
        with open('./tests/fixtures/product_response.json') as prod_response:
            mocked_responses.add(
                method='GET',
                url='https://localhost/public/v1/products/PRD-276-377-545',
                json=json.load(prod_response),
            )

        return load_workbook('./tests/fixtures/comparation_product.xlsx')


@pytest.fixture(scope='function')
def get_sync_media_env(fs, mocked_responses):
    os.mkdir(f'{fs.root_path}/media')
    copy2('./tests/fixtures/image.png', f'{fs.root_path}/media/image.png')
    with open('./tests/fixtures/product_response.json') as prod_response:
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/products/PRD-276-377-545',
            json=json.load(prod_response),
        )

        return load_workbook('./tests/fixtures/media_sync.xlsx')


@pytest.fixture(scope='function')
def get_sync_capabilities_env_ppu_enabled(fs, mocked_responses):
    with open('./tests/fixtures/product_response_modifications.json') as prod_response:
        response = json.load(prod_response)
        response['capabilities']['ppu'] = {
            'schema': 'QT',
            'dynamic': False,
            'future': False,
            'predictive': False,
        }
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/products/PRD-276-377-545',
            json=response,
        )

        return load_workbook('./tests/fixtures/capabilities_sync.xlsx')


@pytest.fixture(scope='function')
def get_sync_translations_env(fs, mocked_responses):
    with open('./tests/fixtures/product_response.json') as prod_response:
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/products/PRD-276-377-545',
            json=json.load(prod_response),
        )

        return load_workbook('./tests/fixtures/translations_sync.xlsx')


@pytest.fixture(scope='function')
def mocked_reports(fs):
    with open('./tests/fixtures/reports.json') as response:
        return json.load(response)


@pytest.fixture
def mocked_customer():
    with open('./tests/fixtures/customer/customer.json') as response:
        return json.load(response)


@pytest.fixture
def mocked_reseller():
    with open('./tests/fixtures/customer/reseller.json') as response:
        return json.load(response)


@pytest.fixture
def customers_workbook(fs):
    return load_workbook('./tests/fixtures/customer/customers.xlsx')


@pytest.fixture(scope='function')
def sample_translation_workbook(fs):
    return load_workbook('./tests/fixtures/translation.xlsx')


@pytest.fixture(scope='function')
def mocked_translation_response():
    with open('./tests/fixtures/translation_response.json') as response:
        return json.load(response)


@pytest.fixture(scope='function')
def mocked_translation_attributes_xlsx_response():
    with open('./tests/fixtures/translation_attributes_response.xlsx') as response:
        yield response.buffer


@pytest.fixture(scope='function')
def mocked_translation_attributes_response():
    with open('./tests/fixtures/translation_attributes_response.json') as response:
        return json.load(response)


@pytest.fixture
def extension_class_declaration():
    def _declaration(extension_name, with_variables=True):
        if with_variables:
            return EXTENSION_VARIABLES_DECLARATION + EXTENSION_CLASS_DECLARATION.format(
                extension_name=extension_name,
            )
        return EXTENSION_CLASS_DECLARATION.format(extension_name=extension_name)

    return _declaration


@pytest.fixture
def extension_imports():
    def _imports(
        with_schedulable=True,
        with_variables=True,
        with_background=True,
        with_interactive=False,
    ):
        variables = ''
        schedulable = ''
        scheduled_response = ''
        background_response = ''
        interactive_response = ''

        if with_variables:
            variables = '\n    variables,'

        if with_schedulable:
            schedulable = '\n    schedulable,'
            scheduled_response = '\n    ScheduledExecutionResponse,'

        if with_background:
            background_response = '\n    BackgroundResponse,'

        if with_interactive:
            interactive_response = '\n    InteractiveResponse,'

        return EXTENSION_IMPORTS.format(
            variables=variables,
            schedulable=schedulable,
            scheduled_response=scheduled_response,
            background_response=background_response,
            interactive_response=interactive_response,
        )

    return _imports


def _event_handler(event_template, async_impl=False):
    async_def = ''
    if async_impl:
        async_def = 'async '
    return event_template.format(async_def=async_def)


def _test_handler(event_template, extension_name, async_impl=False):
    async_def = ''
    pytest_asyncio = ''
    await_keyword = ''
    client_mocker_prefix = ''
    connect_client_prefix = ''
    if async_impl:
        async_def = 'async '
        pytest_asyncio = '@pytest.mark.asyncio\n'
        await_keyword = 'await '
        client_mocker_prefix = 'async_'
        connect_client_prefix = 'async_'
    return event_template.format(
        extension_name=extension_name,
        async_def=async_def,
        pytest_asyncio=pytest_asyncio,
        await_keyword=await_keyword,
        client_mocker_prefix=client_mocker_prefix,
        connect_client_prefix=connect_client_prefix,
    )


@pytest.fixture
def extension_bg_event():
    return functools.partial(_event_handler, EXTENSION_BG_EVENT)


@pytest.fixture
def extension_schedulable_event():
    return functools.partial(_event_handler, EXTENSION_SCHEDULABLE_EVENT)


@pytest.fixture
def extension_interactive_event():
    return functools.partial(_event_handler, EXTENSION_INTERACTIVE_EVENT)


@pytest.fixture
def test_bg_event():
    return functools.partial(_test_handler, TEST_BG_EVENT)


@pytest.fixture
def test_interactive_event():
    return functools.partial(_test_handler, TEST_INTERACTIVE_EVENT)


@pytest.fixture
def test_schedulable_event():
    return functools.partial(_test_handler, TEST_SCHEDULABLE_EVENT)


@pytest.fixture(scope='function')
def console_80_columns(mocker):
    mocker.patch.dict(os.environ, {'COLUMNS': '80'})


@pytest.fixture
def always_yes_console():
    from connect.cli.core.terminal import console

    console.skip_confirm = True
    yield
    console.skip_confirm = False


@pytest.fixture
def client():
    from connect.client import ConnectClient

    return ConnectClient(
        api_key='ApiKey',
        endpoint='https://localhost/public/v1',
        use_specs=False,
    )


@pytest.fixture
def load_stream_sync():
    return load_workbook(
        './tests/fixtures/commerce/stream_sync.xlsx',
        read_only=True,
    )

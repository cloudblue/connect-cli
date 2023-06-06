import os

import pytest
from click import ClickException
from freezegun import freeze_time
from openpyxl import load_workbook
from responses import matchers

from connect.cli.core.config import Config
from connect.cli.plugins.product.clone import ProductCloner
from connect.cli.plugins.shared.sync_stats import SynchronizerStats


def test_dump(mocker, config_mocker):
    mock = mocker.patch(
        'connect.cli.plugins.product.clone.dump_product',
    )
    config = Config()
    config.load('/tmp')

    stats = SynchronizerStats()
    cloner = ProductCloner(
        config=config,
        source_account='VA-000',
        destination_account='VA-000',
        product_id='PRD-123',
        progress=mocker.MagicMock(),
        stats=stats,
    )

    cloner.dump()

    mock.assert_called_once()


def test_clean_wb(
    mocker,
    config_mocker,
    fs,
):
    config = Config()
    config.load('/tmp')

    stats = SynchronizerStats()
    cloner = ProductCloner(
        config=config,
        source_account='VA-000',
        destination_account='VA-000',
        product_id='PRD-123',
        progress=mocker.MagicMock(),
        stats=stats,
    )

    os.mkdir(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
        ),
    )
    wb = load_workbook('./tests/fixtures/comparation_product.xlsx')
    wb.save(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
            'PRD-123.xlsx',
        ),
    )
    cloner.load_wb()
    cloner.clean_wb()

    cloned_wb = load_workbook(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
            'PRD-123.xlsx',
        ),
    )

    for row in range(2, 11):
        assert cloned_wb['Capabilities'][f'B{row}'].value == 'update'
    for row in range(2, 7):
        assert cloned_wb['Embedding Static Resources'][f'C{row}'].value == 'create'
    for row in range(2, 5):
        assert cloned_wb['Media'][f'C{row}'].value == 'create'
    for row in range(2, 8):
        assert cloned_wb['Templates'][f'C{row}'].value == 'create'
    for row in range(2, 20):
        assert cloned_wb['Items'][f'C{row}'].value == 'create'
    for row in range(2, 14):
        assert cloned_wb['Ordering Parameters'][f'C{row}'].value == 'create'
    for row in range(2, 4):
        assert cloned_wb['Fulfillment Parameters'][f'C{row}'].value == 'create'
    for row in range(2, 3):
        assert cloned_wb['Configuration Parameters'][f'C{row}'].value == 'create'
    for row in range(2, 4):
        assert cloned_wb['Actions'][f'C{row}'].value == 'create'
    for row in range(2, 4):
        assert cloned_wb['Translations'][f'B{row}'].value == 'create'
    for row in range(2, 32):
        assert cloned_wb['FA (TRN-1079-0833-9890)'][f'C{row}'].value == '-'
    for row in range(2, 32):
        assert cloned_wb['ES-AR (TRN-1079-0833-9891)'][f'C{row}'].value == 'update'


@freeze_time('2022-04-05 20:15:00')
def test_create_product(
    mocker,
    mocked_responses,
    mocked_categories_response,
    mocked_product_response,
):
    config = Config()
    config.load('/tmp')
    config.add_account('VA-000', 'Account 0', 'Api 0', 'https://localhost/public/v1')

    stats = SynchronizerStats()
    cloner = ProductCloner(
        config=config,
        source_account='VA-000',
        destination_account='VA-000',
        product_id='PRD-123',
        progress=mocker.MagicMock(),
        stats=stats,
    )

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/categories?limit=100&offset=0',
        json=mocked_categories_response,
    )
    mocked_responses.add(
        method='POST',
        url='https://localhost/public/v1/products',
        match=[
            matchers.json_params_matcher(
                {
                    'name': 'Clone of PRD-123 2022-04-05-20:15:00',
                    'category': {
                        'id': 'CAT-59128',
                    },
                    'translations': [
                        {'locale': {'id': 'FA'}, 'primary': True},
                    ],
                },
            ),
        ],
        json=mocked_product_response,
    )

    os.mkdir(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
        ),
    )
    wb = load_workbook('./tests/fixtures/comparation_product.xlsx')
    wb.save(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
            'PRD-123.xlsx',
        ),
    )
    cloner.load_wb()
    cloner.create_product()

    wb = load_workbook(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
            'PRD-123.xlsx',
        ),
    )

    assert wb['General Information']['B5'].value == 'PRD-276-377-545'


def test_create_product_errordef(
    mocker,
    config_mocker,
    fs,
    mocked_responses,
    mocked_categories_response,
    mocked_product_response,
):
    config = Config()
    config.load('/tmp')
    config.add_account('VA-000', 'Account 0', 'Api 0', 'https://localhost/public/v1')

    stats = SynchronizerStats()
    cloner = ProductCloner(
        config=config,
        source_account='VA-000',
        destination_account='VA-000',
        product_id='PRD-123',
        progress=mocker.MagicMock(),
        stats=stats,
    )

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/categories?limit=100&offset=0',
        json=mocked_categories_response,
    )
    mocked_responses.add(
        method='POST',
        url='https://localhost/public/v1/products',
        status=500,
    )

    os.mkdir(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
        ),
    )
    wb = load_workbook('./tests/fixtures/comparation_product.xlsx')
    wb.save(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
            'PRD-123.xlsx',
        ),
    )
    cloner.load_wb()
    with pytest.raises(ClickException) as e:
        cloner.create_product()

    assert 'Error on product creation' in str(e)


def test_inject(
    config_mocker,
    fs,
    mocked_responses,
    mocker,
):
    config = Config()
    config.load('/tmp')
    config.add_account('VA-000', 'Account 0', 'Api 0', 'https://localhost/public/v1')

    stats = SynchronizerStats()
    cloner = ProductCloner(
        config=config,
        source_account='VA-000',
        destination_account='VA-000',
        product_id='PRD-123',
        progress=mocker.MagicMock(),
        stats=stats,
    )

    os.mkdir(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
        ),
    )
    wb = load_workbook('./tests/fixtures/comparation_product.xlsx')
    wb.save(
        os.path.join(
            cloner.fs.root_path,
            'PRD-123',
            'PRD-123.xlsx',
        ),
    )

    mocker.patch(
        'connect.cli.plugins.product.clone.GeneralSynchronizer',
    )
    mocker.patch(
        'connect.cli.plugins.product.clone.CapabilitiesSynchronizer',
    )
    mocker.patch(
        'connect.cli.plugins.product.clone.TemplatesSynchronizer',
    )
    mocker.patch(
        'connect.cli.plugins.product.clone.ParamsSynchronizer',
    )
    mocker.patch(
        'connect.cli.plugins.product.clone.ActionsSynchronizer',
    )
    mocker.patch(
        'connect.cli.plugins.product.clone.MediaSynchronizer',
    )
    mocker.patch(
        'connect.cli.plugins.product.clone.ItemSynchronizer',
        return_value=FakeItemSynchronizer,
    )
    mocker.patch(
        'connect.cli.plugins.product.clone.sync_product_translations',
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545/items',
        json=[],
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545/templates?limit=100&offset=0',
        json=[],
    )
    cloner.inject()


class FakeItemSynchronizer:
    @staticmethod
    def open(first, second):
        return 'PRD-276-377-545'

    @staticmethod
    def sync():
        pass

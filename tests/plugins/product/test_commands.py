import json
import re
import string
import os

import pytest

from click import ClickException
from click.testing import CliRunner

from openpyxl import load_workbook

from connect.cli.core.config import Config
from connect.cli.plugins.product.export import dump_product


def test_sync_general_sync(fs, get_general_env, mocked_responses, ccli):
    config = Config()
    config.load(fs.root_path)
    config.add_account(
        'VA-000',
        'Account 1',
        'ApiKey XXXX:YYYY',
        endpoint='https://localhost/public/v1',
    )
    config.activate('VA-000')
    config.store()
    assert os.path.isfile(f'{fs.root_path}/config.json') is True

    with open('./tests/fixtures/units_response.json') as units_response:
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/settings/units',
            json=json.load(units_response),
        )

        with open('./tests/fixtures/product_response.json') as prod_response:
            mocked_responses.add(
                method='PUT',
                url='https://localhost/public/v1/products/PRD-276-377-545',
                json=json.load(prod_response),
            )
            get_general_env.save(f'{fs.root_path}/test.xlsx')

            runner = CliRunner()
            result = runner.invoke(
                ccli,
                [
                    '-c',
                    fs.root_path,
                    'product',
                    'sync',
                    '--yes',
                    f'{fs.root_path}/test.xlsx',
                ],
            )

            assert result.exit_code == 0


def test_list_products(fs, mocked_responses, ccli):
    with open('./tests/fixtures/product_response.json') as prod_response:
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/products',
            json=[json.load(prod_response)],
        )
    config = Config()
    config.load(fs.root_path)
    config.add_account(
        'VA-000',
        'Account 1',
        'ApiKey XXXX:YYYY',
        endpoint='https://localhost/public/v1',
    )
    config.activate('VA-000')
    config.store()
    assert os.path.isfile(f'{fs.root_path}/config.json') is True
    runner = CliRunner()
    result = runner.invoke(
        ccli,
        [
            '-c',
            fs.root_path,
            'product',
            'list',
        ],
    )

    assert result.exit_code == 0
    assert "PRD-276-377-545 - My Produc" in result.output


def test_export(config_mocker, mocker, ccli):

    mock = mocker.patch(
        'connect.cli.plugins.product.commands.dump_product',
        side_effect=lambda *args: 'PRD-000.xlsx',
    )

    runner = CliRunner()
    result = runner.invoke(
        ccli,
        [
            'product',
            'export',
            'PRD-000',
        ],
    )
    mock.assert_called_once()
    assert mock.mock_calls[0][1][2] == 'PRD-000'
    assert mock.mock_calls[0][1][3] is None
    assert result.exit_code == 0
    assert 'The product PRD-000 has been successfully exported to PRD-000.xlsx.\n' in result.output


def test_export_custom_file(config_mocker, mocker, ccli):
    mock = mocker.patch(
        'connect.cli.plugins.product.commands.dump_product',
        side_effect=lambda *args: '/tmp/my_product.xlsx',
    )

    runner = CliRunner()
    result = runner.invoke(
        ccli,
        [
            'product',
            'export',
            'PRD-000',
            '-o',
            '/tmp/my_product.xlsx',
        ],
    )
    mock.assert_called_once()
    assert mock.mock_calls[0][1][2] == 'PRD-000'
    assert mock.mock_calls[0][1][3] == '/tmp/my_product.xlsx'
    assert result.exit_code == 0
    assert 'The product PRD-000 has been successfully exported to /tmp/my_product.xlsx.\n' in result.output


def test_export_product_not_exists(fs, mocked_responses):
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-0000',
        status=404,
    )
    with pytest.raises(ClickException) as e:
        dump_product(
            api_url='https://localhost/public/v1',
            api_key='ApiKey SU111:1111',
            product_id='PRD-0000',
            output_path=fs.root_path,
            output_file='output.xlsx',
            silent=True,
        )

    assert str(e.value) == '404 - Not Found: Product PRD-0000 not found.'


def test_export_product(
    fs,
    mocked_responses,
    mocked_product_response,
    mocked_categories_response,
    mocked_media_response,
    mocked_templates_response,
    mocked_items_response,
    mocked_ordering_params_response,
    mocked_fulfillment_params_response,
    mocked_configuration_params_response,
    mocked_actions_response,
    mocked_configurations_response,
    mocked_locales_response,
    mocked_primary_translation_response,
    sample_product_workbook,
):
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545',
        json=mocked_product_response,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/media/VA-392-495/PRD-276-377-545/media/PRD-276-377-545'
            '-logo_aJD74iQ.png',
        body=open('./tests/fixtures/image.png', 'rb').read(),
        status=200,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/media/VA-392-495/PRD-276-377-545/media/media.png',
        body=open('./tests/fixtures/image.png', 'rb').read(),
        status=200,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/media/VA-392-495/PRD-276-377-545/media/media_jxr1ifH.png',
        body=open('./tests/fixtures/image.png', 'rb').read(),
        status=200,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/media/VA-392-495/PRD-276-377-545/media/media_cqhgp78.png',
        body=open('./tests/fixtures/image.png', 'rb').read(),
        status=200,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/categories',
        json=mocked_categories_response,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545/media',
        json=mocked_media_response,
        headers={
            'Content-Range': 'items 0-2/3',
        },
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545/templates',
        json=mocked_templates_response,
        headers={
            'Content-Range': 'items 0-5/6',
        },
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545/items',
        json=mocked_items_response,
        headers={
            'Content-Range': 'items 0-17/18',
        },
    )
    ordering_query = re.compile(
        r'https:\/\/localhost\/public\/v1\/products\/PRD-276-377-545\/parameters\?eq\(phase,'
        r'ordering\)&limit\=(100|[1-9]?[0-9])\&offset\=0',
    )
    fulfillment_query = re.compile(
        r'https:\/\/localhost\/public\/v1\/products\/PRD-276-377-545\/parameters\?eq\(phase,'
        r'fulfillment\)&limit\=(100|[1-9]?[0-9])\&offset\=0',
    )
    configuration_query = re.compile(
        r'https:\/\/localhost\/public\/v1\/products\/PRD-276-377-545\/parameters\?eq\(phase,'
        r'configuration\)&limit\=(100|[1-9]?[0-9])\&offset\=0',
    )
    translation_query = re.compile(
        r'https:\/\/localhost\/public\/v1\/localization\/translations\?eq\(context.instance_id,'
        r'PRD-276-377-545\)&limit\=(100|[1-9]?[0-9])\&offset\=0',
    )
    mocked_responses.add(
        method='GET',
        url=ordering_query,
        json=mocked_ordering_params_response,
        headers={
            'Content-Range': 'items 0-11/12',
        },
    )
    mocked_responses.add(
        method='GET',
        url=fulfillment_query,
        json=mocked_fulfillment_params_response,
        headers={
            'Content-Range': 'items 0-1/2',
        },
    )
    mocked_responses.add(
        method='GET',
        url=configuration_query,
        json=mocked_configuration_params_response,
        headers={
            'Content-Range': 'items 0-0/1',
        },
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545/actions',
        json=mocked_actions_response,
        headers={
            'Content-Range': 'items 0-1/2',
        },
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545/configurations',
        json=mocked_configurations_response,
        headers={
            'Content-Range': 'items 0-17/18',
        },
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/localization/locales',
        json=mocked_locales_response,
    )
    mocked_responses.add(
        method='GET',
        url=(
            'https://localhost/public/v1/localization/translations?'
            'and(eq(context.instance_id,PRD-457-715-047),eq(primary,true))&limit=100&offset=0'
        ),
        json=mocked_primary_translation_response,
    )
    mocked_responses.add(
        method='GET',
        url=translation_query,
        json=mocked_primary_translation_response,
        headers={
            'Content-Range': 'items 0-0/1',
        },
    )
    output_file = dump_product(
        api_url='https://localhost/public/v1',
        api_key='ApiKey SU111:1111',
        product_id='PRD-276-377-545',
        output_file='output.xlsx',
        output_path=fs.root_path,
        silent=True,
    )

    product_wb = load_workbook(output_file)
    for name in sample_product_workbook.sheetnames:
        assert name in product_wb.sheetnames

    ignore_sheet_cells = {
        'General Information': ['B7'],
    }
    for sheet in sample_product_workbook.sheetnames:
        sample_sheet = sample_product_workbook[sheet]
        product_sheet = product_wb[sheet]
        letter_limit = _get_col_limit_by_type(sheet)

        for col in string.ascii_uppercase:
            if col == letter_limit:
                break
            for row in range(1, sample_sheet.max_row + 1):
                cell = f'{col}{row}'
                if cell not in ignore_sheet_cells.get(sheet, []):
                    assert product_sheet[cell].value == sample_sheet[cell].value


def _get_col_limit_by_type(ws_type):
    if ws_type == 'General Information':
        return 'C'
    elif ws_type == 'Capabilities':
        return 'C'
    elif ws_type == 'Embedding Static Resources':
        return 'D'
    elif ws_type == 'Media':
        return 'F'
    elif ws_type == 'Templates':
        return 'F'
    elif ws_type == 'Items':
        return 'M'
    elif ws_type == 'Ordering Parameters':
        return 'L'
    elif ws_type == 'Fulfillment Parameters':
        return 'L'
    elif ws_type == 'Configuration Parameters':
        return 'L'
    elif ws_type == 'Actions':
        return 'G'
    elif ws_type == 'Configuration':
        return 'G'
    elif ws_type == 'Translations':
        return 'K'
    return 'Z'

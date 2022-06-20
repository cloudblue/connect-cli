import warnings

from connect.client import ConnectClient
from connect.cli.plugins.customer.export import dump_customers

from openpyxl import load_workbook


def test_dump_customers(fs, mocked_responses, mocked_customer, mocked_reseller):

    warnings.filterwarnings("ignore", category=UserWarning)
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/tier/accounts',
        json=[mocked_customer, mocked_reseller],
        headers={
            'Content-Range': 'items 0-1/1',
        },
    )

    output_file = dump_customers(
        ConnectClient(
            'ApiKey XXX',
            endpoint='https://localhost/public/v1',
            use_specs=False,
        ),
        account_id='PA-1234',
        output_path=fs.root_path,
        output_file='Customers.xlsx',
    )
    customers_wb = load_workbook(
        output_file,
        data_only=True,
    )
    ws = customers_wb['Customers']
    assert len(ws['A']) == 3
    assert ws['A2'].value == mocked_customer['id']


def test_dump_customers_client_error(mocker, fs, mocked_responses):

    warnings.filterwarnings("ignore", category=UserWarning)
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/tier/accounts',
        status=400,
    )
    mocked_handle_error = mocker.patch(
        'connect.cli.plugins.customer.export.handle_http_error',
    )

    dump_customers(
        ConnectClient(
            'ApiKey XXX',
            endpoint='https://localhost/public/v1',
            use_specs=False,
        ),
        account_id='PA-1234',
        output_path=fs.root_path,
        output_file='Customers.xlsx',
    )

    mocked_handle_error.assert_called_once()

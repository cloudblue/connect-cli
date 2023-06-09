from zipfile import BadZipFile

import pytest
from click import ClickException
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

from connect.cli.plugins.customer.sync import CustomerSynchronizer
from connect.cli.plugins.shared.exceptions import SheetNotFoundError


def test_sync_all_skip(fs, customers_workbook, client):
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats.get_counts_as_dict() == {
        'processed': 2,
        'created': 0,
        'updated': 0,
        'deleted': 0,
        'skipped': 2,
        'errors': 0,
    }


def test_bad_action(fs, customers_workbook, client):
    customers_workbook['Customers']['D2'] = 'wrong'
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {2: ['Action wrong is not supported']}


def test_bad_account(fs, customers_workbook, client):
    customers_workbook['Customers']['D2'] = 'create'
    customers_workbook['Customers']['H2'] = 'robot'
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {
        2: ['Customer type must be customer or reseller, not robot'],
    }


def test_empty_address(fs, customers_workbook, client):
    customers_workbook['Customers']['D2'] = 'create'
    customers_workbook['Customers']['K2'] = ''
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {
        2: ['Address line 1, city, state and zip are mandatory'],
    }


def test_create_existing(fs, customers_workbook, client):
    customers_workbook['Customers']['D2'] = 'create'
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {
        2: ['Create action must not have account id, is set to TA-7374-0753-1907'],
    }


def test_create_customer_no_parent(fs, customers_workbook, client):
    customers_workbook['Customers']['D2'] = 'create'
    customers_workbook['Customers']['H2'] = 'customer'
    customers_workbook['Customers']['A2'] = None
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {2: ['Customers requires a parent account']}


def test_update_customer_no_id(fs, customers_workbook, client):
    customers_workbook['Customers']['D2'] = 'update'
    customers_workbook['Customers']['A2'] = 'KA-'
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {2: ['Update operation requires account ID to be set']}


def test_update_customer_no_account_connect(fs, customers_workbook, mocked_responses, client):
    customers_workbook['Customers']['D2'] = 'update'
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/tier/accounts/TA-7374-0753-1907',
        status=404,
    )
    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {
        2: ['Account with id TA-7374-0753-1907 does not exist'],
    }


def test_create_account_connect(fs, customers_workbook, mocked_responses, mocked_reseller, client):
    customers_workbook['Customers']['D2'] = 'create'
    customers_workbook['Customers']['A2'] = None
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    mocked_responses.add(
        method='POST',
        url='https://localhost/public/v1/tier/accounts',
        json=mocked_reseller,
    )
    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._created == 1


def test_create_account_connect_uuid(
    fs,
    customers_workbook,
    mocked_responses,
    mocked_reseller,
    client,
):
    customers_workbook['Customers']['D2'] = 'create'
    customers_workbook['Customers']['A2'] = None
    customers_workbook['Customers']['C2'] = None
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    mocked_responses.add(
        method='POST',
        url='https://localhost/public/v1/tier/accounts',
        json=mocked_reseller,
    )
    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._created == 1


def test_create_account_connect_parent_id(
    fs,
    customers_workbook,
    mocked_responses,
    mocked_reseller,
    client,
):
    customers_workbook['Customers']['D3'] = 'create'
    customers_workbook['Customers']['A3'] = None
    customers_workbook['Customers']['C3'] = None
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    mocked_responses.add(
        method='POST',
        url='https://localhost/public/v1/tier/accounts',
        json=mocked_reseller,
    )
    mocked_responses.add(
        method='GET',
        url=f'https://localhost/public/v1/tier/accounts/{mocked_reseller["id"]}',
        json=mocked_reseller,
    )
    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._created == 1


def test_create_account_connect_parent_id_not_found(
    fs,
    customers_workbook,
    mocked_responses,
    mocked_reseller,
    client,
):
    customers_workbook['Customers']['D3'] = 'create'
    customers_workbook['Customers']['A3'] = None
    customers_workbook['Customers']['C3'] = None
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    mocked_responses.add(
        method='GET',
        url=f'https://localhost/public/v1/tier/accounts/{mocked_reseller["id"]}',
        status=404,
    )
    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {
        3: [f'Parent with id {mocked_reseller["id"]} does not exist'],
    }


def test_create_account_connect_parent_external_id(
    fs,
    customers_workbook,
    mocked_responses,
    mocked_reseller,
    client,
):
    customers_workbook['Customers']['D3'] = 'create'
    customers_workbook['Customers']['A3'] = None
    customers_workbook['Customers']['C3'] = None
    customers_workbook['Customers']['F3'] = 'external_id'
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    mocked_responses.add(
        method='GET',
        url=f'https://localhost/public/v1/tier/accounts?eq(external_id,{mocked_reseller["id"]})&limit=0&offset=0',
        json=[mocked_reseller],
        headers={
            'Content-Range': 'items 0-1/1',
        },
    )
    mocked_responses.add(
        method='GET',
        url=f'https://localhost/public/v1/tier/accounts?eq(external_id,{mocked_reseller["id"]})&limit=1&offset=0',
        json=[mocked_reseller],
        headers={
            'Content-Range': 'items 0-1/1',
        },
    )
    mocked_responses.add(
        method='POST',
        url='https://localhost/public/v1/tier/accounts',
        json=mocked_reseller,
    )
    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._created == 1


def test_create_account_connect_parent_external_id_not_found(
    fs,
    customers_workbook,
    mocked_responses,
    mocked_reseller,
    client,
):
    customers_workbook['Customers']['D3'] = 'create'
    customers_workbook['Customers']['A3'] = None
    customers_workbook['Customers']['C3'] = None
    customers_workbook['Customers']['F3'] = 'external_id'
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    mocked_responses.add(
        method='GET',
        url=f'https://localhost/public/v1/tier/accounts?eq(external_id,{mocked_reseller["id"]})&limit=0&offset=0',
        json=[],
        headers={
            'Content-Range': 'items 0-0/0',
        },
    )
    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {
        3: ['Parent with external_id TA-7374-0753-1907 not found'],
    }


def test_create_account_connect_parent_external_id_more_than_one(
    fs,
    customers_workbook,
    mocked_responses,
    mocked_reseller,
    client,
):
    customers_workbook['Customers']['D3'] = 'create'
    customers_workbook['Customers']['A3'] = None
    customers_workbook['Customers']['C3'] = None
    customers_workbook['Customers']['F3'] = 'external_id'
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    mocked_responses.add(
        method='GET',
        url=f'https://localhost/public/v1/tier/accounts?eq(external_id,{mocked_reseller["id"]})&limit=0&offset=0',
        json=[],
        headers={
            'Content-Range': 'items 0-2/2',
        },
    )
    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {
        3: ['More than one Parent with external_id TA-7374-0753-1907'],
    }


def test_create_account_connect_parent_external_uid(
    fs,
    customers_workbook,
    mocked_responses,
    mocked_reseller,
    client,
):
    customers_workbook['Customers']['D3'] = 'create'
    customers_workbook['Customers']['A3'] = None
    customers_workbook['Customers']['C3'] = None
    customers_workbook['Customers']['F3'] = 'external_uid'
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    mocked_responses.add(
        method='GET',
        url=f'https://localhost/public/v1/tier/accounts?eq(external_uid,{mocked_reseller["id"]})&limit=0&offset=0',
        json=[mocked_reseller],
        headers={
            'Content-Range': 'items 0-1/1',
        },
    )
    mocked_responses.add(
        method='GET',
        url=f'https://localhost/public/v1/tier/accounts?eq(external_uid,{mocked_reseller["id"]})&limit=1&offset=0',
        json=[mocked_reseller],
        headers={
            'Content-Range': 'items 0-1/1',
        },
    )
    mocked_responses.add(
        method='POST',
        url='https://localhost/public/v1/tier/accounts',
        json=mocked_reseller,
    )
    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._created == 1


def test_create_account_connect_parent_external_uid_not_found(
    fs,
    customers_workbook,
    mocked_responses,
    mocked_reseller,
    client,
):
    customers_workbook['Customers']['D3'] = 'create'
    customers_workbook['Customers']['A3'] = None
    customers_workbook['Customers']['C3'] = None
    customers_workbook['Customers']['F3'] = 'external_uid'
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    mocked_responses.add(
        method='GET',
        url=f'https://localhost/public/v1/tier/accounts?eq(external_uid,{mocked_reseller["id"]})&limit=0&offset=0',
        json=[],
        headers={
            'Content-Range': 'items 0-0/0',
        },
    )
    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {
        3: ['Parent with external_uid TA-7374-0753-1907 not found'],
    }


def test_create_account_connect_parent_external_uid_more_than_one(
    fs,
    customers_workbook,
    mocked_responses,
    mocked_reseller,
    client,
):
    customers_workbook['Customers']['D3'] = 'create'
    customers_workbook['Customers']['A3'] = None
    customers_workbook['Customers']['C3'] = None
    customers_workbook['Customers']['F3'] = 'external_uid'
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    mocked_responses.add(
        method='GET',
        url=f'https://localhost/public/v1/tier/accounts?eq(external_uid,{mocked_reseller["id"]})&limit=0&offset=0',
        json=[],
        headers={
            'Content-Range': 'items 0-2/2',
        },
    )
    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {
        3: ['More than one Parent with external_uid TA-7374-0753-1907'],
    }


def test_parent_search_criteria(
    fs,
    mocker,
    client,
    customers_workbook,
):
    mocked_data = mocker.MagicMock()
    mocked_data.parent_search_criteria = 'criteria'
    mocked_data.parent_search_value = None
    mocker.patch('connect.cli.plugins.customer.sync._RowData', return_value=mocked_data)
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    CustomerSynchronizer._validate_row = lambda _, x: None
    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {
        2: ['Parent search value is needed if criteria is set'],
        3: ['Parent search value is needed if criteria is set'],
    }


def test_account_hub_cannot_be_modified(
    fs,
    mocker,
    client,
    customers_workbook,
):
    mocked_data = mocker.MagicMock()
    mocked_data.hub_id = 'HUB'
    mocker.patch('connect.cli.plugins.customer.sync._RowData', return_value=mocked_data)
    customers_workbook.save(f'{fs.root_path}/test.xlsx')

    CustomerSynchronizer._validate_row = lambda _, x: None
    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=client,
    )
    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Customers')
    synchronizer.sync()
    assert synchronizer.stats._row_errors == {
        2: ['Accounts on hub HUB can not be modified'],
        3: ['Accounts on hub HUB can not be modified'],
    }


def test_open_invalid_file(mocker):
    mocker.patch(
        'connect.cli.plugins.customer.sync.load_workbook',
        side_effect=InvalidFileException('it is a json!'),
    )

    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=None,
    )

    with pytest.raises(ClickException) as cv:
        synchronizer.open('file.xlsx', 'Customers')

    assert str(cv.value) == 'it is a json!'


def test_open_bad_zip(mocker):
    mocker.patch(
        'connect.cli.plugins.customer.sync.load_workbook',
        side_effect=BadZipFile(),
    )

    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=None,
    )

    with pytest.raises(ClickException) as cv:
        synchronizer.open('file.xlsx', 'Customers')

    assert str(cv.value) == 'file.xlsx is not a valid xlsx file.'


def test_open_sheet_not_found(mocker):
    mocker.patch(
        'connect.cli.plugins.customer.sync.load_workbook',
        return_value=Workbook(),
    )

    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=None,
    )

    with pytest.raises(SheetNotFoundError) as cv:
        synchronizer.open('file.xlsx', 'Customers')

    assert str(cv.value) == 'File does not contain Customers to synchronize, skipping'


def test_open_invalid_columns(mocker):
    wb = Workbook()
    ws = wb.create_sheet('Customers')
    ws['A1'].value = 'Invalid column'
    mocker.patch(
        'connect.cli.plugins.customer.sync.load_workbook',
        return_value=wb,
    )

    synchronizer = CustomerSynchronizer(
        account_id='VA-123',
        client=None,
    )

    with pytest.raises(ClickException) as cv:
        synchronizer.open('file.xlsx', 'Customers')

    assert str(cv.value) == 'Column `A1` must be `ID` and is `Invalid column`.'


def test_populate_hubs(mocker, mocked_responses, client):
    mocker.patch(
        'connect.cli.plugins.customer.sync.load_workbook',
        return_value=Workbook(),
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/hubs?limit=100&offset=0',
        json=[
            {
                'id': 'HB-001',
                'instance': {'type': 'custom'},
            },
            {
                'id': 'HB-002',
                'instance': {'type': 'OA'},
            },
        ],
        headers={
            'Content-Range': 'items 0-2/2',
        },
    )

    synchronizer = CustomerSynchronizer(
        account_id='PA-123',
        client=client,
    )
    synchronizer.populate_hubs()

    assert synchronizer.hubs == [
        'HB-0000-0000',
        'HB-001',
    ]


def test_save(mocker):
    synchronizer = CustomerSynchronizer(
        account_id='PA-123',
        client=None,
    )
    synchronizer._wb = mocker.MagicMock()

    synchronizer.save('file.xlsx')

    synchronizer._wb.save.assert_called_once_with('file.xlsx')

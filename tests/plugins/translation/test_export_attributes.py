import os.path
from openpyxl import Workbook, load_workbook

from connect.client import ConnectClient
from connect.cli.plugins.translation.export_attributes import dump_translation_attributes


def test_dump_translation_attributes(
    fs,
    mocked_responses,
    mocked_translation_attributes_response,
    sample_translation_workbook,
):
    client = ConnectClient(
        api_key='ApiKey SU:123',
        use_specs=False,
        endpoint='https://localhost/public/v1',
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/localization/translations/TRN-8100-3865-4869/attributes',
        json=mocked_translation_attributes_response,
        headers={
            'Content-Range': 'items 0-29/30',
        }
    )
    wb = Workbook()
    wb.create_sheet('Attributes')
    dump_translation_attributes(
        ws=wb['Attributes'],
        client=client,
        translation_id='TRN-8100-3865-4869',
        silent=True,
    )

    # must save and reload Workbook to fully test comparing with sample file
    filename = os.path.join(fs.root_path, 'xxx.xlsx')
    wb.save(filename)
    wb = load_workbook(filename, data_only=True)

    sample_sheet = sample_translation_workbook['Attributes']
    for col in 'ABCD':
        for row in range(1, sample_sheet.max_row + 1):
            cell = f'{col}{row}'
            assert wb['Attributes'][cell].value == sample_sheet[cell].value

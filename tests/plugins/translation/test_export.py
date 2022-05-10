import string

import pytest
from click import ClickException
from openpyxl import load_workbook

from connect.cli.plugins.translation.export import dump_translation


_LAST_COLUMN_BY_SHEET = {
    'Instructions': 'B',
    'General': 'B',
    'Attributes': 'D',
}


def test_dump_translation(
    fs,
    mocked_responses,
    mocked_translation_attributes_xlsx_response,
    sample_translation_workbook,
):
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/localization/translations/TRN-8100-3865-4869/attributes',
        body=mocked_translation_attributes_xlsx_response,
        headers={
            'Contet-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        },
    )

    output_file = dump_translation(
        api_url='https://localhost/public/v1',
        api_key='ApiKey XXX',
        translation_id='TRN-8100-3865-4869',
        output_file='translation.xlsx',
        output_path=fs.root_path,
    )

    translation_wb = load_workbook(output_file, data_only=True)
    for sheetname in sample_translation_workbook.sheetnames:
        assert sheetname in translation_wb.sheetnames

    for sheetname in sample_translation_workbook.sheetnames:
        sample_sheet = sample_translation_workbook[sheetname]
        translation_sheet = translation_wb[sheetname]
        letter_limit = _LAST_COLUMN_BY_SHEET[sheetname]
        for col in string.ascii_uppercase:
            if col == letter_limit:
                break
            for row in range(1, sample_sheet.max_row + 1):
                cell = f'{col}{row}'
                assert translation_sheet[cell].value == sample_sheet[cell].value


def test_dump_translation_not_exists(fs, mocked_responses):
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/localization/translations/TRN-0000-0000-0000/attributes',
        status=404,
    )
    with pytest.raises(ClickException) as e:
        dump_translation(
            api_url='https://localhost/public/v1',
            api_key='ApiKey XXX',
            translation_id='TRN-0000-0000-0000',
            output_file='translation.xlsx',
            output_path=fs.root_path,
        )

    assert str(e.value) == '404 - Not Found: Translation TRN-0000-0000-0000 not found.'

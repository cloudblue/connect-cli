from openpyxl import Workbook

from connect.cli.plugins.translation.utils import insert_column_ws


def test_insert_column_ws():
    ws = Workbook().active
    ws.append(['Column A', 'Column B', 'Column C'])
    ws.append(['A.1', 'B.1', 'C.1'])
    ws.append(['A.2', 'B.2', 'C.2'])
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 90

    insert_column_ws(ws, 2, 50)

    assert ws.column_dimensions['A'].width == 20
    assert ws.column_dimensions['B'].width == 50
    assert ws.column_dimensions['C'].width == 35
    assert ws.column_dimensions['D'].width == 90


def test_insert_column_ws_default_width():
    ws = Workbook().active
    ws.append(['Column A', 'Column B', 'Column C'])
    ws.append(['A.1', 'B.1', 'C.1'])
    ws.append(['A.2', 'B.2', 'C.2'])
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 90

    insert_column_ws(ws, 2)

    assert ws.column_dimensions['A'].width == 20
    assert ws.column_dimensions['B'].width == 35
    assert ws.column_dimensions['C'].width == 35
    assert ws.column_dimensions['D'].width == 90


def test_insert_column_ws_extra_right():
    ws = Workbook().active
    ws.append(['Column A', 'Column B', 'Column C'])
    ws.append(['A.1', 'B.1', 'C.1'])
    ws.append(['A.2', 'B.2', 'C.2'])
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 90

    insert_column_ws(ws, 4, 50)

    assert ws.column_dimensions['A'].width == 20
    assert ws.column_dimensions['B'].width == 35
    assert ws.column_dimensions['C'].width == 90
    assert ws.column_dimensions['D'].width == 50

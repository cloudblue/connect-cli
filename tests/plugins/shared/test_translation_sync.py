from copy import deepcopy

import pytest
import responses
from openpyxl import load_workbook

from connect.cli.plugins.shared.sync_stats import SynchronizerStats
from connect.cli.plugins.shared.translation_sync import TranslationsSynchronizer
from connect.client import ConnectClient


def test_skipped(mocker, fs, get_sync_translations_env):
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 2, 'created': 0, 'updated': 0,
        'deleted': 0, 'skipped': 2, 'errors': 0,
    }


def test_validate_invalid_action(mocker, fs, get_sync_translations_env):
    get_sync_translations_env['Translations']['B3'] = 'invalid'
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 2, 'created': 0, 'updated': 0,
        'deleted': 0, 'skipped': 1, 'errors': 1,
    }
    assert stats['Translations']._row_errors == {
        3: ['Action must be `-`, `delete`, `update` or `create`. Provided invalid'],
    }


@pytest.mark.parametrize('action', ('-', 'create', 'update', 'delete'))
def test_validate_primary_is_skipped(mocker, fs, get_sync_translations_env, action):
    get_sync_translations_env['Translations']['B2'] = action
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 2, 'created': 0, 'updated': 0,
        'deleted': 0, 'skipped': 2, 'errors': 0,
    }


def test_validate_translation_id(mocker, fs, get_sync_translations_env):
    get_sync_translations_env['Translations']['A3'] = None
    get_sync_translations_env['Translations']['B3'] = 'update'
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 2, 'created': 0, 'updated': 0,
        'deleted': 0, 'skipped': 1, 'errors': 1,
    }
    assert stats['Translations']._row_errors == {
        3: ['Translation ID is required to update or delete a translation'],
    }


def test_validate_autotranslation(mocker, fs, get_sync_translations_env):
    get_sync_translations_env['Translations']['B3'] = 'update'
    get_sync_translations_env['Translations']['I3'] = 'invalid'
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 2, 'created': 0, 'updated': 0,
        'deleted': 0, 'skipped': 1, 'errors': 1,
    }
    assert stats['Translations']._row_errors == {
        3: ['Autotranslation must be `Enabled` or `Disabled`. Provided invalid'],
    }


def test_validate_locale(mocker, fs, get_sync_translations_env):
    get_sync_translations_env['Translations']['B3'] = 'create'
    get_sync_translations_env['Translations']['G3'] = None
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 2, 'created': 0, 'updated': 0,
        'deleted': 0, 'skipped': 1, 'errors': 1,
    }
    assert stats['Translations']._row_errors == {
        3: ['Locale is required to create a translation'],
    }


def test_delete_translation(mocker, fs, get_sync_translations_env, mocked_responses):
    get_sync_translations_env['Translations']['B3'] = 'delete'
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )
    mocked_responses.add(
        method='DELETE',
        url='https://localhost/public/v1/localization/translations/TRN-1079-0833-9891',
        status=204,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 2, 'created': 0, 'updated': 0,
        'deleted': 1, 'skipped': 1, 'errors': 0,
    }


def test_delete_translation_not_found_is_ok(mocker, fs, get_sync_translations_env, mocked_responses):
    get_sync_translations_env['Translations']['B3'] = 'delete'
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )
    mocked_responses.add(
        method='DELETE',
        url='https://localhost/public/v1/localization/translations/TRN-1079-0833-9891',
        status=404,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 2, 'created': 0, 'updated': 0,
        'deleted': 1, 'skipped': 1, 'errors': 0,
    }


def test_delete_translation_fails(mocker, fs, get_sync_translations_env, mocked_responses):
    get_sync_translations_env['Translations']['B3'] = 'delete'
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )
    mocked_responses.add(
        method='DELETE',
        url='https://localhost/public/v1/localization/translations/TRN-1079-0833-9891',
        status=500,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 2, 'created': 0, 'updated': 0,
        'deleted': 0, 'skipped': 1, 'errors': 1,
    }
    assert stats['Translations']._row_errors == {
        3: ['500 Internal Server Error'],
    }


def test_update_translation(
    mocker, fs, get_sync_translations_env, mocked_new_translation_response, mocked_responses,
):
    get_sync_translations_env['Translations']['B3'] = 'update'
    get_sync_translations_env['Translations']['H3'] = 'la nueva descripción'
    get_sync_translations_env['Translations']['I3'] = 'Disabled'
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )
    response = deepcopy(mocked_new_translation_response)
    response['id'] = 'TRN-1079-0833-9891'
    mocked_responses.add(
        method='PUT',
        url='https://localhost/public/v1/localization/translations/TRN-1079-0833-9891',
        status=200,
        json=response,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 2, 'created': 0, 'updated': 1,
        'deleted': 0, 'skipped': 1, 'errors': 0,
    }


def test_create_translation(
    mocker, fs, get_sync_translations_env, mocked_new_translation_response, mocked_responses,
):
    get_sync_translations_env['Translations']['B4'] = 'create'
    get_sync_translations_env['Translations']['G4'] = 'JA (Japanese)'
    get_sync_translations_env['Translations']['H4'] = 'This is the japanese translation'
    get_sync_translations_env['Translations']['I4'] = 'Enabled'
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/localization/contexts?eq(instance_id,PRD-276-377-545)&limit=1&offset=0',
        headers={
            'Content-Range': 'items 0-0/1',
        },
        json=[{
            'id': 'LCX-1111-2222-3333',
            'instance_id': 'PRD-276-377-545', 'name': 'My product',
        }],
    )
    mocked_responses.add(
        method='POST',
        url='https://localhost/public/v1/localization/translations',
        status=201,
        match=[
            responses.matchers.json_params_matcher({
                "auto": {"enabled": True}, "context": {"id": "LCX-1111-2222-3333"},
                "description": "This is the japanese translation", "locale": {"id": "JA"},
            }),
        ],
        json=mocked_new_translation_response,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()
    synchronizer.save(f'{fs.root_path}/test.xlsx')

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 3, 'created': 1, 'updated': 0,
        'deleted': 0, 'skipped': 2, 'errors': 0,
    }
    wb = load_workbook(f'{fs.root_path}/test.xlsx')
    assert wb['Translations']['A4'].value == 'TRN-1079-0833-9999'
    assert wb['Translations']['C4'].value == 'PRD-276-377-545'
    assert wb['Translations']['E4'].value == 'VA-392-495'
    assert wb['Translations']['K4'].value == 'inactive'
    assert wb['Translations']['L4'].value == 'No'
    assert wb['Translations']['M4'].value == '2020-10-29T12:00:00+00:00'
    assert wb['Translations']['N4'].value == '2020-10-29T12:00:00+00:00'


def test_create_translation_locale_not_autotranslation_error(
    mocker, fs, get_sync_translations_env, mocked_responses,
):
    get_sync_translations_env['Translations']['B4'] = 'create'
    get_sync_translations_env['Translations']['G4'] = 'ES-AR (Argentinian Spanish)'
    get_sync_translations_env['Translations']['H4'] = "Argentinian can't be autotranslated!"
    get_sync_translations_env['Translations']['I4'] = 'Enabled'
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/localization/contexts?eq(instance_id,PRD-276-377-545)&limit=1&offset=0',
        headers={
            'Content-Range': 'items 0-0/1',
        },
        json=[{
            'id': 'LCX-1111-2222-3333',
            'instance_id': 'PRD-276-377-545', 'name': 'My product',
        }],
    )
    mocked_responses.add(
        method='POST',
        url='https://localhost/public/v1/localization/translations',
        status=400,
        json={"error_code": "VAL_001", "errors": ["Locale is not available for autotranslation."]},
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 3, 'created': 0, 'updated': 0,
        'deleted': 0, 'skipped': 2, 'errors': 1,
    }
    assert stats['Translations']._row_errors == {
        4: ["400 Bad Request: VAL_001 - Locale is not available for autotranslation."],
    }


def test_create_translation_get_context_error(mocker, fs, get_sync_translations_env, mocked_responses):
    get_sync_translations_env['Translations']['B4'] = 'create'
    get_sync_translations_env['Translations']['G4'] = 'JA (Japanese)'
    get_sync_translations_env['Translations']['H4'] = 'This is the japanese translation'
    get_sync_translations_env['Translations']['I4'] = 'Enabled'
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/localization/contexts?eq(instance_id,PRD-276-377-545)&limit=1&offset=0',
        status=500,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 3, 'created': 0, 'updated': 0,
        'deleted': 0, 'skipped': 2, 'errors': 1,
    }
    assert stats['Translations']._errors == ['500 Internal Server Error']


def test_several_actions_order_is_ok(mocker, fs, mocked_new_translation_response, mocked_responses_ordered):
    wb = load_workbook('./tests/fixtures/translations_sync.xlsx')
    wb['Translations']['B2'] = 'update'
    wb['Translations']['H2'] = 'new description'
    wb['Translations']['L2'] = 'No'
    wb['Translations']['B3'] = 'delete'
    wb['Translations']['B4'] = 'create'
    wb['Translations']['G4'] = 'JA (Japanese)'
    wb['Translations']['H4'] = 'New japanese translation'
    wb['Translations']['I4'] = 'Enabled'
    wb['Translations']['A5'] = 'TRN-1079-0833-9897'
    wb['Translations']['B5'] = 'delete'
    wb.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )
    mocked_responses_ordered.add(
        method='GET',
        url='https://localhost/public/v1/products/PRD-276-377-545',
        json={"id": "PRD-276-377-545"},
    )
    mocked_responses_ordered.add(
        method='DELETE',
        url='https://localhost/public/v1/localization/translations/TRN-1079-0833-9891',
        status=204,
    )
    mocked_responses_ordered.add(
        method='DELETE',
        url='https://localhost/public/v1/localization/translations/TRN-1079-0833-9897',
        status=204,
    )
    response = deepcopy(mocked_new_translation_response)
    response['id'] = 'TRN-1079-0833-9890'
    mocked_responses_ordered.add(
        method='PUT',
        url='https://localhost/public/v1/localization/translations/TRN-1079-0833-9890',
        status=200,
        json=mocked_new_translation_response,
    )
    mocked_responses_ordered.add(
        method='GET',
        url='https://localhost/public/v1/localization/contexts?eq(instance_id,PRD-276-377-545)&limit=1&offset=0',
        headers={
            'Content-Range': 'items 0-0/1',
        },
        json=[{
            'id': 'LCX-1111-2222-3333',
            'instance_id': 'PRD-276-377-545', 'name': 'My product',
        }],
    )
    mocked_responses_ordered.add(
        method='POST',
        url='https://localhost/public/v1/localization/translations',
        status=201,
        json=mocked_new_translation_response,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()

    assert stats['Translations'].get_counts_as_dict() == {
        'processed': 4, 'created': 1, 'updated': 1,
        'deleted': 2, 'skipped': 0, 'errors': 0,
    }


def test_locales_validation_still_present_after_update(mocker, fs, get_sync_translations_env):
    get_sync_translations_env['Translations']['B3'] = 'update'
    get_sync_translations_env.save(f'{fs.root_path}/test.xlsx')

    stats = SynchronizerStats()
    synchronizer = TranslationsSynchronizer(
        client=ConnectClient(
            use_specs=False,
            api_key='ApiKey SU:123',
            endpoint='https://localhost/public/v1',
        ),
        progress=mocker.MagicMock(),
        stats=stats,
    )

    synchronizer.open(f'{fs.root_path}/test.xlsx', 'Translations')
    synchronizer.sync()
    synchronizer.save(f'{fs.root_path}/test.xlsx')

    wb = load_workbook(f'{fs.root_path}/test.xlsx')
    assert any(
        (
            data_validation.formula1 == "'General Information'!$AB$2:$AB$98"
            and data_validation.sqref.ranges[0].coord == 'G2:G3'
        ) for data_validation in wb['Translations'].data_validations.dataValidation
    )

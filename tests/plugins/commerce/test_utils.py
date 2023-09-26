import json
import os
import tempfile

import pytest
from click import ClickException
from connect.client import ConnectClient, ClientError
from openpyxl import Workbook, load_workbook

from connect.cli.plugins.commerce.utils import (
    _validate_header,
    clone_stream,
    clone_transformations,
    create_stream_from_origin,
    fill_and_download_attachments,
    get_work_book,
    guess_if_billing_or_pricing_stream,
    update_attachments,
    update_general_information,
    update_transformations,
)


def test_guess_if_billing_or_pricing_billing(mocked_responses, client):
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams?eq(id,STR-123)&limit=0&offset=0',
        json=[],
        headers={
            'Content-Range': 'items 0-1/1',
        },
    )
    assert guess_if_billing_or_pricing_stream(client, 'STR-123') == 'billing'


def test_guess_if_billing_or_pricing_pricing(mocked_responses, client):
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams?eq(id,STR-567)&limit=0&offset=0',
        json=[],
        headers={
            'Content-Range': 'items 0-0/0',
        },
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/pricing/streams?eq(id,STR-123)&limit=0&offset=0',
        json=[],
        headers={
            'Content-Range': 'items 0-1/1',
        },
    )
    assert guess_if_billing_or_pricing_stream(client, 'STR-123') == 'pricing'


def test_guess_if_billing_or_pricing_none(mocked_responses, client):
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams?eq(id,STR-567)&limit=0&offset=0',
        json=[],
        headers={
            'Content-Range': 'items 0-0/0',
        },
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/pricing/streams?eq(id,STR-123)&limit=0&offset=0',
        json=[],
        headers={
            'Content-Range': 'items 0-0/0',
        },
    )
    assert guess_if_billing_or_pricing_stream(client, 'STR-123') is None


def test_fill_attachments_download_fails(mocked_responses, mocker, client):
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/media/folders/ftype/fname/files/id',
        status=500,
    )
    with tempfile.TemporaryDirectory() as tmpdir:
        wb = Workbook()
        ws = wb.create_sheet('Attachments')
        mocked_progress = mocker.MagicMock()
        with pytest.raises(ClickException) as exc:
            fill_and_download_attachments(
                ws,
                [
                    {
                        'id': 'id',
                        'name': 'name',
                        'folder': {'type': 'ftype', 'name': 'fname'},
                    },
                ],
                client,
                os.path.join(tmpdir, 'media'),
                mocked_progress,
            )
            assert exec.value == 'Error obtaining file name'


def test_clone_stream_not_found(
    mocker,
    mocked_responses,
    client,
):
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams?eq(id,STR-7755-7115-2464)&limit=0&offset=0',
        json=[],
        headers={
            'Content-Range': 'items 0-0/0',
        },
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/pricing/streams?eq(id,STR-7755-7115-2464)&limit=0&offset=0',
        json=[],
        headers={
            'Content-Range': 'items 0-0/0',
        },
    )

    with pytest.raises(ClickException) as e:
        clone_stream(
            mocker.MagicMock(client=client, id='VA-000'),
            'STR-7755-7115-2464',
            None,
        )

    assert str(e.value) == 'Stream STR-7755-7115-2464 not found for the current account VA-000.'


def test_clone_stream_computed_and_diff_accounts(
    mocker,
    mocked_responses,
    client,
):
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams?eq(id,STR-7755-7115-2464)&limit=0&offset=0',
        json=[],
        headers={
            'Content-Range': 'items 0-1/1',
        },
    )
    with open('./tests/fixtures/commerce/stream_retrieve_response.json') as content:
        response = json.load(content)[0]
        response['sources'] = {'some': 'thing'}
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/billing/streams?eq(id,STR-7755-7115-2464)&select(context,samples,sources)&limit=1&offset=0',
            json=[response],
        )

    with pytest.raises(ClickException) as e:
        clone_stream(
            mocker.MagicMock(client=client, id='VA-000'),
            'STR-7755-7115-2464',
            mocker.MagicMock(client=client, id='VA-001'),
        )

    assert str(e.value) == 'You cannot clone a Computed stream between different accounts.'


def test_clone_inbound_stream(
    mocker,
    mocked_responses,
    client,
):
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams?eq(id,STR-7755-7115-2464)&limit=0&offset=0',
        json=[],
        headers={
            'Content-Range': 'items 0-1/1',
        },
    )
    with open('./tests/fixtures/commerce/stream_retrieve_response.json') as content:
        response = json.load(content)[0]
        response['owner']['id'] = 'VA-999'
        mocked_responses.add(
            method='GET',
            url='https://localhost/public/v1/billing/streams?eq(id,STR-7755-7115-2464)&select(context,samples,sources)&limit=1&offset=0',
            json=[response],
        )

    with pytest.raises(ClickException) as e:
        clone_stream(
            mocker.MagicMock(client=client, id='VA-000'),
            'STR-7755-7115-2464',
            mocker.MagicMock(client=client, id='VA-001'),
        )

    assert str(e.value) == 'Inbound streams cannot be cloned.'


def test_create_stream_from_origin_error(
    mocked_responses,
    client,
):
    mocked_responses.add(
        method='POST',
        url='https://localhost/public/v1/billing/streams',
        status=400,
        json={'error_code': 'error_code', 'errors': ['error one']},
    )
    with pytest.raises(ClickException) as e:
        create_stream_from_origin(
            client,
            {
                'id': 'STR-123',
                'context': {},
                'name': 'name',
                'sources': [],
                'owner': {'id': 'VA-000'},
            },
            'billing',
        )
    assert str(e.value) == 'error one'


def test_clone_transformations_with_two_columns_same_name(
    mocker,
):
    mocked_col_mapping = mocker.patch(
        'connect.cli.plugins.commerce.utils.generate_column_mapping',
        return_value=({'name': ['COL-ID-1', 'COL-ID-2']}, {'COL-ID-1': '', 'COL-ID-2': ''}),
    )
    mocked_progress = mocker.MagicMock()
    clone_transformations(
        mocker.MagicMock(),
        'STR-333',
        'billing',
        {},
        [
            {
                'function': {'id': 'fn-123', 'name': 'Lookup Data from a stream attached Excel'},
                'settings': {},
                'overview': 'overview',
                'position': 1000,
                'columns': {'input': [{'id': 'COL-ID-2', 'name': 'name'}], 'output': []},
            },
        ],
        {'name': ['COL-ID-1', 'COL-ID-2']},
        mocked_progress,
    )
    mocked_col_mapping.assert_called()
    mocked_progress.add_task.assert_called_with('Processing transformations', total=1)
    assert mocked_progress.update.call_count == 1


def test_clone_transformations_fix_file_mapping(
    mocker,
    client,
    mocked_responses,
):
    mocked_responses.add(
        method='POST',
        url='https://localhost/public/v1/billing/streams/STR-333/transformations',
    )
    mocked_col_mapping = mocker.patch(
        'connect.cli.plugins.commerce.utils.generate_column_mapping',
        return_value=({'name': ['COL-ID-1', 'COL-ID-2']}, {'COL-ID-1': '', 'COL-ID-2': ''}),
    )
    mocked_progress = mocker.MagicMock()
    clone_transformations(
        client,
        'STR-333',
        'billing',
        {'filepath': 'newfilepath'},
        [
            {
                'function': {'id': 'fn-123', 'name': 'Lookup Data from a stream attached Excel'},
                'settings': {'file': 'filepath'},
                'overview': 'overview',
                'position': 1000,
                'columns': {'input': [{'id': 'COL-ID-2', 'name': 'name'}], 'output': []},
            },
        ],
        {'name': ['COL-ID-1', 'COL-ID-2']},
        mocked_progress,
    )
    mocked_col_mapping.assert_called()
    mocked_progress.add_task.assert_called_with('Processing transformations', total=1)
    assert mocked_progress.update.call_count == 1


def test_get_work_book():
    wb = get_work_book('./tests/fixtures/commerce/stream_sync.xlsx')

    general_info = wb['General Information']
    assert general_info['B2'].value == 'STR-7748-7021-7449'

    wb.close()


def test_get_work_book_invalid_sheets():
    with pytest.raises(ClickException) as ce:
        get_work_book('./tests/fixtures/actions_sync.xlsx')

    assert str(ce.value) == (
        'The file must contain `General Information`, `Columns`, `Transformations` and '
        '`Attachments` sheets.'
    )


def test_get_work_book_non_existing_file():
    with pytest.raises(ClickException) as ce:
        get_work_book('non_existing_file.xlsx')

    assert str(ce.value) == 'The file non_existing_file.xlsx does not exists.'


def test_get_work_book_invalid_format():
    with pytest.raises(ClickException) as ce:
        get_work_book('./tests/fixtures/image.png')

    assert str(ce.value) == 'The file ./tests/fixtures/image.png has invalid format, must be xlsx.'


def test_validate_header_missing():
    with pytest.raises(ClickException) as ce:
        _validate_header(['ID', 'Name'], ['ID', 'Name', 'Description'], 'Streams')

    assert str(ce.value) == 'The Streams sheet header does not contain `Description` header.'


def test_update_general_information_error(sample_stream_workbook, mocker, mocked_responses):
    client = ConnectClient(
        api_key='ApiKey X',
        endpoint='https://localhost/public/v1',
        use_specs=False,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams?'
        'eq(id,STR-7755-7115-2464)&select(context,samples,sources)&limit=1&offset=0',
        json=[{'name': 'Name', 'description': 'Description'}],
    )

    mocked_responses.add(
        method='PUT',
        url='https://localhost/public/v1/billing/streams/STR-7755-7115-2464',
        status=400,
    )

    results = []
    update_general_information(
        client=client,
        collection='billing',
        stream_id='STR-7755-7115-2464',
        sheet=sample_stream_workbook['General Information'],
        results=results,
        errors=[],
        progress=mocker.MagicMock(),
    )

    sample_stream_workbook.close()
    assert len(results) == 1
    assert results[0] == ('General information', 5, 0, 0, 0, 0, 5)


def test_update_transformations(sample_stream_workbook, mocker, mocked_responses):
    client = ConnectClient(
        api_key='ApiKey X',
        endpoint='https://localhost/public/v1',
        use_specs=False,
    )

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams/STR-7755-7115-2464/transformations',
        json=[
            {'id': 'STRA-774-870-217-449-001'},
            {'id': 'STRA-774-870-217-449-002'},
        ],
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams/STR-7755-7115-2464'
        '/transformations/STRA-774-870-217-449-001',
        json={
            'id': 'STRA-774-870-217-449-001',
            'settings': {
                'from': 'id',
                'regex': {
                    'groups': {'1': {'name': 'first_name', 'type': 'string'}},
                    'pattern': '(?P<first_name>\\w+)',
                },
            },
            'description': 'Old description',
            'position': 10000,
        },
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams/STR-7755-7115-2464'
        '/transformations/STRA-774-870-217-449-002',
        json={
            'id': 'STRA-774-870-217-449-002',
            'settings': {
                'additional_values': [],
                'from': 'position',
                'match_condition': True,
                'value': '200',
            },
            'description': 'edeeasd',
            'position': 20000,
        },
    )
    mocked_responses.add(
        method='PUT',
        url='https://localhost/public/v1/billing/streams/STR-7755-7115-2464'
        '/transformations/STRA-774-870-217-449-001',
        status=200,
    )

    results = []
    errors = []
    update_transformations(
        client=client,
        collection='billing',
        stream_id='STR-7755-7115-2464',
        sheet=sample_stream_workbook['Transformations'],
        results=results,
        errors=errors,
        progress=mocker.MagicMock(),
    )

    assert len(errors) == 0
    assert len(results) == 1
    assert results[0] == ('Transformations', 2, 0, 1, 0, 0, 0)


def test_update_transformations_not_exists(sample_stream_workbook, mocker, mocked_responses):
    client = ConnectClient(
        api_key='ApiKey X',
        endpoint='https://localhost/public/v1',
        use_specs=False,
    )

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams/STR-7755-7115-2464'
        '/transformations/STRA-774-870-217-449-001',
        status=404,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams/STR-7755-7115-2464'
        '/transformations/STRA-774-870-217-449-002',
        status=404,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams/STR-7755-7115-2464/transformations',
        json=[],
    )

    results = []
    errors = []
    update_transformations(
        client=client,
        collection='billing',
        stream_id='STR-7755-7115-2464',
        sheet=sample_stream_workbook['Transformations'],
        results=results,
        errors=errors,
        progress=mocker.MagicMock(),
    )

    assert len(errors) == 2
    assert (
        'The transformation STRA-774-870-217-449-001'
        ' cannot be updated because it does not exist.'
    ) in errors
    assert (
        'The transformation STRA-774-870-217-449-002'
        ' cannot be updated because it does not exist.'
    ) in errors

    assert results[0] == ('Transformations', 2, 0, 0, 0, 0, 2)


def test_update_transformations_with_errors(sample_stream_workbook, mocker, mocked_responses):
    client = ConnectClient(
        api_key='ApiKey X',
        endpoint='https://localhost/public/v1',
        use_specs=False,
    )

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams/STR-7755-7115-2464'
        '/transformations/STRA-774-870-217-449-001',
        json={
            'id': 'STRA-774-870-217-449-001',
            'settings': {
                'from': 'id',
                'regex': {
                    'groups': {'1': {'name': 'first_name', 'type': 'string'}},
                    'pattern': '(?P<first_name>\\w+)',
                },
            },
            'description': 'Old description',
            'position': 10000,
        },
    )
    mocked_responses.add(
        method='PUT',
        url='https://localhost/public/v1/billing/streams/STR-7755-7115-2464'
        '/transformations/STRA-774-870-217-449-001',
        status=400,
    )

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams/STR-7755-7115-2464'
        '/transformations/STRA-774-870-217-449-002',
        status=404,
    )

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/billing/streams/STR-7755-7115-2464/transformations',
        json=[{'id': 'STRA-774-870-217-449-003'}],
    )
    mocked_responses.add(
        method='DELETE',
        url='https://localhost/public/v1/billing/streams/STR-7755-7115-2464'
        '/transformations/STRA-774-870-217-449-003',
        status=500,
    )

    results = []
    errors = []
    update_transformations(
        client=client,
        collection='billing',
        stream_id='STR-7755-7115-2464',
        sheet=sample_stream_workbook['Transformations'],
        results=results,
        errors=errors,
        progress=mocker.MagicMock(),
    )

    assert len(errors) == 3
    assert results[0] == ('Transformations', 2, 0, 0, 0, 0, 3)


def test_update_attachments_with_errors(sample_stream_workbook, mocker, mocked_responses):
    client = ConnectClient(
        api_key='ApiKey X',
        endpoint='https://localhost/public/v1',
        use_specs=False,
    )

    mocker.patch('connect.cli.plugins.commerce.utils.upload_attachment', side_effect=ClientError())

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/media/folders/streams_attachments/STR-7755-7115-2464'
        '/files?eq(id,ID)&limit=0&offset=0',
        headers={'Content-Range': 'items 0-0/0'},
        status=200,
    )
    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/media/folders/streams_attachments/STR-7755-7115-2464'
        '/files?eq(id,ID-EXISTS)&limit=0&offset=0',
        headers={'Content-Range': 'items 0-0/0'},
        status=200,
    )

    mocked_responses.add(
        method='GET',
        url='https://localhost/public/v1/media/folders/streams_attachments/'
        'STR-7755-7115-2464/files',
        json=[{'id': 'MLF-123'}],
    )
    mocked_responses.add(
        method='DELETE',
        url='https://localhost/public/v1/media/folders/streams_attachments/STR-7755-7115-2464'
        '/files/MLF-123',
        status=400,
    )

    results = []
    errors = []
    update_attachments(
        client=client,
        stream_id='STR-7755-7115-2464',
        sheet=sample_stream_workbook['Attachments'],
        results=results,
        errors=errors,
        progress=mocker.MagicMock(),
    )

    assert len(errors) == 3
    assert results[0] == ('Attachment files', 2, 0, 0, 0, 0, 3)

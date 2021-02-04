import os
import sys
import json

from datetime import datetime, timezone
from importlib import import_module

import click
from click import ClickException

from interrogatio import dialogus
from tqdm import tqdm

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import Color, WHITE
from cnct import ConnectClient
from cmr import render

from cnctcli.actions.products.constants import DEFAULT_BAR_FORMAT
from cnctcli.constants import REPORT_PARAM_TYPES
from cnctcli.actions.reports_params import (
    date_params,
    dynamic_params,
    static_params,
)

from cnct import ClientError


def get_report_entrypoint(func_fqn):
    module_name, func_name = func_fqn.rsplit('.', 1)
    try:
        module = import_module(module_name)
        return getattr(module, func_name)
    except ImportError:
        pass
    except AttributeError:
        pass


def get_report_id(func_fqn):
    tokens = func_fqn.split('.')
    if len(tokens) < 3:
        raise ClickException("Reports project does not conform with specification")
    return tokens[1]


def handle_param_input(config, client, param):
    questions = []
    if date_params.get(param['type']):
        handler = date_params[param['type']]
        questions.append(
            handler(param)
        )
    elif dynamic_params.get(param['type']):
        handler = dynamic_params[param['type']]
        questions.append(
            handler(config.active, client, param)
        )
    elif static_params.get(param['type']):
        handler = static_params[param['type']]
        questions.append(
            handler(param)
        )
    else:
        raise ClickException(f'Unknown parameter type {param["type"]}')

    return questions


def _convert_to_utc_input(date_string):
    date = datetime.strptime(date_string, '%Y-%m-%d')
    return date.astimezone().astimezone(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S')


def get_report_inputs(config, client, parameters):
    parameters_values = {}
    i = 0
    for param in parameters:
        questions = handle_param_input(config, client, param)

        answers = dialogus(
            questions,
            param['name'],
            confirm='Run' if i == len(parameters) else 'Next',
        )

        if not answers:
            raise ClickException("Aborted by user input")

        if param['type'] == 'date_range':
            if answers[param["id"]]["from"]:
                after_str = answers[param["id"]]["from"]
                before_str = answers[param["id"]]["to"]
                after = _convert_to_utc_input(after_str)
                before = _convert_to_utc_input(before_str)
                parameters_values[param['id']] = {
                    'after': after,
                    'before': before,
                }
            continue
        elif questions[0]['type'] == 'selectmany' and \
                len(questions[0]['values']) == len(answers[param['id']]):
            parameters_values[param['id']] = {
                "all": True,
                "choices": [],
            }
        elif questions[0]['type'] == 'selectmany' and \
                len(questions[0]['values']) != len(answers[param['id']]):
            parameters_values[param['id']] = {
                "all": False,
                "choices": answers[param['id']],
            }
        else:
            parameters_values[param['id']] = answers[param['id']]
        i += 1
    return parameters_values


def execute_report(config, reports_dir, report, output_file):
    sys.path.append(reports_dir)
    validate_report_definition(report, reports_dir)

    if config.active.id.startswith('VA') and 'vendor' not in report['audience']:
        raise ClickException(
            "This report is not expected to be executed on vendor accounts"
        )

    if config.active.id.startswith('PA') and 'provider' not in report['audience']:
        raise ClickException(
            "This report is not expected to be executed on provider accounts"
        )

    client = ConnectClient(
        config.active.api_key,
        endpoint=config.active.endpoint,
        use_specs=False,
        default_limit=10,
    )

    entrypoint = get_report_entrypoint(report['entrypoint'])
    report_id = get_report_id(report['entrypoint'])
    inputs = get_report_inputs(config, client, report['parameters'])

    template_workbook = load_workbook(
        os.path.join(
            reports_dir,
            report['template'],
        ),
    )

    ws = template_workbook['Data']

    row_idx = report['start_row']
    start_col_idx = report['start_col']

    class Progress(tqdm):
        def __init__(self):
            super().__init__()
            self.desc = f'Processing report {report["name"]}...'
            self.leave = True
            self.bar_format = DEFAULT_BAR_FORMAT
            self.current_value = 0
            self.monitor_interval = 1
            self.miniters = 1

        def __call__(self, value, max_value):
            self.total = max_value
            self.update(value - self.current_value)
            self.current_value = value

    click.echo(f'Preparing to run report {report_id}. Please wait...\n')

    progress = Progress()

    start_time = datetime.now().isoformat()
    try:
        for row in entrypoint(client, inputs, progress):
            for col_idx, cell_value in enumerate(row, start=start_col_idx):
                ws.cell(row_idx, col_idx, value=cell_value)
            row_idx += 1
    except (ClientError, Exception) as e:
        if isinstance(e, ClientError):
            message = f'\nError returned by Connect when executing the report: {str(e)}'
        elif isinstance(e, RuntimeError):
            message = f'\nReport error: {str(e)}'
        else:
            message = f'\nUnknown error while executing the report: {str(e)}'
        trace = []
        tb = e.__traceback__
        while tb is not None:
            trace.append({
                "filename": tb.tb_frame.f_code.co_filename,
                "name": tb.tb_frame.f_code.co_name,
                "lineno": tb.tb_lineno
            })
            tb = tb.tb_next
        trace_string = json.dumps(trace, sort_keys=True, indent=4)
        raise ClickException(
            f'{message}\n'
            f'Trace: {trace_string}'
        )

    add_info_sheet(
        template_workbook.create_sheet('Info'),
        config,
        report,
        inputs,
        start_time,
        report_id,
    )
    template_workbook.save(output_file)
    progress.close()


def add_info_sheet(ws, config, report, report_values, start_time, report_id):
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 180
    ws['A1'].value = "Report Execution Information"
    ws.merge_cells('A1:B1')
    ws['A1'].fill = PatternFill('solid', start_color=Color('1565C0'))
    ws['A1'].font = Font(sz=24, color=WHITE)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    for i in range(2, 10):
        ws[f'A{i}'].alignment = Alignment(
            horizontal='left',
            vertical='top',
        )
        ws[f'B{i}'].alignment = Alignment(
            horizontal='left',
            vertical='top',
        )
    ws['A2'].value = "Report Start time"
    ws['B2'].value = start_time
    ws['A3'].value = "Report Finish time"
    ws['B3'].value = datetime.now().isoformat()
    ws['A4'].value = "Account ID"
    ws['B4'].value = config.active.id
    ws['A5'].value = "Account Name"
    ws['B5'].value = config.active.name
    ws['A6'].value = "Report ID"
    ws['B6'].value = report_id
    ws['A7'].value = "Report Name"
    ws['B7'].value = report['name']
    ws['A8'].value = "Runtime environment"
    ws['B8'].value = "CLI Tool"
    ws['A9'].value = "Report execution paramters"
    ws['B9'].value = json.dumps(report_values, indent=4, sort_keys=True)
    ws['B9'].alignment = Alignment(
        horizontal='left',
        vertical='top',
        wrap_text=True,
    )


def get_report_description(reports_dir, readme_file):
    with open(
        os.path.join(
            reports_dir,
            readme_file
        )
    ) as file:
        readme = file.read()
        file.close()
        return render(readme)


def validate_report_json(descriptor, reports_dir):
    if 'name' not in descriptor:
        raise ClickException('Name property is required for reports.json')
    if 'readme_file' not in descriptor:
        raise ClickException('Description property is required for reports.json')
    if 'version' not in descriptor:
        raise ClickException('Version property is required for reports.json')
    if 'reports' not in descriptor:
        raise ClickException('"No reports declared in reports.json')
    for report in descriptor['reports']:
        validate_report_definition(report, reports_dir)
    validate_entry_points(descriptor['reports'])


def validate_report_definition(definition, reports_dir):
    if 'name' not in definition:
        raise ClickException(
            'Property name not found in report'
        )

    required_properties = [
        'name',
        'readme_file',
        'template',
        'start_row',
        'start_col',
        'entrypoint',
        'report_spec',
        'audience',
    ]
    for required_prop in required_properties:
        if required_prop not in definition:
            raise ClickException(
                f'Property {required_prop} not found for report {definition["name"]}'
            )
    if definition['report_spec'] != "1":
        raise ClickException(
            'Supported report specification by Connect CLI tool is 1'
        )
    if 'parameters' not in definition:
        raise ClickException(
            f'Missing parameters list for report {definition["name"]}'
        )
    for param in definition['parameters']:
        validate_report_parameters(param, definition['name'])
    if not os.path.isfile(os.path.join(
        reports_dir,
        definition['template']
    )):
        raise ClickException(
            f'Template {definition["template"]} not found on {reports_dir}'
        )
    if not os.path.isfile(os.path.join(
        reports_dir,
        definition['readme_file']
    )):
        raise ClickException(
            f'Template {definition["readme_file"]} not found on {reports_dir}'
        )


def validate_report_parameters(report_parameters, report_name):
    if 'id' not in report_parameters:
        raise ClickException(
            f'Missing id on parameters for report {report_name}'
        )
    if 'type' not in report_parameters:
        raise ClickException(
            f'Missing type for input parameter {report_parameters["id"]} on report {report_name}'
        )
    if report_parameters['type'] == 'choice' or report_parameters['type'] == 'checkbox':
        validate_choices(report_parameters, report_name)

    if 'name' not in report_parameters:
        raise ClickException(
            f'Missing name for input parameter {report_parameters["id"]} on report {report_name}'
        )
    if 'description' not in report_parameters:
        raise ClickException(
            f'Missing description for input parameter {report_parameters["id"]} '
            f'on report {report_name}'
        )
    if report_parameters['type'] not in REPORT_PARAM_TYPES:
        raise ClickException(
            f'Report {report_name} has a unknown parameter type {report_parameters["type"]} '
            f'defined for parameter {report_parameters["id"]}'
        )


def validate_choices(param, report_name):
    if 'choices' not in param:
        raise ClickException(
            f'Missing choices for parameter {param["id"]} on report {report_name}'
        )
    for choice in param['choices']:
        if 'value' not in choice:
            raise ClickException(
                f'Missing value in one of the choices for {param["id"]} on report {report_name}'
            )
        if 'label' not in choice:
            raise ClickException(
                f'Missing label in one of the choices for {param["id"]} on report {report_name}'
            )


def validate_entry_points(reports):
    entry_points = []
    for report in reports:
        if report['entrypoint'] not in entry_points:
            entry_points.append(report['entrypoint'])
        else:
            raise ClickException("There is a duplicated entrypoint on reports.json")

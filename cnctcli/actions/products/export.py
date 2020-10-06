# -*- coding: utf-8 -*-

# This file is part of the Ingram Micro Cloud Blue Connect connect-cli.
# Copyright (c) 2019-2020 Ingram Micro. All Rights Reserved.

import os
from datetime import datetime

from click import ClickException

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.colors import Color, WHITE
from openpyxl.worksheet.datavalidation import DataValidation

from tqdm import trange

from cnctcli.actions.products.constants import ITEMS_COLS_HEADERS
from cnctcli.api.products import get_items, get_product


def _setup_cover_sheet(ws, product):
    ws.title = 'product_info'
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 50
    ws.merge_cells('A1:B1')
    cell = ws['A1']
    cell.fill = PatternFill('solid', start_color=Color('1565C0'))
    cell.font = Font(sz=24, color=WHITE)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value = 'Product information'
    for i in range(3, 9):
        ws[f'A{i}'].font = Font(sz=14)
        ws[f'B{i}'].font = Font(sz=14, bold=True)
    ws['A3'].value = 'Account ID'
    ws['B3'].value = product['owner']['id']
    ws['A4'].value = 'Account Name'
    ws['B4'].value = product['owner']['name']
    ws['A5'].value = 'Product ID'
    ws['B5'].value = product['id']
    ws['A6'].value = 'Product Name'
    ws['B6'].value = product['name']
    ws['A7'].value = 'Export datetime'
    ws['B7'].value = datetime.now().isoformat()


def _setup_items_header(ws):
    color = Color('d3d3d3')
    fill = PatternFill('solid', color)
    cels = ws['A1': 'M1']
    for cel in cels[0]:
        ws.column_dimensions[cel.column_letter].width = 25
        ws.column_dimensions[cel.column_letter].auto_size = True
        cel.fill = fill
        cel.value = ITEMS_COLS_HEADERS[cel.column_letter]


def _calculate_commitment(item):
    period = item.get('period')
    if not period:
        return '-'
    commitment = item.get('commitment')
    if not commitment:
        return '-'
    count = commitment['count']
    if count == 1:
        return '-'

    multiplier = commitment['multiplier']

    if multiplier == 'billing_period':
        if period == 'monthly':
            years = count // 12
            return '{} year{}'.format(
                years,
                's' if years > 1 else '',
            )
        else:
            return '{} years'.format(count)

    # One-time
    return '-'


def _fill_item_row(ws, row_idx, item):
    ws.cell(row_idx, 1, value=item['id'])
    ws.cell(row_idx, 2, value=item['mpn'])
    ws.cell(row_idx, 3, value='-')
    ws.cell(row_idx, 4, value=item['display_name'])
    ws.cell(row_idx, 5, value=item['description'])
    ws.cell(row_idx, 6, value=item['type'])
    ws.cell(row_idx, 7, value=item['precision'])
    ws.cell(row_idx, 8, value=item['unit']['unit'])
    period = item.get('period', 'monthly')
    if period.startswith('years_'):
        period = f"{period.rsplit('_')[-1]} years"
    ws.cell(row_idx, 9, value=period)
    ws.cell(row_idx, 10, value=_calculate_commitment(item))
    ws.cell(row_idx, 11, value=item['status'])
    ws.cell(row_idx, 12, value=item['events']['created']['at'])
    ws.cell(row_idx, 13, value=item['events'].get('updated', {}).get('at'))


def _dump_items(ws, api_url, api_key, product_id, silent):
    _setup_items_header(ws)

    processed_items = 0
    row_idx = 2
    limit = 2
    offset = 0

    count, items = get_items(api_url, api_key, product_id, limit, offset)

    if count == 0:
        raise ClickException(f"The product {product_id} doesn't have items.")

    action_validation = DataValidation(
        type='list',
        formula1='"-,create,update,delete"',
        allow_blank=False,
    )
    type_validation = DataValidation(
        type='list',
        formula1='"reservation,ppu"',
        allow_blank=False,
    )
    period_validation = DataValidation(
        type='list',
        formula1='"onetime,monthly,yearly,2 years,3 years,4 years,5 years"',
        allow_blank=False,
    )

    precision_validation = DataValidation(
        type='list',
        formula1='"integer,decimal(1),decimal(2),decimal(4),decimal(8)"',
        allow_blank=False,
    )

    commitment_validation = DataValidation(
        type='list',
        formula1='"-,1 year,2 years,3 years,4 years,5 years"',
        allow_blank=False,
    )

    ws.add_data_validation(action_validation)
    ws.add_data_validation(type_validation)
    ws.add_data_validation(period_validation)
    ws.add_data_validation(precision_validation)
    ws.add_data_validation(commitment_validation)

    items = iter(items)

    progress = trange(0, count, position=0, disable=silent)

    while True:
        try:
            item = next(items)
            progress.set_description(f"Processing item {item['id']}")
            progress.update(1)
            _fill_item_row(ws, row_idx, item)
            action_validation.add(f'C{row_idx}')
            type_validation.add(f'F{row_idx}')
            precision_validation.add(f'G{row_idx}')
            period_validation.add(f'I{row_idx}')
            commitment_validation.add(f'J{row_idx}')
            processed_items += 1
            row_idx += 1
        except StopIteration:
            if processed_items < count:
                offset += limit
                _, items = get_items(api_url, api_key, product_id, limit, offset)
                items = iter(items)
                continue
            break


def dump_product(api_url, api_key, product_id, output_file, silent):
    if not output_file:
        output_file = os.path.abspath(
            os.path.join('.', f'{product_id}.xlsx'),
        )

    product = get_product(api_url, api_key, product_id)
    wb = Workbook()
    _setup_cover_sheet(wb.active, product)

    _dump_items(wb.create_sheet('product_items'), api_url, api_key, product_id, silent)
    wb.save(output_file)

    return output_file

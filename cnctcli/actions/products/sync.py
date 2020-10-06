# -*- coding: utf-8 -*-

# This file is part of the Ingram Micro Cloud Blue Connect connect-cli.
# Copyright (c) 2019-2020 Ingram Micro. All Rights Reserved.

from collections import namedtuple
# from pprint import pprint
from zipfile import BadZipFile

from click import ClickException

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from tqdm import trange

from cnctcli.actions.products.constants import (
    BILLING_PERIOD,
    COMMITMENT,
    ITEMS_COLS_HEADERS,
    PRECISIONS,
)
from cnctcli.api.products import (
    create_item,
    create_unit,
    delete_item,
    get_item,
    get_item_by_mpn,
    get_product,
    get_units,
    update_item,
)

fields = (v.replace(' ', '_').lower() for v in ITEMS_COLS_HEADERS.values())

_RowData = namedtuple('RowData', fields)


class ProductSynchronizer:
    def __init__(self, endpoint, api_key, silent):
        self._endpoint = endpoint
        self._api_key = api_key
        self._silent = silent
        self._units = get_units(self._endpoint, self._api_key)
        self._product_id = None
        self._wb = None

    def open(self, input_file):
        self._open_workbook(input_file)
        if len(self._wb.sheetnames) != 2:
            raise ClickException('Invalid input file: not enough sheets.')
        ws = self._wb[self._wb.sheetnames[0]]
        product_id = ws['B5'].value
        get_product(self._endpoint, self._api_key, product_id)
        ws = self._wb[self._wb.sheetnames[1]]
        self._validate_item_sheet(ws)

        self._product_id = product_id
        return self._product_id

    def sync(self):
        ws = self._wb[self._wb.sheetnames[1]]
        errors = {}
        skipped_count = 0
        created_items = []
        updated_items = []
        deleted_items = []

        row_indexes = trange(2, ws.max_row + 1, position=0, disable=self._silent)
        for row_idx in row_indexes:
            data = _RowData(*[ws.cell(row_idx, col_idx).value for col_idx in range(1, 14)])
            row_indexes.set_description(f'Processing item {data.id or data.mpn}')
            if data.action == '-':
                skipped_count += 1
                continue
            row_errors = self._validate_row(data)
            if row_errors:
                errors[row_idx] = row_errors
                continue

            if data.action == 'create':
                item = get_item_by_mpn(self._endpoint, self._api_key, self._product_id, data.mpn)
                if item:
                    errors[row_idx] = [
                        f'Cannot create item: item with MPN `{data.mpn}`'
                        f' already exists with ID `{item["id"]}`.'
                    ]
                    continue
                row_indexes.set_description(f"Creating item {data[1]}")
                try:
                    item = create_item(
                        self._endpoint,
                        self._api_key,
                        self._product_id,
                        self._get_item_payload(data),
                    )
                    created_items.append(item)
                    self._update_sheet_row(ws, row_idx, item)
                except Exception as e:
                    errors[row_idx] = [str(e)]

            if data.action == 'update':
                item = self._get_item(data)

                if not item:
                    field = 'ID' if data.id else 'MPN'
                    value = data.id if data.id else data.mpn
                    errors[row_idx] = [
                        f'Cannot update item: item with {field} `{value}` '
                        f'the item does not exist.'
                    ]
                    continue

                row_indexes.set_description(f"Updating item {item['id']}")
                if item['status'] == 'published':
                    payload = {
                        'name': data.name,
                        'mpn': data.mpn,
                        'description': data.description,
                        'ui': {'visibility': True},
                    }
                else:
                    payload = self._get_item_payload(data)
                    if item['type'] == 'ppu':
                        del payload['period']
                try:
                    item = update_item(
                        self._endpoint,
                        self._api_key,
                        self._product_id,
                        item['id'],
                        payload,
                    )
                    updated_items.append(item)
                    self._update_sheet_row(ws, row_idx, item)
                except Exception as e:
                    errors[row_idx] = [str(e)]

            if data.action == 'delete':
                item = self._get_item(data)

                if not item:
                    field = 'ID' if data.id else 'MPN'
                    value = data.id if data.id else data.mpn
                    errors[row_idx] = [
                        f'Cannot update item: item with {field} `{value}` '
                        f'the item does not exist.'
                    ]
                    continue
                try:
                    delete_item(
                        self._endpoint,
                        self._api_key,
                        self._product_id,
                        item['id'],
                    )
                    deleted_items.append(item)
                except Exception as e:
                    errors[row_idx] = [str(e)]
        return (
            skipped_count,
            len(created_items),
            len(updated_items),
            len(deleted_items),
            errors,
        )

    def save(self, output_file):
        self._wb.save(output_file)

    def _open_workbook(self, input_file):
        try:
            self._wb = load_workbook(input_file)
        except InvalidFileException as ife:
            raise ClickException(str(ife))
        except BadZipFile:
            raise ClickException(f'{input_file} is not a valid xlsx file.')

    def _validate_item_sheet(self, ws):
        cels = ws['A1': 'M1']
        for cel in cels[0]:
            if cel.value != ITEMS_COLS_HEADERS[cel.column_letter]:
                raise ClickException(
                    f'Invalid input file: column {cel.column_letter} '
                    f'must be {ITEMS_COLS_HEADERS[cel.column_letter]}'
                )

    def _validate_commitment(self, row):
        if row.commitment not in COMMITMENT:
            valid_commitment = ', '.join([f'`{name}`' for name in COMMITMENT])
            return [
                f'the item `Commitment` must be one between '
                f'{valid_commitment}, not `{row.commitment}`.'
            ]

        if row.type == 'ppu':
            if row.commitment != '-':
                return [
                    f'the commitment `{row.commitment}` '
                    'is invalid for `ppu` items.'
                ]

        if row.billing_period == 'onetime' and row.commitment != '-':
            return [
                f'the commitment `{row.commitment}` '
                'is invalid for `onetime` items.'
            ]

        if row.billing_period == '2 years':
            if row.commitment not in ('-', '4 years'):
                return [
                    f'for a billing period of `2 years` the commitment '
                    'must be one between `-`, `4 years`, '
                    f' not {row.commitment}.'
                ]
        if row.billing_period in (
            '3 years',
            '4 years',
            '5 years',
        ):
            if row.commitment != '-':
                return [
                    f'for a billing period of `{row.billing}` the commitment '
                    f'must be `-`, not {row.commitment}.'
                ]

        return []

    def _validate_row(self, row):
        errors = []
        if row.action == ('delete', 'update'):
            if not (row.id or row.mpn):
                errors.append(
                    'one between the item `ID` or '
                    f'`MPN` is required for the `{row.action}` action.'
                )
                return errors
            if row.status == 'published':
                errors.append(
                    'the item status must be `draft` '
                    'for the `delete` action.'
                )
                return errors

        if row.action == 'create':
            if row.id:
                errors.append(
                    'the `ID` must not be specified '
                    'for the `create` action.'
                )
                return errors

        if not row.mpn:
            errors.append(
                'the item `MPN` is required.'
            )

        if not row.name:
            errors.append(
                f'the item `Name` is required for the `{row.action}` action.'
            )
        if not row.description:
            errors.append(
                f'the item `Description` is required for the `{row.action}` action.'
            )

        if row.type not in ('reservation', 'ppu'):
            errors.append(
                f'the item `Type` must be one between'
                f' `reservation` or `ppu`, not `{row.type}`.'
            )

        if row.type == 'reservation':
            if row.precision != 'integer':
                errors.append(
                    f'for items of type `reservation` the `Precision` '
                    f'must be `integer`, not `{row.precision}`.'
                )
        else:
            if row.precision not in PRECISIONS:
                valid_precision = ', '.join([f'`{name}`' for name in PRECISIONS])
                errors.append(
                    f'the item `Precision` must be one between '
                    f'{valid_precision}, not `{row.precision}`.'
                )

        if row.type == 'ppu':
            if row.billing_period != 'monthly':
                errors.append(
                    f'for items of type `ppu` the `Billing period` '
                    f'must be `monthly`, not `{row.billing_period}`.'
                )
        else:
            if row.billing_period not in BILLING_PERIOD:
                valid_period = ', '.join([f'`{name}`' for name in BILLING_PERIOD])
                errors.append(
                    f'the item `Billing period` must be one between'
                    f' {valid_period}, not `{row.billing_period}`.'
                )

        errors.extend(self._validate_commitment(row))
        return errors

    def _get_commitment_count(self, data):
        if data.billing_period == 'onetime':
            return 1
        if data.commitment == '-':
            return 1
        try:
            years, _ = data.commitment.split()
            years = int(years)
            if data.billing_period == 'monthly':
                return years * 12
            return years
        except:  # noqa
            return 1

    def _get_billing_period(self, data):
        if data.billing_period in ('onetime', 'monthly', 'yearly'):
            return data.billing_period
        count, _ = data.billing_period.split()
        return f'years_{count}'

    def _get_or_create_unit(self, data):
        for unit in self._units:
            if unit['id'] == data.unit:
                return unit['id']
            if unit['type'] == data.type and unit['description'] == data.unit:
                return unit['id']

        created = create_unit(
            self._endpoint,
            self._api_key,
            {
                'description': data.unit,
                'type': data.type,
                'unit': 'unit' if data.type == 'reservation' else 'unit-h'
            },
        )
        return created['id']

    def _get_item(self, data):
        if data.id:
            return get_item(self._endpoint, self._api_key, self._product_id, data.id)
        elif data.mpn:
            return get_item_by_mpn(self._endpoint, self._api_key, self._product_id, data.mpn)

    def _get_item_payload(self, data):
        commitment = {
            'commitment': {
                'count': self._get_commitment_count(data),
            },
        }
        payload = {
            'mpn': data.mpn,
            'name': data.name,
            'description': data.description,
            'type': data.type,
            'precision': data.precision,
            'unit': {'id': self._get_or_create_unit(data)},
            'period': self._get_billing_period(data),
            'ui': {'visibility': True},
        }
        if data.type == 'reservation':
            payload.update(commitment)

        return payload

    def _update_sheet_row(self, ws, row_idx, item):
        ws.cell(row_idx, 1, value=item['id'])
        ws.cell(row_idx, 11, value=item['status'])
        ws.cell(row_idx, 12, value=item['events']['created']['at'])
        ws.cell(row_idx, 13, value=item['events'].get('updated', {}).get('at'))

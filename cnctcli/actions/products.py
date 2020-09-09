# -*- coding: utf-8 -*-

# This file is part of the Ingram Micro Cloud Blue Connect product-sync.
# Copyright (c) 2019-2020 Ingram Micro. All Rights Reserved.

from zipfile import BadZipFile

import click
from connect.config import Config
from connect.exceptions import ServerError
from connect.resources.product import ProductsResource
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from tqdm import tqdm, trange

_COLS_HEADERS = {
    'A': 'Name',
    'B': 'MPN',
    'C': 'Billing Period',
    'D': 'Reservation',
    'E': 'Description',
    'F': 'Yearly Commitment',
    'G': 'Unit',
    'H': 'Connect Item ID',
    'I': 'Error Code',
    'J': 'Error Message',
}


def _setup_excel_sheet_header(ws):
    ws.sheet_properties.tabColor = '67389A'
    color = Color('d3d3d3')
    fill = PatternFill('solid', color)
    cels = ws['A1': 'H1']
    for cel in cels[0]:
        ws.column_dimensions[cel.column_letter].width = 25
        ws.column_dimensions[cel.column_letter].auto_size = True
        cel.fill = fill
        cel.value = _COLS_HEADERS[cel.column_letter]


def _check_skipped(skipped):
    if skipped:
        click.echo(
            click.style('The following products have been skipped:', fg='yellow')
        )
        for pinfo in skipped:
            click.echo(f'\t{pinfo[0]}: {pinfo[1]}')


def dump_products(api_url, api_key, product_ids, output_file):
    skipped = []
    config = Config(api_url=api_url, api_key=api_key)
    products = ProductsResource(config)
    wb = Workbook()
    ids = tqdm(product_ids, position=0)
    need_save = False
    for product_id in ids:
        ids.set_description('Processing product {}'.format(product_id))
        try:
            items = products.items(product_id).search()
        except:  # noqa: E722
            skipped.append(
                (
                    product_id,
                    f'Product "{product_id}"" does not exist.',
                ),
            )
            continue
        need_save = True
        ws = wb.create_sheet(product_id)
        _setup_excel_sheet_header(ws)
        processing_items = tqdm(items, position=1, leave=None)
        for row_idx, item in enumerate(processing_items, start=2):
            processing_items.set_description('Processing item {}'.format(item.id))
            ws.cell(row_idx, 1, value=item.display_name)
            ws.cell(row_idx, 2, value=item.mpn)
            ws.cell(row_idx, 3, value=item.period)
            ws.cell(row_idx, 4, value=item.type == 'reservation')
            ws.cell(row_idx, 5, value=item.description)
            commitment = item.commitment.count == 12 if item.commitment else False
            ws.cell(row_idx, 6, value=commitment)
            ws.cell(row_idx, 7, value=item.unit.unit)
            ws.cell(row_idx, 8, value=item.id)
            for i in range(1, 9):
                ws.column_dimensions[get_column_letter(i)].auto_size = True

    if need_save:
        wb.remove_sheet(wb.worksheets[0])
        wb.save(output_file)

    _check_skipped(skipped)


def _validate_sheet(ws):
    cels = ws['A1': 'H1']
    for cel in cels[0]:
        if cel.value != _COLS_HEADERS[cel.column_letter]:
            return _COLS_HEADERS[cel.column_letter]


def _report_exception(ws, row_idx, exc):
    color = Color('d3d3d3')
    fill = PatternFill('solid', color)
    cels = ws['I1': 'J1']
    for cel in cels[0]:
        ws.column_dimensions[cel.column_letter].width = 25
        ws.column_dimensions[cel.column_letter].auto_size = True
        cel.fill = fill
        cel.value = _COLS_HEADERS[cel.column_letter]
    code = '-'
    msg = str(exc)
    if isinstance(exc, ServerError):
        code = exc.error.error_code
        msg = '\n'.join(exc.error.errors)
    ws.cell(row_idx, 9, value=code)
    ws.cell(row_idx, 10, value=msg)
    return code, msg


def _get_product_item_by_id(items, item_id):
    try:
        return items.get(item_id)
    except:  # noqa: E722
        pass


def _get_product_item_by_mpn(items, mpn):
    try:
        results = items.search(filters={'mpn': mpn})
        return results[0] if results else None
    except:  # noqa: E722
        pass


def _create_product_item(items, data):
    if data[3] is True:
        # reservation
        period = 'monthly'
        if data[2].lower() == 'yearly':
            period = 'yearly'
        if data[2].lower() == 'onetime':
            period = 'onetime'
        item = {
            'name': data[0],
            'mpn': data[1],
            'description': data[4],
            'ui': {'visibility': True},
            'commitment': {
                'count': 12 if data[5] is True else 1,
            },
            'unit': {'id': data[6]},
            'type': 'reservation',
            'period': period,
        }
    else:
        # PPU
        item = {
            'name': data[0],
            'mpn': data[1],
            'description': data[4],
            'ui': {'visibility': True},
            'unit': {'id': data[6]},
            'type': 'ppu',
            'precision': 'decimal(2)',
        }

    return items.create(item)


def sync_products(api_url, api_key, input_file):
    skipped = []
    items_errors = []
    need_save = False
    config = Config(api_url=api_url, api_key=api_key)
    products = ProductsResource(config)
    wb = None
    try:
        wb = load_workbook(input_file)
    except InvalidFileException as ife:
        click.echo(
            click.style(str(ife), fg='red')
        )
        return
    except BadZipFile:
        click.echo(
            click.style(f'{input_file} is not a valid xlsx file.', fg='red')
        )
        return
    product_ids = wb.sheetnames
    ids = tqdm(product_ids, position=0)
    for product_id in ids:
        ids.set_description('Syncing product {}'.format(product_id))
        try:
            products.get(product_id)
        except:   # noqa: E722
            skipped.append(
                (
                    product_id,
                    f'The product "{product_id}" does not exist.',
                ),
            )
            continue
        items = products.items(product_id)
        ws = wb[product_id]
        invalid_column = _validate_sheet(ws)
        if invalid_column:
            skipped.append(
                (
                    product_id,
                    f'The worksheet "{product_id}" does not have the column {invalid_column}.',
                ),
            )
            continue

        row_indexes = trange(2, ws.max_row + 1, position=1, leave=None)
        for row_idx in row_indexes:
            row_indexes.set_description('Processing row {}'.format(row_idx))
            data = [ws.cell(row_idx, col_idx).value for col_idx in range(1, 9)]
            if data[7]:
                item = _get_product_item_by_id(items, data[7])
            else:
                item = _get_product_item_by_mpn(items, data[1])

            if item:
                row_indexes.set_description('Updating item {}'.format(item.mpn))
                try:
                    items.update(
                        item.id,
                        {
                            'name': data[0],
                            'mpn': data[1],
                            'description': data[4],
                            'ui': {'visibility': True},
                        },
                    )
                except ServerError as se:
                    code, msg = _report_exception(ws, row_idx, se)
                    items_errors.append(
                        f"\t{product_id}: mpn={data[1]}, code={code}, message={msg}"
                    )
                    need_save = True
            else:
                try:
                    result = _create_product_item(items, data)
                    ws.cell(row_idx, 8, value=result.id)
                except ServerError as se:
                    code, msg = _report_exception(ws, row_idx, se)
                    items_errors.append(
                        f"\t{product_id}: mpn={data[1]}, code={code}, message={msg}"
                    )
                need_save = True
    if need_save:
        wb.save(input_file)

    _check_skipped(skipped)
    if items_errors:
        click.echo(
            click.style('\n\nThe following items have not been synced:', fg='yellow')
        )

        for i in items_errors:
            click.echo(i)
